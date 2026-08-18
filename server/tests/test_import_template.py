"""Tệp mẫu Excel: sinh từ registry, và kiểm cấu trúc từ chối đúng thứ phải từ chối.

Không cần PostgreSQL — descriptor, bộ sinh tệp và bộ đọc đều là hàm thuần trên
`CatalogSpec`. Phần chạm DB (kiểm dữ liệu, ghi) nằm ở `test_import_pipeline.py`.
"""

from __future__ import annotations

from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

import pytest
from openpyxl import Workbook, load_workbook

from ket.kernel.errors import (
    ImportFileUnreadableError,
    ImportSheetMissingError,
    ImportTemplateMismatchError,
    ImportTooManyRowsError,
)
from ket.kernel.excel.descriptors import (
    INSTRUCTIONS_SHEET,
    CellKind,
    template_for,
)
from ket.kernel.excel.reader import (
    MAX_ROWS,
    MAX_SCANNED_ROWS,
    ensure_structure,
    iter_rows,
)
from ket.kernel.excel.template import build_template
from ket.kernel.master_data.registry import REGISTRY


def _spec(slug: str):  # type: ignore[no-untyped-def] # noqa: ANN202 — trả về CatalogSpec đã khẳng định khác None
    spec = REGISTRY.get(slug)
    assert spec is not None, f"danh mục {slug} không còn đăng ký"
    return spec


def _template_bytes(slug: str) -> bytes:
    spec = _spec(slug)
    return build_template(spec, template_for(spec))


def _filled(slug: str, rows: list[list[object]]) -> BytesIO:
    """Một tệp đúng cấu trúc mẫu, kèm các dòng dữ liệu đã cho."""
    spec = _spec(slug)
    descriptor = template_for(spec)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = descriptor.sheet_name
    sheet.append(list(descriptor.headers))
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


# --------------------------------------------------------------- descriptor


def test_every_registered_catalog_has_a_template() -> None:
    """H77: thêm danh mục = thêm model + descriptor, tệp mẫu có sẵn.

    Canh trên **toàn bộ** registry chứ không một vài danh mục mẫu: đây chính là
    tính chất mà H77 mua, và nó chỉ có giá trị nếu không có ngoại lệ nào.
    """
    for spec in REGISTRY.specs():
        descriptor = template_for(spec)
        assert descriptor.columns, f"{spec.slug} không có cột nào"
        assert descriptor.sheet_name, f"{spec.slug} không có tên sheet"
        # Tên sheet phải vừa trần của định dạng xlsx, nếu không Excel tự cắt và
        # tệp mẫu sinh ra mang một tên mà chính phép kiểm của ta từ chối.
        assert len(descriptor.sheet_name) <= 31, spec.slug
        assert len(set(descriptor.headers)) == len(descriptor.headers), (
            f"{spec.slug} có hai cột trùng tiêu đề — dòng tiêu đề là hợp đồng, "
            "trùng nghĩa là một cột không bao giờ đọc được"
        )


def test_required_columns_are_marked_with_a_star() -> None:
    """FR-SYS-083: cột bắt buộc đánh dấu `*`, và dấu nằm trong chính tiêu đề."""
    descriptor = template_for(_spec("items"))
    code = descriptor.column("code")
    name_en = descriptor.column("name_en")
    assert code is not None and name_en is not None
    assert code.required and code.display_header.endswith(" *")
    assert not name_en.required and not name_en.display_header.endswith("*")


def test_foreign_keys_become_code_columns_not_id_columns() -> None:
    """H79: người điền tệp gõ mã, không gõ khóa ngoại.

    Vật tư hàng hóa là ca đáng canh nhất — nó có hai `CatalogReference`
    (`base_unit_id`, `warehouse_id`), và cả hai phải xuất hiện dưới dạng mã.
    """
    descriptor = template_for(_spec("items"))
    fields = {column.field for column in descriptor.columns}
    assert "base_unit_code" in fields
    assert "warehouse_code" in fields
    assert not any(field.endswith("_id") for field in fields), (
        f"tệp mẫu không được có cột id: {sorted(f for f in fields if f.endswith('_id'))}"
    )
    for column in descriptor.references:
        assert column.kind is CellKind.CODE_REF
        assert column.reference_slug is not None
        assert column.target_field is not None


def test_branch_is_not_a_per_row_column() -> None:
    """Chi nhánh là lựa chọn của cả lượt nhập, không của từng dòng.

    Một cột chi nhánh trên từng dòng là đường để một tệp ghi vào chi nhánh mà
    người nhập không đứng ở đó — phép kiểm phạm vi của tầng HTTP đọc
    `acting_branch_id`, không đọc tệp.
    """
    for spec in REGISTRY.specs():
        fields = {column.field for column in template_for(spec).columns}
        assert "branch_id" not in fields and "branch_code" not in fields, spec.slug


# ------------------------------------------------------------------- tệp mẫu


def test_template_has_an_instructions_sheet_and_a_data_sheet() -> None:
    """FR-SYS-080: bộ tệp mẫu có sheet **Hướng dẫn**."""
    spec = _spec("warehouses")
    workbook = load_workbook(BytesIO(_template_bytes("warehouses")))
    descriptor = template_for(spec)
    assert workbook.sheetnames == [INSTRUCTIONS_SHEET, descriptor.sheet_name]
    assert workbook[descriptor.sheet_name][1][0].value == descriptor.headers[0]


def test_template_data_sheet_has_no_example_row() -> None:
    """Một dòng ví dụ là một dòng người dùng quên xóa rồi nhập vào danh mục thật."""
    descriptor = template_for(_spec("warehouses"))
    sheet = load_workbook(BytesIO(_template_bytes("warehouses")))[descriptor.sheet_name]
    assert sheet.max_row == 1


def test_generated_template_passes_its_own_structure_check() -> None:
    """Vòng khép kín: tệp mẫu ta phát ra phải là tệp ta nhận vào.

    Không có phép kiểm này thì hai bên (`template.py` dựng tiêu đề,
    `reader.py` so tiêu đề) trôi lệch nhau và triệu chứng là người dùng tải mẫu
    về, không sửa gì, nộp lên và bị từ chối.
    """
    for spec in REGISTRY.specs():
        descriptor = template_for(spec)
        content = BytesIO(build_template(spec, descriptor))
        ensure_structure(content, descriptor)


# --------------------------------------------------------- kiểm cấu trúc


def test_renamed_sheet_is_refused_and_the_message_names_the_right_one() -> None:
    """FR-SYS-082 + bước 13: từ chối, và nói tên đúng."""
    descriptor = template_for(_spec("warehouses"))
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(list(descriptor.headers))
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    with pytest.raises(ImportSheetMissingError) as error:
        ensure_structure(buffer, descriptor)
    assert error.value.details["expected"] == descriptor.sheet_name
    assert error.value.problem_extra()["found_sheets"] == ["Sheet1"]


def test_renamed_column_is_refused_and_the_message_names_what_is_missing() -> None:
    """FR-SYS-082: đổi tên cột thì không nhận, và thông điệp nêu cột thiếu."""
    descriptor = template_for(_spec("warehouses"))
    headers = list(descriptor.headers)
    headers[1] = "Tên kho"
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = descriptor.sheet_name
    sheet.append(headers)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    with pytest.raises(ImportTemplateMismatchError) as error:
        ensure_structure(buffer, descriptor)
    extra = error.value.problem_extra()
    assert descriptor.headers[1] in extra["missing_columns"]
    assert "Tên kho" in extra["unexpected_columns"]


def test_reordered_columns_are_refused() -> None:
    """Đổi chỗ hai cột cho nhau: tập tên vẫn đủ, nhưng dữ liệu sẽ vào sai ô.

    Đây là ca mà một phép kiểm "so tập hợp tên" bỏ lọt — và hậu quả của nó im
    lặng hoàn toàn: mã đi vào cột tên, tên đi vào cột mã, không lỗi nào nổ.
    """
    descriptor = template_for(_spec("warehouses"))
    headers = list(descriptor.headers)
    headers[0], headers[1] = headers[1], headers[0]
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = descriptor.sheet_name
    sheet.append(headers)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    with pytest.raises(ImportTemplateMismatchError):
        ensure_structure(buffer, descriptor)


def test_trailing_whitespace_in_a_header_is_not_a_rename() -> None:
    """`"Mã "` không phải là đổi tên cột — Excel thêm khoảng trắng một cách vô hình."""
    descriptor = template_for(_spec("warehouses"))
    headers = [f"  {descriptor.headers[0]}  ", *descriptor.headers[1:]]
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = descriptor.sheet_name
    sheet.append(headers)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    ensure_structure(buffer, descriptor)


def test_a_file_that_is_not_a_workbook_is_refused_with_a_usable_message() -> None:
    with pytest.raises(ImportFileUnreadableError):
        ensure_structure(BytesIO(b"day khong phai xlsx"), template_for(_spec("warehouses")))


# ------------------------------------------------------------- đọc dòng


def test_blank_rows_are_skipped_but_partially_blank_rows_are_not() -> None:
    """Dòng rỗng hoàn toàn bỏ qua; dòng thiếu một ô thì vẫn phải đi tiếp để bị bắt lỗi."""
    descriptor = template_for(_spec("warehouses"))
    source = _filled(
        "warehouses",
        [
            ["KHO1", "Kho chính", None, None, None, None],
            [None, None, None, None, None, None],
            [None, "Kho thiếu mã", None, None, None, None],
        ],
    )
    rows = list(iter_rows(source, descriptor))
    assert [row for row, _ in rows] == [2, 4]
    assert rows[1][1]["code"] is None
    assert rows[1][1]["name"] == "Kho thiếu mã"


def test_row_numbers_match_what_the_user_sees_in_excel() -> None:
    """FR-SYS-081: "lỗi theo từng dòng" chỉ dùng được nếu số dòng khớp thanh dòng Excel."""
    descriptor = template_for(_spec("warehouses"))
    source = _filled("warehouses", [["K1", "Kho 1", None, None, None, None]])
    ((row_number, _),) = list(iter_rows(source, descriptor))
    assert row_number == 2, "dòng dữ liệu đầu tiên là dòng 2, ngay dưới dòng tiêu đề"


def test_a_numeric_cell_keeps_the_digits_the_user_typed() -> None:
    """Excel lưu số dưới dạng số thực; ép thẳng sang chuỗi làm hỏng chữ số.

    `0.1 + 0.2` là ví dụ kinh điển, nhưng ca thật ở đây tầm thường hơn nhiều:
    một tỷ lệ quy đổi `1.15` gõ vào ô sẽ đi vào bảng đệm, và nếu nó đến đó dưới
    dạng `1.1499999999999999` thì phép kiểm kiểu vẫn cho qua và cột `NUMERIC`
    nhận một con số khác thứ người dùng nhìn thấy.
    """
    descriptor = template_for(_spec("warehouses"))
    source = _filled("warehouses", [["K1", 0.1 + 0.2, None, None, None, None]])
    ((_, values),) = list(iter_rows(source, descriptor))
    assert values["name"] == "0.30000000000000004" or values["name"] == "0.3"
    assert "e" not in (values["name"] or ""), "không được rơi sang ký hiệu khoa học"


def test_an_integer_cell_does_not_grow_a_decimal_tail() -> None:
    descriptor = template_for(_spec("warehouses"))
    source = _filled("warehouses", [["K1", 5, None, None, None, None]])
    ((_, values),) = list(iter_rows(source, descriptor))
    assert values["name"] == "5"


def test_a_file_over_the_row_ceiling_is_refused() -> None:
    """Trần dòng bảo vệ bộ nhớ máy chủ khỏi một tệp dán nhầm."""
    descriptor = template_for(_spec("warehouses"))
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(descriptor.sheet_name)
    sheet.append(list(descriptor.headers))
    for index in range(MAX_ROWS + 5):
        sheet.append([f"K{index}", f"Kho {index}", None, None, None, None])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    with pytest.raises(ImportTooManyRowsError):
        list(iter_rows(buffer, descriptor))


def test_a_file_padded_with_blank_rows_is_refused_before_it_burns_a_worker() -> None:
    """H4: `MAX_ROWS` một mình không chặn được gì — nó chỉ đếm dòng **có nội dung**.

    Một tệp vài trăm kilobyte khai vùng đã dùng tới hàng triệu dòng rỗng giữ vòng
    đọc chạy hàng chục giây mà `emitted` không bao giờ chạm trần. Dòng rỗng không
    tốn bộ nhớ, nhưng nó tốn đúng thứ FR-NFR-044 bảo vệ: một worker — và hàng đợi
    chỉ có vài worker cho cả phòng kế toán.

    Trần thứ hai đếm dòng **quét qua**, kể cả rỗng.
    """
    descriptor = template_for(_spec("warehouses"))
    workbook = Workbook(write_only=True)
    sheet = workbook.create_sheet(descriptor.sheet_name)
    sheet.append(list(descriptor.headers))
    for _ in range(MAX_SCANNED_ROWS + 10):
        sheet.append([None] * len(descriptor.columns))
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    with pytest.raises(ImportTooManyRowsError):
        list(iter_rows(buffer, descriptor))


def _workbook_without_dimension(row_count: int) -> BytesIO:
    """Tệp xlsx **không** khai thẻ `<dimension>`, với `row_count` dòng rỗng.

    Dựng bằng tay vì openpyxl **luôn** ghi thẻ đó — nên một test dùng
    `Workbook(write_only=True)` không bao giờ chạm đường tấn công thật, và bản sửa
    đầu tiên cho H4 đã xanh mà vẫn để nguyên chi phí 9 giây trong `load_workbook`.
    """
    sheet_xml = (
        '<?xml version="1.0"?><worksheet '
        'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"><sheetData>'
        + "<row/>" * row_count
        + "</sheetData></worksheet>"
    )
    buffer = BytesIO()
    with ZipFile(buffer, "w", ZIP_DEFLATED) as archive:
        archive.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0"?><Types '
            'xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.'
            'relationships+xml"/><Default Extension="xml" ContentType="application/xml"/>'
            "</Types>",
        )
        archive.writestr(
            "_rels/.rels",
            '<?xml version="1.0"?><Relationships '
            'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
            '2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>',
        )
        archive.writestr(
            "xl/workbook.xml",
            '<?xml version="1.0"?><workbook '
            'xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
            '<sheets><sheet name="Kho" sheetId="1" r:id="rId1"/></sheets></workbook>',
        )
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            '<?xml version="1.0"?><Relationships '
            'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/'
            '2006/relationships/worksheet" Target="worksheets/sheet1.xml"/></Relationships>',
        )
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    buffer.seek(0)
    return buffer


def test_a_file_that_explodes_when_decompressed_is_refused_before_it_is_opened() -> None:
    """R2-4: chi phí thật nằm **trước** vòng lặp đọc, không trong nó.

    Khi bảng tính không khai `<dimension>`, openpyxl phải quét trọn XML của sheet
    ngay trong `load_workbook` để biết vùng dữ liệu — một tệp 172 KB chứa hàng
    chục triệu thẻ `<row/>` rỗng ngốn ~9 giây CPU **trước khi** dòng đầu tiên được
    phát ra. `MAX_SCANNED_ROWS` không chạm tới được, và `IMPORT_MAX_BYTES` thì đo
    tệp đã nén.

    Phép đo là kích thước **sau giải nén**, đọc từ mục lục zip mà không giải nén
    byte nào.
    """
    descriptor = template_for(_spec("warehouses"))
    bomb = _workbook_without_dimension(20_000_000)
    assert len(bomb.getvalue()) < 5 * 1024 * 1024, "tệp nén phải nhỏ, nếu không test đo nhầm thứ"

    with pytest.raises(ImportTooManyRowsError):
        ensure_structure(bomb, descriptor)


def test_an_enum_column_publishes_its_allowed_values() -> None:
    """R2-1: tập giá trị đọc từ **kiểu cột thật**, không khai tay.

    Nhờ vậy thêm một tính chất ở phase sau không phải nhớ cập nhật khung nhập liệu.
    """
    nature = template_for(_spec("items")).column("nature")
    assert nature is not None
    assert set(nature.allowed_values) == {"goods", "finished_goods", "service", "description_only"}


def test_a_column_learns_its_length_limit_from_the_real_column() -> None:
    """R2-3: `swift_code` khai `String(11)` ở ORM nhưng không khai `max_length` ở API.

    Không đọc trần từ cột thì bước kiểm không có gì để so, và giá trị dài quá chỉ
    đổ ở tầng DB — sau khi người dùng đã được báo "hợp lệ".
    """
    swift = template_for(_spec("banks")).column("swift_code")
    assert swift is not None and swift.max_length == 11
