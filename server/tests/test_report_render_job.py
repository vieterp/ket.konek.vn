"""Job render nền + ngưỡng chuyển-job trên PostgreSQL thật (lát 5E, bước 19).

Bất biến bám tiêu chí phase-05:

* **FR-NFR-042/044**: render vượt ngưỡng KHÔNG chạy trong request — endpoint
  trả `202` + job; UI theo dõi tiến độ và hủy được qua hàng đợi phase 2.
* **Ngưỡng là cấu hình** (`report.{pdf,xlsx}_job_threshold_rows`): mỗi bản cài
  một tốc độ máy chủ, đổi ngưỡng không sửa code.
* **Tệp kết quả chỉ người yêu cầu tải được**: tệp mang số liệu theo phạm vi
  RLS của người bấm render — người khác thấy dòng job (RLS chi nhánh) không
  có nghĩa đọc được tệp.
* Job render là phép ĐỌC (`idempotent-restart`), hủy dừng ở ranh giới lô.

Dataset riêng (`rpt5e`) — cùng lý do mọi tệp reporting tách dataset: bộ sổ
khép kín, không cộng chéo với fixture của tệp khác.
"""

from __future__ import annotations

import io
from collections.abc import Iterator
from datetime import date
from decimal import Decimal
from pathlib import Path
from uuid import UUID, uuid4

import openpyxl
import pytest
from fastapi.testclient import TestClient
from sqlalchemy import Engine
from sqlalchemy.orm import Session, sessionmaker

import ket.reporting.render_job as render_job_module
from catalog_api_support import UserFactory, actor, all_branch_codes, ensure_role
from conftest import api_test_client
from import_support import FakeProgress
from ket.api.dependencies import BRANCH_HEADER
from ket.kernel.config.catalog import (
    REPORT_XLSX_JOB_THRESHOLD_KEY,
    SettingScope,
)
from ket.kernel.config.settings_service import set_setting
from ket.kernel.datasets.provisioning import DatasetRef, drop_dataset_schema, provision_dataset
from ket.kernel.errors import AttachmentStorageNotConfiguredError
from ket.kernel.jobs import queue
from ket.kernel.jobs.models import JobStatus
from ket.kernel.jobs.registry import REGISTRY, JobCancelled, JobContext
from ket.kernel.persistence.unit_of_work import unit_of_work
from ket.main import create_app
from ket.modules.general_ledger.journal.schemas import JournalLineIn, JournalVoucherIn
from ket.modules.general_ledger.journal.service import JournalVoucherService
from ket.reporting.engine import REPORT_EXPORT, REPORT_VIEW
from ket.reporting.render_job import RENDER_JOB_CODE, ReportRenderJobParams, _run_render
from ket.settings import Settings
from ket.worker.runner import resolve_body_scope
from posting_support import PostingContext, posting_scope, seed_posting_context

pytestmark = pytest.mark.db

ACTOR_ID = 1
VND = "VND"
_DATASET_CODE = "rpt5e"
RANGE = {"from_date": "2026-01-01", "to_date": "2026-12-31"}


@pytest.fixture(scope="module")
def render_storage(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return tmp_path_factory.mktemp("render-blobs")


@pytest.fixture(scope="module")
def render_settings(test_settings: Settings, render_storage: Path) -> Settings:
    """App + worker cùng nhìn MỘT kho tệp — điều kiện để download thấy blob."""
    return test_settings.model_copy(update={"attachments_dir": render_storage})


@pytest.fixture
def client(
    render_settings: Settings, app_engine: Engine, session_factory: sessionmaker[Session]
) -> Iterator[TestClient]:
    with api_test_client(create_app(render_settings)) as instance:
        yield instance


@pytest.fixture(scope="module")
def render_dataset(owner_engine: Engine) -> Iterator[DatasetRef]:
    ref = provision_dataset(owner_engine, code=_DATASET_CODE, name="Render job 5E", scheme="TT99")
    yield ref
    drop_dataset_schema(owner_engine, _DATASET_CODE)


@pytest.fixture(scope="module")
def context(session_factory: sessionmaker[Session], render_dataset: DatasetRef) -> PostingContext:
    return seed_posting_context(session_factory, render_dataset)


@pytest.fixture(scope="module")
def exporter_role(session_factory: sessionmaker[Session], render_dataset: DatasetRef) -> str:
    return ensure_role(
        session_factory, render_dataset, "xuat_bao_cao_5e", [REPORT_VIEW, REPORT_EXPORT]
    )


_seeded: set[int] = set()


def _seed_one_voucher(
    session_factory: sessionmaker[Session],
    render_dataset: DatasetRef,
    context: PostingContext,
) -> None:
    """Một chứng từ đã ghi sổ (2 dòng) trong chi nhánh của `context`."""
    if context.branch_id in _seeded:
        return
    _seeded.add(context.branch_id)
    with unit_of_work(
        session_factory, posting_scope(render_dataset, context, user_id=ACTOR_ID)
    ) as session:
        service = JournalVoucherService(session)
        voucher = service.create(
            JournalVoucherIn(
                branch_id=context.branch_id,
                document_date=date(2026, 1, 10),
                posting_date=date(2026, 1, 10),
                currency_code=VND,
                exchange_rate=Decimal(1),
                description="chứng từ test render job",
                lines=(
                    JournalLineIn(
                        account_id=context.accounts["111"],
                        debit_fc=Decimal(1_000_000),
                        credit_fc=Decimal(0),
                    ),
                    JournalLineIn(
                        account_id=context.accounts["511"],
                        debit_fc=Decimal(0),
                        credit_fc=Decimal(1_000_000),
                    ),
                ),
            ),
            user_id=ACTOR_ID,
        )
        service.post(voucher.id, user_id=ACTOR_ID)


@pytest.fixture(scope="module")
def seeded_books(
    session_factory: sessionmaker[Session],
    render_dataset: DatasetRef,
    context: PostingContext,
) -> None:
    _seed_one_voucher(session_factory, render_dataset, context)


@pytest.fixture(scope="module")
def context_b(session_factory: sessionmaker[Session], render_dataset: DatasetRef) -> PostingContext:
    """Chi nhánh thứ hai — cho ca đa chi nhánh của C1."""
    return seed_posting_context(session_factory, render_dataset)


@pytest.fixture(scope="module")
def seeded_books_b(
    session_factory: sessionmaker[Session],
    render_dataset: DatasetRef,
    context_b: PostingContext,
) -> None:
    _seed_one_voucher(session_factory, render_dataset, context_b)


@pytest.fixture(scope="module")
def low_xlsx_threshold(
    session_factory: sessionmaker[Session],
    render_dataset: DatasetRef,
    context: PostingContext,
) -> None:
    """Hạ ngưỡng XLSX xuống 1 dòng — mọi lượt render XLSX của tệp này đi job.

    Ngưỡng PDF giữ mặc định (2.000) nên nhánh đồng bộ vẫn có test phủ ở
    `test_reports_api.py`; tệp này chỉ cần chứng minh CÔNG TẮC.
    """
    with unit_of_work(
        session_factory, posting_scope(render_dataset, context, user_id=ACTOR_ID)
    ) as session:
        set_setting(
            session,
            key=REPORT_XLSX_JOB_THRESHOLD_KEY,
            scope=SettingScope.SYSTEM,
            user_id=ACTOR_ID,
            raw_value="1",
            expected_row_version=None,
        )


def _headers(
    client: TestClient,
    session_factory: sessionmaker[Session],
    dataset: DatasetRef,
    user_factory: UserFactory,
    role: str,
    prefix: str,
    password: str,
    branch_id: int | None,
) -> dict[str, str]:
    headers = actor(
        client,
        session_factory,
        dataset,
        user_factory,
        role,
        prefix,
        password,
        branch_codes=all_branch_codes(session_factory, dataset),
    )
    # `branch_id=None` = KHÔNG gửi X-Branch — đường hợp lệ của user đa chi
    # nhánh (`_acting_branch` trả None), chính ca đã phơi C1.
    if branch_id is None:
        return headers
    return {**headers, BRANCH_HEADER: str(branch_id)}


def _enqueue_render(client: TestClient, headers: dict[str, str]) -> tuple[UUID, int]:
    response = client.post(
        "/api/v1/reports/S03b-DN/render",
        json={"format": "xlsx", "params": RANGE},
        headers=headers,
    )
    assert response.status_code == 202, response.text
    body = response.json()
    return UUID(body["job_id"]), body["estimated_rows"]


def _run_render_job(
    session_factory: sessionmaker[Session],
    dataset: DatasetRef,
    context: PostingContext,
    storage_root: Path,
    job_id: UUID,
) -> None:
    """Vòng worker thu nhỏ: giành job → chạy thân → ghi kết quả.

    Ba transaction tách bạch đúng như `ket.worker.runner`. Phạm vi của THÂN job
    dựng bằng CHÍNH `resolve_body_scope` của runner từ dòng job (H1 review 5E):
    harness tự chế scope là đúng khoảng cách đã cho C1 trốn toàn bộ test.
    """
    claim_scope = posting_scope(dataset, context, user_id=ACTOR_ID)
    with unit_of_work(session_factory, claim_scope) as session:
        job = queue.claim_next(session)
        assert job is not None and job.id == job_id, "hàng đợi phải chỉ có job vừa xếp"
        attempt = job.attempt
        params = dict(job.params or {})
        requested_by = job.requested_by
        job_branch_id = job.branch_id
        job_type = REGISTRY.get(job.type)
    body_scope = resolve_body_scope(
        session_factory,
        dataset,
        job_type=job_type,
        job_id=job_id,
        requested_by=requested_by,
        branch_id=job_branch_id,
    )
    with unit_of_work(session_factory, body_scope) as session:
        result = job_type.run(
            JobContext(
                job_id=job_id,
                session=session,
                progress=FakeProgress(reports=[]),
                attempt=attempt,
                dataset_schema=dataset.schema_name,
                branch_id=body_scope.acting_branch_id,
                requested_by=requested_by,
                storage_root=storage_root,
            ),
            params,
        )
    with unit_of_work(session_factory, claim_scope) as session:
        assert queue.finish(session, job_id, attempt=attempt, result=result)


def test_render_job_is_registered_but_not_directly_enqueueable() -> None:
    job_type = REGISTRY.get(RENDER_JOB_CODE)
    assert job_type.resume_semantics.value == "idempotent-restart"
    assert job_type.direct_enqueue is False


def test_render_job_without_storage_root_refuses_with_actionable_error() -> None:
    fake_dataset = DatasetRef(id=0, code="x", schema_name="ds_x", scheme="TT99")
    context = JobContext(
        job_id=uuid4(),
        # Lỗi phải nổ TRƯỚC mọi lượt chạm session — session giả là chủ đích.
        session=None,  # type: ignore[arg-type]
        progress=FakeProgress(reports=[]),
        attempt=1,
        dataset_schema=fake_dataset.schema_name,
        branch_id=1,
        requested_by=ACTOR_ID,
        storage_root=None,
    )
    with pytest.raises(AttachmentStorageNotConfiguredError):
        _run_render(
            context,
            ReportRenderJobParams(code="S03b-DN", format="xlsx", params={}),
        )


class TestRenderJobFlow:
    def test_over_threshold_returns_202_then_file_downloads(
        self,
        client: TestClient,
        session_factory: sessionmaker[Session],
        render_dataset: DatasetRef,
        context: PostingContext,
        user_factory: UserFactory,
        exporter_role: str,
        test_password: str,
        render_storage: Path,
        seeded_books: None,
        low_xlsx_threshold: None,
    ) -> None:
        headers = _headers(
            client,
            session_factory,
            render_dataset,
            user_factory,
            exporter_role,
            "rj-chinh",
            test_password,
            context.branch_id,
        )
        job_id, estimated = _enqueue_render(client, headers)
        assert estimated >= 2, "chứng từ 2 dòng phải cho ≥2 dòng dataset"

        # Chưa chạy → 409 kèm trạng thái để client biết "chờ thêm".
        pending = client.get(f"/api/v1/reports/render-jobs/{job_id}/file", headers=headers)
        assert pending.status_code == 409, pending.text
        assert pending.json()["error_code"] == "report.render_not_ready"

        _run_render_job(session_factory, render_dataset, context, render_storage, job_id)

        done = client.get(f"/api/v1/reports/render-jobs/{job_id}/file", headers=headers)
        assert done.status_code == 200, done.text
        assert "S03b-DN" in done.headers["content-disposition"]
        workbook = openpyxl.load_workbook(io.BytesIO(done.content), data_only=True)
        sheet = workbook.active
        assert sheet is not None
        cells = [row[0] for row in sheet.iter_rows(values_only=True)]
        assert "Tổng cộng" in cells, "tệp từ job phải là cùng bản kết xuất với đường đồng bộ"

    def test_file_of_someone_elses_job_is_not_found(
        self,
        client: TestClient,
        session_factory: sessionmaker[Session],
        render_dataset: DatasetRef,
        context: PostingContext,
        user_factory: UserFactory,
        exporter_role: str,
        test_password: str,
        render_storage: Path,
        seeded_books: None,
        low_xlsx_threshold: None,
    ) -> None:
        owner_headers = _headers(
            client,
            session_factory,
            render_dataset,
            user_factory,
            exporter_role,
            "rj-chu",
            test_password,
            context.branch_id,
        )
        job_id, _ = _enqueue_render(client, owner_headers)
        _run_render_job(session_factory, render_dataset, context, render_storage, job_id)

        other_headers = _headers(
            client,
            session_factory,
            render_dataset,
            user_factory,
            exporter_role,
            "rj-khac",
            test_password,
            context.branch_id,
        )
        # Người khác — kể cả cùng chi nhánh, cùng quyền export — nhận 404:
        # tệp mang số liệu theo phạm vi RLS của NGƯỜI YÊU CẦU.
        stolen = client.get(f"/api/v1/reports/render-jobs/{job_id}/file", headers=other_headers)
        assert stolen.status_code == 404, stolen.text
        # Chủ job vẫn tải được — 404 ở trên là vì danh tính, không phải vì tệp.
        owned = client.get(f"/api/v1/reports/render-jobs/{job_id}/file", headers=owner_headers)
        assert owned.status_code == 200

    def test_cancel_stops_render_at_batch_boundary(
        self,
        client: TestClient,
        session_factory: sessionmaker[Session],
        render_dataset: DatasetRef,
        context: PostingContext,
        user_factory: UserFactory,
        exporter_role: str,
        test_password: str,
        render_storage: Path,
        seeded_books: None,
        low_xlsx_threshold: None,
        monkeypatch: pytest.MonkeyPatch,
    ) -> None:
        headers = _headers(
            client,
            session_factory,
            render_dataset,
            user_factory,
            exporter_role,
            "rj-huy",
            test_password,
            context.branch_id,
        )
        job_id, _ = _enqueue_render(client, headers)
        # Nhịp kiểm thật là 1.000 dòng — fixture chỉ có vài dòng nên hạ nhịp
        # xuống 1 để cờ hủy được đọc ngay ở dòng đầu; thứ đang đo là ĐƯỜNG
        # DỪNG (JobCancelled thoát khỏi pipeline streaming), không phải nhịp.
        monkeypatch.setattr(render_job_module, "CANCEL_CHECK_EVERY_ROWS", 1)

        scope = posting_scope(render_dataset, context, user_id=ACTOR_ID)
        with unit_of_work(session_factory, scope) as session:
            job = queue.claim_next(session)
            assert job is not None and job.id == job_id
            attempt = job.attempt
            params = dict(job.params or {})
        with (
            unit_of_work(session_factory, scope) as session,
            pytest.raises(JobCancelled),
        ):
            REGISTRY.get(RENDER_JOB_CODE).run(
                JobContext(
                    job_id=job_id,
                    session=session,
                    progress=FakeProgress(reports=[], cancelled=True),
                    attempt=attempt,
                    dataset_schema=render_dataset.schema_name,
                    branch_id=context.branch_id,
                    requested_by=ACTOR_ID,
                    storage_root=render_storage,
                ),
                params,
            )
        with unit_of_work(session_factory, scope) as session:
            assert queue.mark_cancelled(session, job_id, attempt=attempt)
        with unit_of_work(session_factory, scope) as session:
            assert queue.get_job(session, job_id).status == JobStatus.CANCELLED.value

        cancelled = client.get(f"/api/v1/reports/render-jobs/{job_id}/file", headers=headers)
        assert cancelled.status_code == 409
        assert cancelled.json()["details"]["job_status"] == "cancelled"

    def test_multi_branch_requester_gets_full_scope_file(
        self,
        client: TestClient,
        session_factory: sessionmaker[Session],
        render_dataset: DatasetRef,
        context: PostingContext,
        context_b: PostingContext,
        user_factory: UserFactory,
        exporter_role: str,
        test_password: str,
        render_storage: Path,
        seeded_books: None,
        seeded_books_b: None,
        low_xlsx_threshold: None,
    ) -> None:
        """C1 review 5E: tệp của job = bộ số của lượt render đồng bộ cùng người.

        User gán HAI chi nhánh, KHÔNG gửi `X-Branch` (`acting_branch_id=None` —
        đường hợp lệ): trước vá, `jobs.branch_id=None` làm thân job chạy phạm vi
        RỖNG → tệp 0 dòng dù ước lượng 4; sau vá, thân job chạy dưới danh sách
        chi nhánh HIỆN HÀNH của người yêu cầu (`REQUESTER_BRANCHES`) nên tệp
        mang đủ cả hai chi nhánh, khớp ước lượng.
        """
        headers = _headers(
            client,
            session_factory,
            render_dataset,
            user_factory,
            exporter_role,
            "rj-hai-cn",
            test_password,
            None,
        )
        job_id, estimated = _enqueue_render(client, headers)
        assert estimated >= 4, "hai chi nhánh × chứng từ 2 dòng phải cho ≥4 dòng"

        with unit_of_work(
            session_factory, posting_scope(render_dataset, context, user_id=ACTOR_ID)
        ) as session:
            assert queue.get_job(session, job_id).branch_id is None

        _run_render_job(session_factory, render_dataset, context, render_storage, job_id)

        with unit_of_work(
            session_factory, posting_scope(render_dataset, context, user_id=ACTOR_ID)
        ) as session:
            result = queue.get_job(session, job_id).result or {}
        assert result["row_count"] == estimated, (
            "thân job phải thấy ĐÚNG số dòng mà endpoint đã ước lượng dưới scope request"
        )

        done = client.get(f"/api/v1/reports/render-jobs/{job_id}/file", headers=headers)
        assert done.status_code == 200, done.text
        workbook = openpyxl.load_workbook(io.BytesIO(done.content), data_only=True)
        sheet = workbook.active
        assert sheet is not None
        totals = [row for row in sheet.iter_rows(values_only=True) if row and row[0] == "Tổng cộng"]
        assert totals, "thiếu dòng Tổng cộng"
        numbers = [value for value in totals[0] if isinstance(value, (int, float, Decimal))]
        # Mỗi chi nhánh một chứng từ 1.000.000 — tổng Nợ = tổng Có = 2.000.000
        # chỉ khi tệp mang CẢ HAI chi nhánh.
        assert [Decimal(str(value)) for value in numbers] == [
            Decimal(2_000_000),
            Decimal(2_000_000),
        ]

    def test_download_refuses_job_of_another_type(
        self,
        client: TestClient,
        session_factory: sessionmaker[Session],
        render_dataset: DatasetRef,
        context: PostingContext,
        user_factory: UserFactory,
        exporter_role: str,
        test_password: str,
        render_storage: Path,
        seeded_books: None,
        low_xlsx_threshold: None,
    ) -> None:
        """M-4 review 5E: endpoint tải chỉ phục vụ job `reporting.report.render`
        — một job loại khác (dù `done`, dù của chính mình) là 404 như không tồn
        tại, không phải 503 "tệp hỏng"."""
        headers = _headers(
            client,
            session_factory,
            render_dataset,
            user_factory,
            exporter_role,
            "rj-loai",
            test_password,
            context.branch_id,
        )
        job_id, _ = _enqueue_render(client, headers)
        _run_render_job(session_factory, render_dataset, context, render_storage, job_id)
        scope = posting_scope(render_dataset, context, user_id=ACTOR_ID)
        with unit_of_work(session_factory, scope) as session:
            job = queue.get_job(session, job_id)
            job.type = "system.diagnostic.slow_task"
            session.flush()
        try:
            response = client.get(f"/api/v1/reports/render-jobs/{job_id}/file", headers=headers)
            assert response.status_code == 404, response.text
        finally:
            with unit_of_work(session_factory, scope) as session:
                job = queue.get_job(session, job_id)
                job.type = RENDER_JOB_CODE
                session.flush()

    def test_render_job_cannot_be_enqueued_directly(
        self,
        client: TestClient,
        session_factory: sessionmaker[Session],
        render_dataset: DatasetRef,
        context: PostingContext,
        user_factory: UserFactory,
        test_password: str,
    ) -> None:
        """`direct_enqueue=False`: đường vào duy nhất là endpoint render — nơi
        mã báo cáo + tham số + ngưỡng đều được kiểm bằng metadata trong DB."""
        role = ensure_role(
            session_factory,
            render_dataset,
            "xep_hang_5e",
            [REPORT_VIEW, REPORT_EXPORT, "system.job.view", "system.job.create"],
        )
        headers = _headers(
            client,
            session_factory,
            render_dataset,
            user_factory,
            role,
            "rj-tay",
            test_password,
            context.branch_id,
        )
        response = client.post(
            "/api/v1/jobs",
            json={"type": RENDER_JOB_CODE, "params": {"code": "S03b-DN", "format": "xlsx"}},
            headers=headers,
        )
        assert response.status_code in (409, 422), response.text
        assert "job" in response.json()["error_code"]
