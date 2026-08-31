"""Sheet ngân hàng của số dư ban đầu (kind 1, lát 6D — nợ hoãn từ 4C).

Ba lời hứa mới của lát:

* Sheet "Số dư ngân hàng" nhập được: dòng kind 1 mang `bank_account_id`, tiền
  tệ phải khớp tiền tệ TK ngân hàng, TK kế toán phải nhóm 112; TK 112 nhập ở
  sheet Số dư tài khoản chỉ còn là CẢNH BÁO (dữ liệu trước 6D vẫn đọc được).
* Carry-forward giữ nhóm 1 qua năm VÀ chia phát sinh 112 theo TK ngân hàng
  (qua `DepositMovementSource` module bank cài); phát sinh 112 không qua chứng
  từ tiền gửi (GLE gõ thẳng) ở lại nhóm 0 — không bịa chủ tài khoản.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal

import pytest
from sqlalchemy import func, select, update
from sqlalchemy.orm import Session, sessionmaker

from bank_support import ensure_company_bank_account, seed_bank_package_data
from cash_book_support import seed_cash_book_package_data
from ket.kernel.contracts import PartnerKind
from ket.kernel.datasets.provisioning import DatasetRef
from ket.kernel.master_data.models.company_bank_account import CompanyBankAccount
from ket.kernel.persistence.unit_of_work import unit_of_work
from ket.modules.bank.models import BankVoucherKind
from ket.modules.bank.schemas import BankVoucherIn, BankVoucherLineIn
from ket.modules.bank.service import BankVoucherService
from ket.modules.cash_book.models import CashVoucherKind
from ket.modules.cash_book.schemas import CashVoucherIn, CashVoucherLineIn
from ket.modules.cash_book.service import CashVoucherService
from ket.modules.general_ledger.journal.schemas import JournalLineIn, JournalVoucherIn
from ket.modules.general_ledger.journal.service import JournalVoucherService
from ket.posting.engine.models import GlPosting, Ledger
from ket.posting.opening_balances.models import OpeningBalance, OpeningDetailKind
from ket.posting.opening_balances.template import ACCOUNT_SHEET, BANK_SHEET, build_opening_template
from posting_support import PostingContext, posting_scope, seed_posting_context
from test_opening_balance_parsing import workbook_of
from test_opening_balances_carry_forward import carry_forward, ensure_fiscal_year
from test_opening_balances_import import opening_rows, run_opening_import

pytestmark = pytest.mark.db

CUSTOMER_ID = 515151
"""Id đối tác bất kỳ: TK 131 của gói test đòi CHIỀU `customer`, và chiều nguồn
`master` chỉ được kiểm tồn tại ở tầng API — validator ghi sổ chỉ hỏi "đã điền
chưa" (xem `dimension_required`)."""

ACTOR_ID = 1
JAN_15 = date(2026, 1, 15)


@pytest.fixture
def context(session_factory: sessionmaker[Session], dataset_alpha: DatasetRef) -> PostingContext:
    return seed_posting_context(session_factory, dataset_alpha)


@pytest.fixture
def accounts(
    session_factory: sessionmaker[Session], dataset_alpha: DatasetRef, context: PostingContext
) -> dict[str, int]:
    seeded = seed_cash_book_package_data(session_factory, dataset_alpha, context)
    seed_bank_package_data(session_factory, dataset_alpha, context)
    return seeded


@pytest.fixture
def vnd_account(
    session_factory: sessionmaker[Session], dataset_alpha: DatasetRef, context: PostingContext
) -> tuple[int, str]:
    code = f"OB-{context.branch_code}-VND"
    account_id = ensure_company_bank_account(session_factory, dataset_alpha, context, code=code)
    return account_id, code


@pytest.fixture
def vnd_account_2(
    session_factory: sessionmaker[Session], dataset_alpha: DatasetRef, context: PostingContext
) -> tuple[int, str]:
    code = f"OB-{context.branch_code}-VND2"
    account_id = ensure_company_bank_account(session_factory, dataset_alpha, context, code=code)
    return account_id, code


@pytest.fixture
def usd_account(
    session_factory: sessionmaker[Session], dataset_alpha: DatasetRef, context: PostingContext
) -> tuple[int, str]:
    code = f"OB-{context.branch_code}-USD"
    account_id = ensure_company_bank_account(
        session_factory, dataset_alpha, context, code=code, currency_code="USD"
    )
    return account_id, code


def test_template_contains_bank_sheet() -> None:
    """Tệp mẫu tải về phải có sheet ngân hàng — hợp đồng FR-SYS-082."""
    from io import BytesIO

    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(build_opening_template()))
    assert "Số dư ngân hàng" in workbook.sheetnames
    header = [cell.value for cell in next(workbook["Số dư ngân hàng"].iter_rows(max_row=1))]
    assert header[0] == "Số tài khoản ngân hàng *"
    assert header[1] == "Số hiệu tài khoản *"


def test_bank_sheet_writes_kind1_rows_with_bank_account(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    context: PostingContext,
    vnd_account: tuple[int, str],
    vnd_account_2: tuple[int, str],
    usd_account: tuple[int, str],
    tmp_path: object,
) -> None:
    vnd_id, vnd_code = vnd_account
    vnd2_id, vnd2_code = vnd_account_2
    usd_id, usd_code = usd_account
    report = run_opening_import(
        session_factory,
        dataset_alpha,
        context,
        tmp_path,  # type: ignore[arg-type]
        workbook_of(
            {
                BANK_SHEET: [
                    [vnd_code, "112", None, None, None, None, 800_000, None],
                    # Cùng tiền tệ, cùng TK kế toán — khóa gộp PHẢI giữ hai
                    # dòng theo hai TK ngân hàng (review 6D, M18: bỏ chiều
                    # bank_account_id là gán nhầm toàn bộ cho TK đầu).
                    [vnd2_code, "112", None, None, None, None, 300_000, None],
                    [usd_code, "112", "USD", 25_000, 100, None, 2_500_000, None],
                ],
            }
        ),
    )
    assert report.committed, report.errors
    assert report.rows_by_kind == {OpeningDetailKind.BANK: 3}

    scope = posting_scope(dataset_alpha, context, user_id=ACTOR_ID)
    with unit_of_work(session_factory, scope) as session:
        rows = [
            row
            for row in opening_rows(session, context)
            if row.detail_kind == OpeningDetailKind.BANK
        ]
        by_account = {row.bank_account_id: row for row in rows}
        assert set(by_account) == {vnd_id, vnd2_id, usd_id}
        assert by_account[vnd_id].debit == Decimal(800_000)
        assert by_account[vnd2_id].debit == Decimal(300_000)
        assert by_account[vnd_id].currency_code == "VND"
        assert by_account[usd_id].debit_fc == Decimal(100)
        assert by_account[usd_id].debit == Decimal(2_500_000)
        assert by_account[usd_id].currency_code == "USD"


def test_bank_sheet_lookup_and_currency_validation(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    context: PostingContext,
    usd_account: tuple[int, str],
    tmp_path: object,
) -> None:
    """Mã lạ, sai tiền tệ, TK ngoài nhóm 112 — mỗi lỗi phải chỉ đúng dòng."""
    _, usd_code = usd_account
    report = run_opening_import(
        session_factory,
        dataset_alpha,
        context,
        tmp_path,  # type: ignore[arg-type]
        workbook_of(
            {
                BANK_SHEET: [
                    ["KHONG-TON-TAI", "112", None, None, None, None, 100_000, None],
                    # TK USD nhưng dòng để trống loại tiền (= VND).
                    [usd_code, "112", None, None, None, None, 100_000, None],
                    # TK kế toán ngoài nhóm 112.
                    [usd_code, "111", "USD", 25_000, 4, None, 100_000, None],
                ],
            }
        ),
        commit=False,
    )
    codes = {error.code for error in report.errors}
    assert "opening.bank_account_unknown" in codes
    assert "opening.bank_account_currency_mismatch" in codes
    assert "opening.bank_account_not_deposit" in codes
    assert not report.committed


def test_deposit_account_on_plain_sheet_warns_but_commits(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    context: PostingContext,
    tmp_path: object,
) -> None:
    """Dữ liệu kiểu trước-6D (112 ở sheet Số dư tài khoản) vẫn ghi được —
    nhưng lượt kiểm phải nhắc chuyển sang sheet ngân hàng (BR-BNK-01)."""
    report = run_opening_import(
        session_factory,
        dataset_alpha,
        context,
        tmp_path,  # type: ignore[arg-type]
        workbook_of({ACCOUNT_SHEET: [["112", None, None, None, None, 500_000, None]]}),
    )
    assert report.committed
    assert any(warning.code == "opening.deposit_on_account_sheet" for warning in report.warnings)


def test_carry_forward_splits_deposit_movement_by_bank_account(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    context: PostingContext,
    accounts: dict[str, int],
    vnd_account: tuple[int, str],
    vnd_account_2: tuple[int, str],
    tmp_path: object,
) -> None:
    """Năm nguồn: dư đầu ngân hàng 800k + báo có 700k + chuyển nội bộ 250k
    sang TK thứ hai + GLE gõ thẳng 112 100k.

    Năm nhận phải ra: kind 1 TK nguồn = 800 + 700 − 250 = 1.250k; kind 1 TK
    đích = 250k (CTNB quy chủ theo CHIỀU dòng — review 6D, M9); phần GLE 100k
    ở lại kind 0 vì không ai biết nó thuộc tài khoản nào.
    """
    vnd_id, vnd_code = vnd_account
    vnd2_id, _vnd2_code = vnd_account_2
    report = run_opening_import(
        session_factory,
        dataset_alpha,
        context,
        tmp_path,  # type: ignore[arg-type]
        workbook_of({BANK_SHEET: [[vnd_code, "112", None, None, None, None, 800_000, None]]}),
    )
    assert report.committed, report.errors

    scope = posting_scope(dataset_alpha, context, user_id=ACTOR_ID)
    with unit_of_work(session_factory, scope) as session:
        bank_service = BankVoucherService(session)
        voucher = bank_service.create(
            BankVoucherIn(
                kind=BankVoucherKind.CREDIT_ADVICE,
                operation_code="thu-khac",
                bank_account_id=vnd_id,
                branch_id=context.branch_id,
                document_date=JAN_15,
                posting_date=JAN_15,
                currency_code="VND",
                exchange_rate=Decimal(1),
                description="báo có carry-forward",
                lines=(
                    BankVoucherLineIn(
                        debit_account_id=accounts["112"],
                        credit_account_id=accounts["3381"],
                        amount_fc=Decimal(700_000),
                    ),
                ),
            ),
            user_id=ACTOR_ID,
        )
        bank_service.post(voucher.id, user_id=ACTOR_ID)
        transfer = bank_service.create(
            BankVoucherIn(
                kind=BankVoucherKind.INTERNAL_TRANSFER,
                operation_code=None,
                bank_account_id=vnd_id,
                counter_bank_account_id=vnd2_id,
                branch_id=context.branch_id,
                document_date=JAN_15,
                posting_date=JAN_15,
                currency_code="VND",
                exchange_rate=Decimal(1),
                description="chuyển nội bộ carry-forward",
                lines=(
                    BankVoucherLineIn(
                        debit_account_id=accounts["112"],
                        credit_account_id=accounts["112"],
                        amount_fc=Decimal(250_000),
                    ),
                ),
            ),
            user_id=ACTOR_ID,
        )
        bank_service.post(transfer.id, user_id=ACTOR_ID)

        journal = JournalVoucherService(session)
        gle = journal.create(
            JournalVoucherIn(
                branch_id=context.branch_id,
                document_date=JAN_15,
                posting_date=JAN_15,
                currency_code="VND",
                exchange_rate=Decimal(1),
                description="GLE gõ thẳng 112",
                lines=(
                    JournalLineIn(account_id=accounts["112"], debit_fc=Decimal(100_000)),
                    JournalLineIn(account_id=accounts["3381"], credit_fc=Decimal(100_000)),
                ),
            ),
            user_id=ACTOR_ID,
        )
        journal.post(gle.id, user_id=ACTOR_ID)

    with unit_of_work(session_factory, scope) as session:
        target_id = ensure_fiscal_year(session, "2027", date(2027, 1, 1))
        result = carry_forward(session, dataset_alpha, context)
        assert result["target_fiscal_year_id"] == target_id

    with unit_of_work(session_factory, scope) as session:
        rows = (
            session.execute(
                select(OpeningBalance)
                .where(OpeningBalance.fiscal_year_id == target_id)
                .where(OpeningBalance.branch_id == context.branch_id)
                .where(OpeningBalance.ledger == Ledger.FINANCIAL)
                .where(OpeningBalance.account_id == accounts["112"])
            )
            .scalars()
            .all()
        )
        by_kind = {(row.detail_kind, row.bank_account_id): row for row in rows}
        source_row = by_kind[(OpeningDetailKind.BANK, vnd_id)]
        assert source_row.debit == Decimal(1_250_000)
        target_row = by_kind[(OpeningDetailKind.BANK, vnd2_id)]
        assert target_row.debit == Decimal(250_000)
        plain_row = by_kind[(OpeningDetailKind.ACCOUNT, None)]
        assert plain_row.debit == Decimal(100_000)


def test_a_deposit_line_carrying_another_dimension_yields_exactly_one_row(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    accounts: dict[str, int],
    vnd_account: tuple[int, str],
) -> None:
    """Review 6G-1 H-2 — một dòng 112 mang THÊM một chiều khác chỉ được sinh
    MỘT dòng dư đầu năm.

    Bản 6D chia chiều TK ngân hàng bằng cách cộng vào xô (TK, TK ngân hàng) rồi
    TRỪ khỏi xô "chưa gắn". Dòng trừ khai mọi chiều = NULL còn dòng gốc giữ
    nguyên chiều của nó, nên hai dòng rơi vào hai xô của `GROUP BY` và phép trừ
    không triệt tiêu: kết quả là BA dòng — dòng gốc còn nguyên, một dòng ÂM bịa
    ở nhóm chưa-gắn, và dòng ngân hàng đúng. Tổng TK 112 vẫn khớp nên không
    cổng nào kêu; chỉ chi tiết dư đầu năm sau là sai, và BR-BNK-01 đọc trên
    chính tập ấy.

    Lát 6G-1 làm lỗi này lộ ra vì nguồn không còn giới hạn ở chứng từ tiền gửi
    (bên tiền của chúng mang bộ chiều RỖNG nên hầu như không ai gặp).
    """
    vnd_id, _vnd_code = vnd_account
    cost_object_id = 424242
    # Chi nhánh RIÊNG: `carry_forward` chạy theo (sổ, chi nhánh) và xóa sạch năm
    # nhận trước khi ghi, nên gieo thêm chứng từ vào chi nhánh dùng chung sẽ đổi
    # kết quả của bài kiểm ngay bên trên tùy thứ tự chạy.
    context = seed_posting_context(session_factory, dataset_alpha)
    scope = posting_scope(dataset_alpha, context, user_id=ACTOR_ID)
    with unit_of_work(session_factory, scope) as session:
        journal = JournalVoucherService(session)
        gle = journal.create(
            JournalVoucherIn(
                branch_id=context.branch_id,
                document_date=JAN_15,
                posting_date=JAN_15,
                currency_code="VND",
                exchange_rate=Decimal(1),
                description="112 kèm khoản mục chi phí",
                lines=(
                    JournalLineIn(
                        account_id=accounts["112"],
                        debit_fc=Decimal(100_000),
                        bank_account_id=vnd_id,
                        cost_object_id=cost_object_id,
                    ),
                    JournalLineIn(account_id=accounts["3381"], credit_fc=Decimal(100_000)),
                ),
            ),
            user_id=ACTOR_ID,
        )
        journal.post(gle.id, user_id=ACTOR_ID)

    with unit_of_work(session_factory, scope) as session:
        # Năm nhận là năm bắt đầu ngay sau năm nguồn — không chọn được số khác.
        target_id = ensure_fiscal_year(session, "2027", date(2027, 1, 1))
        carry_forward(session, dataset_alpha, context)

    with unit_of_work(session_factory, scope) as session:
        rows = (
            session.execute(
                select(OpeningBalance)
                .where(OpeningBalance.fiscal_year_id == target_id)
                .where(OpeningBalance.branch_id == context.branch_id)
                .where(OpeningBalance.ledger == Ledger.FINANCIAL)
                .where(OpeningBalance.account_id == accounts["112"])
                .where(OpeningBalance.cost_object_id == cost_object_id)
            )
            .scalars()
            .all()
        )
        assert len(rows) == 1, [
            (row.detail_kind, row.bank_account_id, row.debit, row.credit) for row in rows
        ]
        only = rows[0]
        assert only.detail_kind == OpeningDetailKind.BANK
        assert only.bank_account_id == vnd_id
        assert only.debit == Decimal(100_000)
        assert only.credit == Decimal(0)

        # Và KHÔNG có dòng âm bịa nào ở nhóm chưa-gắn cho chính TK ấy.
        orphan = session.execute(
            select(func.count())
            .select_from(OpeningBalance)
            .where(OpeningBalance.fiscal_year_id == target_id)
            .where(OpeningBalance.branch_id == context.branch_id)
            .where(OpeningBalance.ledger == Ledger.FINANCIAL)
            .where(OpeningBalance.account_id == accounts["112"])
            .where(OpeningBalance.bank_account_id.is_(None))
            .where(OpeningBalance.credit > 0)
        ).scalar_one()
        assert orphan == 0


def test_a_receivable_line_paired_with_112_stays_a_receivable(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    accounts: dict[str, int],
    vnd_account: tuple[int, str],
) -> None:
    """Review pre-landing H-A — hồi quy do CHÍNH bản vá H-2 tạo ra.

    Dòng phiếu quỹ `Nợ 131 / Có 112` không chạm TK quỹ, nên luật chung cho **cả
    hai** bên nhận trọn chiều — kể cả `bank_account`. Bản viết lại
    `carry_forward.sql` đọc cột ấy trên MỌI dòng và gán `detail_kind = 1` cho
    bất kỳ dòng nào có nó, nên dư PHẢI THU bị xếp sang nhóm tiền gửi. Lượt
    chuyển hóa đơn chỉ nhận nhóm 2–4 ⇒ chi tiết hóa đơn rơi im lặng, và khóa
    ngoại `RESTRICT` ghim luôn TK ngân hàng vào một dòng công nợ.

    Nay chiều bị lọc ở CẢ HAI đầu: mapper không dán nó lên bên không phải 112x,
    và câu chuyển năm chỉ coi nó là chiều ngân hàng khi số hiệu TK bắt đầu 112.
    """
    vnd_id, _vnd_code = vnd_account
    context = seed_posting_context(session_factory, dataset_alpha)
    scope = posting_scope(dataset_alpha, context, user_id=ACTOR_ID)

    with unit_of_work(session_factory, scope) as session:
        service = CashVoucherService(session)
        voucher = service.create(
            CashVoucherIn(
                kind=CashVoucherKind.RECEIPT,
                operation_code="thu-khac",
                cash_account_id=accounts["111"],
                branch_id=context.branch_id,
                document_date=JAN_15,
                posting_date=JAN_15,
                currency_code="VND",
                exchange_rate=Decimal(1),
                payer_receiver_name="Khách trả qua ngân hàng",
                description="Nợ 131 / Có 112 — không chạm TK quỹ",
                lines=(
                    CashVoucherLineIn(
                        debit_account_id=accounts["131"],
                        credit_account_id=accounts["112"],
                        amount_fc=Decimal(500_000),
                        partner_kind=PartnerKind.CUSTOMER,
                        partner_id=CUSTOMER_ID,
                        bank_account_id=vnd_id,
                    ),
                ),
            ),
            user_id=ACTOR_ID,
        )
        service.post(voucher.id, user_id=ACTOR_ID)

        # Ở SỔ: chiều chỉ nằm trên bên 112, không đổ sang bên 131.
        owners = dict(
            session.execute(
                select(GlPosting.account_id, GlPosting.bank_account_id).where(
                    GlPosting.voucher_id == voucher.id,
                    GlPosting.ledger == Ledger.FINANCIAL,
                )
            ).all()
        )
        assert owners[accounts["112"]] == vnd_id
        assert owners[accounts["131"]] is None

    with unit_of_work(session_factory, scope) as session:
        target_id = ensure_fiscal_year(session, "2027", date(2027, 1, 1))
        carry_forward(session, dataset_alpha, context)

    with unit_of_work(session_factory, scope) as session:
        receivable = (
            session.execute(
                select(OpeningBalance)
                .where(OpeningBalance.fiscal_year_id == target_id)
                .where(OpeningBalance.branch_id == context.branch_id)
                .where(OpeningBalance.ledger == Ledger.FINANCIAL)
                .where(OpeningBalance.account_id == accounts["131"])
            )
            .scalars()
            .all()
        )
        assert len(receivable) == 1, [(row.detail_kind, row.bank_account_id) for row in receivable]
        assert receivable[0].detail_kind == OpeningDetailKind.RECEIVABLE
        assert receivable[0].bank_account_id is None


def test_carry_forward_repairs_rows_written_before_the_dimension_was_filtered(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    accounts: dict[str, int],
    vnd_account: tuple[int, str],
) -> None:
    """Lớp phòng thủ THỨ HAI của review pre-landing H-A + H-B, kiểm riêng.

    Hai bản vá chồng nhau: mapper thôi dán `bank_account` lên bên không phải
    112x, và câu chuyển năm chỉ coi cột ấy là chiều ngân hàng khi số hiệu TK bắt
    đầu `112` **và** TK ngân hàng còn trong danh mục. Với bản vá mapper tại chỗ,
    lớp SQL không bao giờ được chạm tới — nên bài kiểm này ghi thẳng vào
    `gl_postings` hai hình dạng mà lớp ấy tồn tại để chịu:

    * dòng công nợ mang `bank_account_id` — đúng thứ dữ liệu đã ghi sổ TRƯỚC bản
      vá mapper đang mang, và cũng là thứ một module tương lai quên lọc sẽ tạo;
    * dòng 112 trỏ một TK ngân hàng đã bị xóa khỏi danh mục — `gl_postings`
      không có khóa ngoại còn `opening_balances` thì có, nên thiếu lớp này lượt
      khóa sổ cuối năm đổ `ForeignKeyViolation` (H-B).
    """
    vnd_id, _vnd_code = vnd_account
    context = seed_posting_context(session_factory, dataset_alpha)
    scope = posting_scope(dataset_alpha, context, user_id=ACTOR_ID)
    missing_bank_id = 987_654_321

    with unit_of_work(session_factory, scope) as session:
        journal = JournalVoucherService(session)
        voucher = journal.create(
            JournalVoucherIn(
                branch_id=context.branch_id,
                document_date=JAN_15,
                posting_date=JAN_15,
                currency_code="VND",
                exchange_rate=Decimal(1),
                description="dựng hai hình dạng cho lớp phòng thủ SQL",
                lines=(
                    JournalLineIn(
                        account_id=accounts["131"],
                        debit_fc=Decimal(500_000),
                        partner_kind=PartnerKind.CUSTOMER,
                        partner_id=CUSTOMER_ID,
                    ),
                    JournalLineIn(
                        account_id=accounts["112"],
                        credit_fc=Decimal(500_000),
                        bank_account_id=vnd_id,
                    ),
                ),
            ),
            user_id=ACTOR_ID,
        )
        journal.post(voucher.id, user_id=ACTOR_ID)

        # Ghi THẲNG vào sổ, vòng qua mapper: đúng hình dạng dữ liệu cũ.
        session.execute(
            update(GlPosting)
            .where(GlPosting.voucher_id == voucher.id)
            .where(GlPosting.account_id == accounts["131"])
            .values(bank_account_id=vnd_id)
        )
        session.execute(
            update(GlPosting)
            .where(GlPosting.voucher_id == voucher.id)
            .where(GlPosting.account_id == accounts["112"])
            .values(bank_account_id=missing_bank_id)
        )

    with unit_of_work(session_factory, scope) as session:
        target_id = ensure_fiscal_year(session, "2027", date(2027, 1, 1))
        carry_forward(session, dataset_alpha, context)  # H-B: không được đổ

    with unit_of_work(session_factory, scope) as session:
        rows = {
            row.account_id: row
            for row in session.execute(
                select(OpeningBalance)
                .where(OpeningBalance.fiscal_year_id == target_id)
                .where(OpeningBalance.branch_id == context.branch_id)
                .where(OpeningBalance.ledger == Ledger.FINANCIAL)
                .where(OpeningBalance.account_id.in_([accounts["131"], accounts["112"]]))
            )
            .scalars()
            .all()
        }
        # H-A: dòng công nợ giữ nguyên nhóm phải thu dù mang cột ấy.
        assert rows[accounts["131"]].detail_kind == OpeningDetailKind.RECEIVABLE
        assert rows[accounts["131"]].bank_account_id is None
        # H-B: TK ngân hàng mồ côi hạ về nhóm chưa-gắn, tiền vẫn chuyển đủ.
        assert rows[accounts["112"]].detail_kind == OpeningDetailKind.ACCOUNT
        assert rows[accounts["112"]].bank_account_id is None
        assert rows[accounts["112"]].credit == Decimal(500_000)


def test_bank_account_with_opening_rows_cannot_be_deleted(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    context: PostingContext,
    vnd_account: tuple[int, str],
    tmp_path: object,
) -> None:
    """FK RESTRICT: còn số dư treo thì không xóa được TK ngân hàng."""
    from sqlalchemy.exc import IntegrityError

    vnd_id, vnd_code = vnd_account
    report = run_opening_import(
        session_factory,
        dataset_alpha,
        context,
        tmp_path,  # type: ignore[arg-type]
        workbook_of({BANK_SHEET: [[vnd_code, "112", None, None, None, None, 250_000, None]]}),
    )
    assert report.committed

    scope = posting_scope(dataset_alpha, context, user_id=ACTOR_ID)
    with pytest.raises(IntegrityError):
        with unit_of_work(session_factory, scope) as session:
            account = session.get(CompanyBankAccount, vnd_id)
            assert account is not None
            session.delete(account)
            session.flush()
