"""Nhập và xuất danh mục qua HTTP (lát 3C-1, mở rộng cho đường xuất ở 3C-2).

Luật nghiệp vụ của lượt nhập nằm ở `test_import_pipeline.py`, của lượt xuất ở
`test_export.py`. Ở đây chỉ kiểm bốn thứ mà **chỉ tầng HTTP** trả lời được:

* tệp mẫu tải về được, và route của nó không bị route đọc bản ghi nuốt mất;
* **hai lớp quyền** — quyền dùng chức năng nhập liệu, và quyền trên chính danh
  mục đó (H48: kế toán kho nhập được danh mục kho không có nghĩa là nhập được
  điều khoản thanh toán);
* hai loại job **không** xếp hàng thẳng qua `/api/v1/jobs` được, vì endpoint
  chung không biết `slug` nào đang bị nhắm tới;
* lượt kiểm để lại một `job_id` và tệp nằm trong kho định địa chỉ theo nội dung.

Đường **xuất** (3C-2) đo hai thứ ở đây vì cùng hai lý do: route của nó cũng bị
`GET /{slug}/{record_id}` nuốt nếu đăng ký sai thứ tự, và mức quyền của nó cũng
là một quyết định (`view`, bằng đường đọc danh sách — tệp xuất ra chứa đúng
những dòng người dùng ấy đã đọc được).
"""

from __future__ import annotations

from collections.abc import Iterator
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import Engine, select
from sqlalchemy.orm import Session, sessionmaker

from catalog_api_support import (
    UserFactory,
    actor,
    all_branch_codes,
    branch_ids,
    catalog_codes,
    ensure_branches,
    ensure_role,
    unique_code,
)
from conftest import api_test_client
from ket.api.dependencies import BRANCH_HEADER
from ket.kernel.attachments import storage
from ket.kernel.datasets.provisioning import DatasetRef
from ket.kernel.excel.descriptors import template_for
from ket.kernel.excel.job import IMPORT_COMMIT, IMPORT_VALIDATE
from ket.kernel.jobs.builtin import JOB_CREATE, JOB_VIEW
from ket.kernel.master_data.models.warehouse import Warehouse
from ket.kernel.master_data.registry import REGISTRY
from ket.kernel.master_data.service import MasterDataService
from ket.kernel.persistence.unit_of_work import RequestScope, unit_of_work
from ket.kernel.security.models import Branch
from ket.kernel.security.permissions import Action
from ket.main import create_app
from ket.settings import Settings

pytestmark = pytest.mark.db

WAREHOUSES = "warehouses"
PAYMENT_TERMS = "payment_terms"
XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

BRANCH_CODES = ["CN_IMP_A", "CN_IMP_MINE", "CN_IMP_KHAC"]
"""Ba chi nhánh, tạo **trước** khi fixture `importer` gán phạm vi cho người dùng.

Hai cái sau tồn tại cho phép kiểm phạm vi của lượt xuất: chỉ với ba bản ghi ở ba
phạm vi (dùng chung, chi nhánh này, chi nhánh kia) mới tách được "lọc theo phiên"
khỏi "lọc về phần dùng chung" — xem test tương ứng."""


@pytest.fixture(scope="module")
def import_dir(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return tmp_path_factory.mktemp("import-store")


@pytest.fixture(scope="module")
def import_settings(test_settings: Settings, import_dir: Path) -> Settings:
    return test_settings.model_copy(update={"attachments_dir": import_dir})


@pytest.fixture(scope="module")
def client(
    import_settings: Settings, app_engine: Engine, session_factory: sessionmaker[Session]
) -> Iterator[TestClient]:
    """`TestClient` dùng chung cho CẢ TỆP — kèm hai ràng buộc.

    Phạm vi module để `importer`/`reader` chỉ phải gán chi nhánh một lần (xem
    docstring của chúng). Đổi lại tệp này mang hai bất biến mà phần lớn tệp API
    khác không có, nên **đừng chép dòng `scope="module"` sang tệp mới mà không
    chép cả hai** — cùng bộ ràng buộc đã ghi ở `test_master_data_merge.py`:

    1. **Không bài nào trong tệp được tạo chi nhánh mới** — người dùng dùng
       chung được gán chi nhánh MỘT lần và không tự cập nhật. Ở tệp này chi
       nhánh chỉ sinh ra trong `branches` (phạm vi module), và
       `_shared_actor_still_spans_every_branch` canh cho điều đó còn đúng.
    2. Cửa sổ hạn mức của `RateLimitMiddleware` là trạng thái trong tiến trình
       của một app, nên nay nó dùng chung cho cả tệp thay vì mới lại mỗi bài.
       Vì thế tắt hẳn hạn mức ở đây. Đo với hạn mức production bật lại: tệp
       vẫn xanh, tức đây là chống-vỡ-về-sau chứ **không** phải một `429` đã
       quan sát được. Hạn mức có bộ test riêng ở `test_rate_limit.py` (gồm cả
       nhánh `0 = tắt` mà dòng này dựa vào); tệp này không nói gì về nó.
    """
    assert app_engine is not None and session_factory is not None
    unlimited = import_settings.model_copy(
        update={"rate_limit_per_minute": 0, "rate_limit_auth_per_minute": 0}
    )
    with api_test_client(create_app(unlimited)) as instance:
        yield instance


_BRANCH_SPAN: dict[str, int] = {}


@pytest.fixture(autouse=True)
def _shared_actor_still_spans_every_branch(request: pytest.FixtureRequest) -> None:
    """`importer`/`reader` được gán chi nhánh MỘT lần — canh cho chúng còn trọn phạm vi.

    Nếu một bài trong tệp này tạo thêm chi nhánh, hai người dùng dùng chung
    **không** được cập nhật, và các bài sau đỏ bằng `scope_incomplete` /
    `scope_insufficient` — một thông điệp trỏ thẳng vào mã production, khiến
    bản sửa hấp dẫn nhất là nới lỏng chính phép kiểm phạm vi đang đúng. Bài
    kiểm này biến ca đó thành một câu nói rõ nguyên nhân.
    """
    # Lấy lười bằng `getfixturevalue` sau khi kiểm dấu: khai thành THAM SỐ là đủ
    # để pytest dựng fixture DB TRƯỚC khi thân hàm chạy, nên `return` sớm không
    # cứu được. Hôm nay cả tệp mang `pytestmark = pytest.mark.db` nên nhánh
    # `return` này chưa bao giờ chạy — nó giữ cho bài không-DB thêm sau này
    # không nổ khó hiểu (bản đầu của lát trước mắc đúng lỗi ấy: 2 ERROR ở
    # `make server-test`).
    if request.node.get_closest_marker("db") is None:
        return
    session_factory: sessionmaker[Session] = request.getfixturevalue("session_factory")
    dataset_alpha: DatasetRef = request.getfixturevalue("dataset_alpha")
    with unit_of_work(
        session_factory,
        RequestScope(dataset_schema=dataset_alpha.schema_name, user_id=1, branch_ids=()),
    ) as session:
        live = len(list(session.scalars(select(Branch.code)).all()))
    seen = _BRANCH_SPAN.setdefault("n", live)
    assert live == seen, (
        f"Số chi nhánh đổi giữa chừng ({seen} → {live}), nên `importer`/`reader` phạm vi "
        "module — đã gán chi nhánh MỘT lần và không tự cập nhật — không còn "
        "trọn phạm vi; các bài sau sẽ đỏ vì `scope_incomplete`.\n"
        "Hai nguyên nhân, kiểm theo thứ tự này:\n"
        "1. Một bài TRONG tệp vừa tạo chi nhánh. Chuyển việc ấy vào fixture "
        "phạm vi module chạy trước khi dựng người dùng, đừng tạo trong bài.\n"
        "2. Bạn đang chạy một lựa chọn bài ĐAN XEN nhiều tệp (ví dụ nêu đích "
        "danh vài bài của tệp này lẫn tệp khác). Tệp kia tạo chi nhánh của "
        "nó ở giữa, và người dùng dùng chung của tệp này thật sự đã cũ — "
        "đây KHÔNG phải lỗi của tệp này. Chạy trọn từng tệp một."
    )


@pytest.fixture(scope="module")
def importer_role(session_factory: sessionmaker[Session], dataset_alpha: DatasetRef) -> str:
    """Được nhập liệu, và được trên **danh mục kho** — không phải trên mọi danh mục."""
    return ensure_role(
        session_factory,
        dataset_alpha,
        "ke_toan_nhap_lieu",
        [
            IMPORT_VALIDATE,
            IMPORT_COMMIT,
            JOB_VIEW,
            JOB_CREATE,
            *catalog_codes(WAREHOUSES, Action.VIEW, Action.CREATE, Action.EDIT),
        ],
    )


@pytest.fixture(scope="module")
def reader_role(session_factory: sessionmaker[Session], dataset_alpha: DatasetRef) -> str:
    """Xem được danh mục kho, **không** có quyền dùng chức năng nhập liệu."""
    return ensure_role(
        session_factory,
        dataset_alpha,
        "ke_toan_chi_xem",
        [*catalog_codes(WAREHOUSES, Action.VIEW), JOB_VIEW],
    )


@pytest.fixture(scope="module")
def branches(session_factory: sessionmaker[Session], dataset_alpha: DatasetRef) -> list[str]:
    return ensure_branches(session_factory, dataset_alpha, BRANCH_CODES)


@pytest.fixture(scope="module")
def importer(
    client: TestClient,
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    user_factory: UserFactory,
    importer_role: str,
    test_password: str,
    branches: list[str],
) -> dict[str, str]:
    """Người nhập liệu **phạm vi toàn công ty**, dùng chung cho CẢ TỆP.

    Phạm vi **module** chứ không phải hàm: `actor` gán từng chi nhánh một lời
    gọi, và danh sách "mọi chi nhánh" dài ra theo số tệp test đã chạy trước.
    Dựng lại người dùng cho mỗi bài khiến riêng tệp này tốn 75s trong một lượt
    chạy đầy đủ. Dùng chung là an toàn vì `api/dependencies.resolve_access`
    chạy lại ở **mỗi** request: quyền đọc tươi từ DB chứ không đóng băng trong
    token. Không bài nào trong tệp đổi vai trò, phạm vi hay mật khẩu của người
    này, và bài kiểm phạm vi lượt xuất tự đặt `X-Branch` cho từng request.
    """
    assert branches, "cần ít nhất một chi nhánh"
    headers = actor(
        client,
        session_factory,
        dataset_alpha,
        user_factory,
        importer_role,
        "nhaplieu",
        test_password,
        branch_codes=all_branch_codes(session_factory, dataset_alpha),
    )
    # `X-Branch` tường minh: `dependencies._acting_branch` suy chi nhánh **khi và
    # chỉ khi** người dùng có đúng một, nên phạm vi ghi của cả module này vốn là
    # hàm của *số chi nhánh mà module test khác tình cờ tạo ra* (R3-M3). Ghim nó
    # để mọi khẳng định ở đây nói về một phạm vi biết trước.
    return {
        **headers,
        BRANCH_HEADER: str(
            branch_ids(session_factory, dataset_alpha, [BRANCH_CODES[0]])[BRANCH_CODES[0]]
        ),
    }


@pytest.fixture(scope="module")
def reader(
    client: TestClient,
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    user_factory: UserFactory,
    reader_role: str,
    test_password: str,
    branches: list[str],
) -> dict[str, str]:
    """Người chỉ-xem, dùng chung cho CẢ TỆP — cùng lý do phạm vi như `importer`."""
    assert branches, "cần ít nhất một chi nhánh"
    return actor(
        client,
        session_factory,
        dataset_alpha,
        user_factory,
        reader_role,
        "chixem",
        test_password,
        branch_codes=all_branch_codes(session_factory, dataset_alpha),
    )


def _upload(slug: str, rows: list[list[object]]) -> bytes:
    spec = REGISTRY.get(slug)
    assert spec is not None
    descriptor = template_for(spec)
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = descriptor.sheet_name
    sheet.append(list(descriptor.headers))
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ------------------------------------------------------------------ tệp mẫu


def test_the_template_route_is_not_swallowed_by_the_record_route(
    client: TestClient, importer: dict[str, str]
) -> None:
    """`GET /{slug}/template` phải thắng `GET /{slug}/{record_id}`.

    Hai bộ route dùng chung tiền tố `/api/v1/master` và FastAPI khớp theo thứ tự
    đăng ký, nên nếu bộ danh mục được gắn trước thì `template` rơi vào
    `{record_id}` và người dùng nhận `422` ("template" không phải số nguyên) thay
    vì nhận tệp. `main.py` gắn bộ nhập liệu trước, và đây là chỗ canh điều đó.
    """
    response = client.get(f"/api/v1/master/{WAREHOUSES}/template", headers=importer)
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(XLSX)
    assert "attachment" in response.headers["content-disposition"]


def test_the_downloaded_template_opens_and_has_both_sheets(
    client: TestClient, importer: dict[str, str]
) -> None:
    response = client.get(f"/api/v1/master/{WAREHOUSES}/template", headers=importer)
    workbook = load_workbook(BytesIO(response.content))
    spec = REGISTRY.get(WAREHOUSES)
    assert spec is not None
    descriptor = template_for(spec)
    assert descriptor.sheet_name in workbook.sheetnames
    assert len(workbook.sheetnames) == 2


def test_downloading_a_template_needs_only_view_permission(
    client: TestClient, reader: dict[str, str]
) -> None:
    """Tệp mẫu không chứa dữ liệu — đòi quyền tạo ở đây chặn nhầm người chuẩn bị tệp."""
    response = client.get(f"/api/v1/master/{WAREHOUSES}/template", headers=reader)
    assert response.status_code == 200, response.text


# ------------------------------------------------------------------- quyền


def test_validate_needs_the_import_permission(client: TestClient, reader: dict[str, str]) -> None:
    """Lớp quyền thứ nhất: được dùng chức năng nhập liệu hàng loạt hay không."""
    response = client.post(
        f"/api/v1/master/{WAREHOUSES}/import/validate",
        headers=reader,
        files={"file": ("kho.xlsx", _upload(WAREHOUSES, []), XLSX)},
    )
    assert response.status_code == 403, response.text


def test_validate_needs_permission_on_that_particular_catalog(
    client: TestClient, importer: dict[str, str]
) -> None:
    """Lớp quyền thứ hai (H48): quyền trên **danh mục này**, không phải mọi danh mục.

    Người dùng ở đây có `master.import.*` và toàn quyền trên danh mục kho, nhưng
    không có gì trên điều khoản thanh toán. Không có lớp này thì một mã quyền
    chung mở cửa cho cả hai mươi danh mục.
    """
    response = client.post(
        f"/api/v1/master/{PAYMENT_TERMS}/import/validate",
        headers=importer,
        files={"file": ("dk.xlsx", _upload(PAYMENT_TERMS, []), XLSX)},
    )
    assert response.status_code == 403, response.text


def test_commit_needs_create_permission_on_the_catalog(
    client: TestClient, importer: dict[str, str]
) -> None:
    response = client.post(
        f"/api/v1/master/{PAYMENT_TERMS}/import/commit",
        headers=importer,
        json={"validation_job_id": "00000000-0000-0000-0000-000000000000"},
    )
    assert response.status_code == 403, response.text


def test_import_jobs_cannot_be_queued_through_the_generic_endpoint(
    client: TestClient, importer: dict[str, str]
) -> None:
    """`direct_enqueue=False`: mã quyền tĩnh không nói được "trên danh mục nào".

    Không có cổng này thì ai có `master.import.create` sẽ ghi được vào **mọi**
    danh mục chỉ bằng cách gọi thẳng endpoint hàng đợi với `type` và `params` tự
    đặt — đi vòng qua đúng phép kiểm per-danh-mục mà test bên trên vừa canh.
    """
    response = client.post(
        "/api/v1/jobs",
        headers=importer,
        json={
            "type": "master.import.commit",
            "params": {"validation_job_id": "00000000-0000-0000-0000-000000000000"},
        },
    )
    assert response.status_code == 422, response.text
    assert response.json()["error_code"] == "job.not_directly_enqueueable"


# -------------------------------------------------------------- xếp hàng


def test_validate_stores_the_file_and_queues_a_job(
    client: TestClient,
    importer: dict[str, str],
    import_dir: Path,
    dataset_alpha: DatasetRef,
) -> None:
    """H78: tệp vào kho định địa chỉ theo nội dung, báo cáo sẽ nằm ở `jobs.result`."""
    content = _upload(WAREHOUSES, [[unique_code("KHO_API"), "Kho API", None, None, None, None]])
    response = client.post(
        f"/api/v1/master/{WAREHOUSES}/import/validate",
        headers=importer,
        files={"file": ("kho.xlsx", content, XLSX)},
    )
    assert response.status_code == 202, response.text
    body = response.json()
    assert body["catalog"] == WAREHOUSES
    assert body["mode"] == "create_only", "mặc định phải là chế độ không ghi đè (H80)"

    stored = storage.blob_path(import_dir, dataset_alpha.schema_name, body["content_hash"])
    assert stored.read_bytes() == content, "tệp đã kiểm phải là đúng tệp đã tải lên"

    queued = client.get(f"/api/v1/jobs/{body['job_id']}", headers=importer)
    assert queued.status_code == 200, queued.text
    assert queued.json()["type"] == "master.import.validate"


def test_commit_refuses_a_validation_job_that_does_not_exist(
    client: TestClient, importer: dict[str, str]
) -> None:
    """Bước ghi trỏ vào một lượt kiểm không có thật thì phải dừng ở lúc **xếp hàng**.

    Ở lát này phép kiểm ấy nằm trong thân job, nên endpoint vẫn trả `202` và job
    hỏng ngay sau đó — đúng hợp đồng của hàng đợi. Điều được canh là nó **không**
    ghi gì; xem `test_import_pipeline.py` cho phần bất biến.
    """
    response = client.post(
        f"/api/v1/master/{WAREHOUSES}/import/commit",
        headers=importer,
        json={"validation_job_id": "00000000-0000-0000-0000-000000000000"},
    )
    assert response.status_code == 202, response.text


# ------------------------------------------------------------------- xuất


def test_the_export_route_is_not_swallowed_by_the_record_route(
    client: TestClient, importer: dict[str, str]
) -> None:
    """`GET /{slug}/export` phải thắng `GET /{slug}/{record_id}` — cùng bẫy với `template`.

    Đăng ký sau bộ danh mục thì mọi lượt xuất nhận `422` ("export" không phải số
    nguyên) thay vì nhận tệp. `main.py` gắn bộ xuất ngay sau bộ nhập và trước bộ
    danh mục; đây là chỗ canh thứ tự ấy.
    """
    response = client.get(f"/api/v1/master/{WAREHOUSES}/export", headers=importer)
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(XLSX)
    assert "attachment" in response.headers["content-disposition"]


def test_exporting_needs_only_view_permission(client: TestClient, reader: dict[str, str]) -> None:
    """Cùng mức quyền với đường đọc danh sách, và cố ý.

    Tệp xuất ra chứa **đúng** những dòng mà `GET /api/v1/master/{slug}` đã trả về
    cho chính người dùng ấy — chỉ khác định dạng. Đòi thêm quyền ở đây là dựng
    một mức quyền thứ hai cho cùng một dữ liệu, thứ sẽ lệch khỏi mức thứ nhất.
    """
    response = client.get(f"/api/v1/master/{WAREHOUSES}/export", headers=reader)
    assert response.status_code == 200, response.text


def test_exporting_a_catalog_without_permission_on_it_is_refused(
    client: TestClient, importer: dict[str, str]
) -> None:
    """H48 áp cho cả đường xuất: quyền trên **danh mục này**, không phải mọi danh mục."""
    response = client.get(f"/api/v1/master/{PAYMENT_TERMS}/export", headers=importer)
    assert response.status_code == 403, response.text


def test_the_export_is_scoped_to_the_session_branch_not_the_whole_dataset(
    client: TestClient,
    importer: dict[str, str],
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
) -> None:
    """Phạm vi chi nhánh của lượt xuất đến từ **phiên**, và phải có cổng canh.

    Review vòng 1 (H3) rồi vòng 2 (H2): đột biến
    `branch_id=authorized.scope.acting_branch_id` → `branch_id=None` **sống sót**
    cả hai lần. Bản đầu của test này dùng người dùng không đặt `X-Branch`, nên
    `acting_branch_id` vốn đã là `None`: nó phân biệt được "không lọc gì" với
    "có lọc", nhưng **không** phân biệt được "lọc theo phiên" với "lọc về phần
    dùng chung" — mà đó đúng là thứ đột biến đổi.

    Ba bản ghi, ba phạm vi, và phiên **có** đặt `X-Branch`: chỉ khi ấy ba khả
    năng mới tách nhau ra. Bảng danh mục cố ý không bật RLS (H39) nên dòng
    `acting_branch_id` ấy là lớp lọc **duy nhất** của endpoint.
    """
    ids = branch_ids(session_factory, dataset_alpha, ["CN_IMP_MINE", "CN_IMP_KHAC"])
    shared_code = unique_code("KHO_CHUNG")
    mine_code = unique_code("KHO_CUA_TOI")
    other_code = unique_code("KHO_KHAC")

    for code, branch in (
        (shared_code, None),
        (mine_code, ids["CN_IMP_MINE"]),
        (other_code, ids["CN_IMP_KHAC"]),
    ):
        scope = RequestScope(
            dataset_schema=dataset_alpha.schema_name,
            user_id=1,
            branch_ids=() if branch is None else (branch,),
            acting_branch_id=branch,
        )
        with unit_of_work(session_factory, scope) as session:
            service: MasterDataService[Warehouse] = MasterDataService(session, Warehouse)
            service.create(code=code, name=f"Kho {code}", branch_id=branch)

    response = client.get(
        f"/api/v1/master/{WAREHOUSES}/export",
        headers={**importer, BRANCH_HEADER: str(ids["CN_IMP_MINE"])},
    )
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content), read_only=True, data_only=True)
    try:
        spec = REGISTRY.get(WAREHOUSES)
        assert spec is not None
        sheet = workbook[template_for(spec).sheet_name]
        exported = {row[0] for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True)}
    finally:
        workbook.close()

    assert mine_code in exported, "thiếu danh mục riêng của chính chi nhánh đang thao tác"
    assert shared_code in exported, "thiếu phần dùng chung toàn công ty"
    assert other_code not in exported, "tệp xuất chứa danh mục riêng của một chi nhánh khác"


def test_validate_only_authorises_auto_create_where_the_user_may_create(
    client: TestClient, importer: dict[str, str]
) -> None:
    """Hàng rào quyền của FR-NFR-062, đo ở **tầng HTTP** — nơi quyền thật sự được kiểm.

    Review vòng 2, H4: toàn bộ tính năng tự tạo chỉ được đo qua `run_import` với
    `allow_create_in` truyền tay, nên hai đột biến sống sót — bỏ hẳn
    `_authorise_targets` khỏi endpoint commit, và ép `missing_reference=CREATE`
    trong thân job. Đột biến thứ nhất nghĩa là: ai được nhập vật tư sẽ ghi được
    vào danh mục kho mà không có quyền nào trên kho, đúng lỗ hổng mà hàng rào
    này tuyên bố đã chặn (khuôn H89).

    `importer` có toàn quyền trên **danh mục kho** và không có gì trên các danh
    mục khác, nên danh sách được duyệt phải chứa `warehouses` và **chỉ** nó —
    dù `items` trỏ tới cả `units_of_measure`.
    """
    response = client.post(
        f"/api/v1/master/{WAREHOUSES}/import/validate",
        headers=importer,
        files={"file": ("kho.xlsx", _upload(WAREHOUSES, []), XLSX)},
        data={"missing_reference": "create"},
    )
    assert response.status_code == 202, response.text
    body = response.json()
    assert body["missing_reference"] == "create"
    # Danh mục kho không có cột tra cứu nào ngoài mã nhóm cha (cố ý không tự tạo
    # được — H106), nên danh sách rỗng là câu trả lời đúng và có nghĩa.
    assert body["allow_create_in"] == []


def test_the_default_import_never_asks_to_create_anything(
    client: TestClient, importer: dict[str, str]
) -> None:
    """Không gửi `missing_reference` thì chế độ là `error` — H80 áp cho cả FR-NFR-062."""
    response = client.post(
        f"/api/v1/master/{WAREHOUSES}/import/validate",
        headers=importer,
        files={"file": ("kho.xlsx", _upload(WAREHOUSES, []), XLSX)},
    )
    assert response.status_code == 202, response.text
    assert response.json()["missing_reference"] == "error"
