"""Số dư ban đầu qua HTTP (slice 4C) — chỉ những điều tầng HTTP mới trả lời được.

Luật nghiệp vụ nằm ở `test_opening_balances_import.py` /
`..._carry_forward.py`. Ở đây kiểm năm thứ:

* tệp mẫu tải về được với quyền `view`, đủ bốn sheet + Hướng dẫn;
* hai mức quyền: kiểm tệp cần `view`, ghi cần `create`;
* thiếu chi nhánh đang thao tác bị chặn NGAY ở endpoint (review 4C, L3) —
  không phải một `202` rồi job fail;
* hai job import **không** xếp thẳng qua `/api/v1/jobs`; job chuyển năm thì
  **được** (tham số tĩnh, quyền tĩnh) và vẫn đòi đúng quyền `create`;
* hai endpoint đọc trả đúng dữ liệu đã ghi (trang + tổng cân đối, hóa đơn
  của một dòng).
"""

from __future__ import annotations

from collections.abc import Iterator
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import Engine
from sqlalchemy.orm import Session, sessionmaker

from catalog_api_support import (
    UserFactory,
    actor,
    all_branch_codes,
    ensure_role,
)
from conftest import api_test_client
from ket.api.dependencies import BRANCH_HEADER
from ket.kernel.datasets.provisioning import DatasetRef
from ket.kernel.excel.descriptors import INSTRUCTIONS_SHEET
from ket.kernel.jobs.builtin import JOB_CREATE, JOB_VIEW
from ket.kernel.jobs.models import Job
from ket.kernel.master_data.models.partner import Partner
from ket.kernel.master_data.service import MasterDataService
from ket.kernel.persistence.unit_of_work import RequestScope, unit_of_work
from ket.main import create_app
from ket.posting.opening_balances.carry_forward_job import CARRY_FORWARD_CODE
from ket.posting.opening_balances.import_job import (
    COMMIT_CODE,
    OPENING_VIEW,
    OPENING_WRITE,
    VALIDATE_CODE,
)
from ket.posting.opening_balances.template import ACCOUNT_SHEET, PAYABLE_SHEET, SHEETS
from ket.settings import Settings
from posting_support import PostingContext, posting_scope, seed_posting_context
from test_opening_balance_parsing import workbook_of
from test_opening_balances_import import run_opening_import

pytestmark = pytest.mark.db

XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def opening_dir(tmp_path_factory: pytest.TempPathFactory) -> Path:
    return tmp_path_factory.mktemp("opening-store")


@pytest.fixture(scope="module")
def opening_settings(test_settings: Settings, opening_dir: Path) -> Settings:
    return test_settings.model_copy(update={"attachments_dir": opening_dir})


@pytest.fixture
def client(
    opening_settings: Settings, app_engine: Engine, session_factory: sessionmaker[Session]
) -> Iterator[TestClient]:
    assert app_engine is not None and session_factory is not None
    with api_test_client(create_app(opening_settings)) as instance:
        yield instance


@pytest.fixture
def context(session_factory: sessionmaker[Session], dataset_alpha: DatasetRef) -> PostingContext:
    """Chi nhánh + năm tài chính TRƯỚC khi actor được gán chi nhánh — actor vì
    thế nhìn thấy đúng chi nhánh test này."""
    return seed_posting_context(session_factory, dataset_alpha)


@pytest.fixture(scope="module")
def writer_role(session_factory: sessionmaker[Session], dataset_alpha: DatasetRef) -> str:
    return ensure_role(
        session_factory,
        dataset_alpha,
        "ke_toan_so_du",
        [OPENING_VIEW, OPENING_WRITE, JOB_VIEW, JOB_CREATE],
    )


@pytest.fixture(scope="module")
def viewer_role(session_factory: sessionmaker[Session], dataset_alpha: DatasetRef) -> str:
    return ensure_role(
        session_factory,
        dataset_alpha,
        "ke_toan_xem_so_du",
        [OPENING_VIEW, JOB_VIEW, JOB_CREATE],
    )


def _actor_headers(
    client: TestClient,
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    user_factory: UserFactory,
    role: str,
    prefix: str,
    test_password: str,
) -> dict[str, str]:
    return actor(
        client,
        session_factory,
        dataset_alpha,
        user_factory,
        role,
        prefix,
        test_password,
        branch_codes=all_branch_codes(session_factory, dataset_alpha),
    )


@pytest.fixture
def writer(
    client: TestClient,
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    user_factory: UserFactory,
    writer_role: str,
    test_password: str,
    context: PostingContext,
) -> dict[str, str]:
    headers = _actor_headers(
        client, session_factory, dataset_alpha, user_factory, writer_role, "sodughi", test_password
    )
    return {**headers, BRANCH_HEADER: str(context.branch_id)}


@pytest.fixture
def viewer(
    client: TestClient,
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    user_factory: UserFactory,
    viewer_role: str,
    test_password: str,
    context: PostingContext,
) -> dict[str, str]:
    headers = _actor_headers(
        client, session_factory, dataset_alpha, user_factory, viewer_role, "soduxem", test_password
    )
    return {**headers, BRANCH_HEADER: str(context.branch_id)}


def _small_workbook() -> BytesIO:
    return workbook_of({ACCOUNT_SHEET: [["111", None, None, None, None, 1_000, None]]})


def test_template_downloads_with_view_permission_and_has_all_sheets(
    client: TestClient, viewer: dict[str, str]
) -> None:
    response = client.get("/api/v1/opening-balances/template", headers=viewer)
    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith(XLSX)
    workbook = load_workbook(BytesIO(response.content))
    assert INSTRUCTIONS_SHEET in workbook.sheetnames
    for descriptor in SHEETS.values():
        assert descriptor.sheet_name in workbook.sheetnames


def test_validate_needs_view_and_enqueues_a_job(
    client: TestClient,
    viewer: dict[str, str],
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    context: PostingContext,
) -> None:
    response = client.post(
        "/api/v1/opening-balances/import/validate",
        headers=viewer,
        data={"fiscal_year_id": str(context.fiscal_year_id), "ledger": "0"},
        files={"file": ("so-du.xlsx", _small_workbook().getvalue(), XLSX)},
    )
    assert response.status_code == 202, response.text
    body = response.json()
    assert body["fiscal_year_id"] == context.fiscal_year_id
    # Job mang chi nhánh đang thao tác nên phải đọc bằng scope thấy chi nhánh đó
    # (RLS hàng đợi fail-closed với scope rỗng — đó là một khẳng định phụ miễn phí).
    scope = RequestScope(
        dataset_schema=dataset_alpha.schema_name, user_id=1, branch_ids=(context.branch_id,)
    )
    with unit_of_work(session_factory, scope) as session:
        job = session.get(Job, body["job_id"])
        assert job is not None and job.type == VALIDATE_CODE
        assert job.branch_id == context.branch_id


def test_validate_without_acting_branch_is_rejected_upfront(
    client: TestClient, viewer: dict[str, str], context: PostingContext
) -> None:
    """Review 4C, L3: thiếu X-Branch phải là `422` ngay, không phải job fail sau."""
    headers = {key: value for key, value in viewer.items() if key != BRANCH_HEADER}
    response = client.post(
        "/api/v1/opening-balances/import/validate",
        headers=headers,
        data={"fiscal_year_id": str(context.fiscal_year_id), "ledger": "0"},
        files={"file": ("so-du.xlsx", _small_workbook().getvalue(), XLSX)},
    )
    assert response.status_code == 422, response.text
    assert response.json()["type"].endswith("opening.branch_required")


def test_commit_requires_create_permission(
    client: TestClient, viewer: dict[str, str], context: PostingContext
) -> None:
    response = client.post(
        "/api/v1/opening-balances/import/commit",
        headers=viewer,
        json={
            "validation_job_id": "018f0000-0000-7000-8000-000000000000",
            "fiscal_year_id": context.fiscal_year_id,
            "ledger": 0,
        },
    )
    assert response.status_code == 403, response.text


def test_import_jobs_cannot_be_enqueued_directly(
    client: TestClient, writer: dict[str, str]
) -> None:
    for code in (VALIDATE_CODE, COMMIT_CODE):
        response = client.post("/api/v1/jobs", headers=writer, json={"type": code})
        assert response.status_code == 422, response.text
        assert response.json()["type"].endswith("job.not_directly_enqueueable")


def test_carry_forward_enqueues_via_generic_jobs_endpoint(
    client: TestClient,
    writer: dict[str, str],
    viewer: dict[str, str],
    context: PostingContext,
) -> None:
    payload = {
        "type": CARRY_FORWARD_CODE,
        "params": {"from_fiscal_year_id": context.fiscal_year_id},
    }
    denied = client.post("/api/v1/jobs", headers=viewer, json=payload)
    assert denied.status_code == 403, denied.text
    accepted = client.post("/api/v1/jobs", headers=writer, json=payload)
    assert accepted.status_code == 202, accepted.text


def test_read_endpoints_return_written_rows_and_invoices(
    client: TestClient,
    viewer: dict[str, str],
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    context: PostingContext,
    opening_dir: Path,
) -> None:
    """`GET /opening-balances` + `GET /{id}/invoices` trên dữ liệu ghi thật."""
    vendor_code = f"NCC-API-{context.branch_code}"
    scope = posting_scope(dataset_alpha, context, user_id=1)
    with unit_of_work(session_factory, scope) as session:
        MasterDataService(session, Partner).create(
            code=vendor_code, name="NCC api", extra={"is_vendor": True}
        )
    source = workbook_of(
        {
            ACCOUNT_SHEET: [["111", None, None, None, None, 500_000, None]],
            PAYABLE_SHEET: [
                [
                    vendor_code,
                    "331",
                    None,
                    None,
                    "HD-API",
                    "20/11/2025",
                    None,
                    None,
                    None,
                    None,
                    500_000,
                ]
            ],
        }
    )
    run_opening_import(session_factory, dataset_alpha, context, opening_dir, source)

    # `branch_id` tường minh: actor thấy MỌI chi nhánh của dataset dùng chung,
    # nên không lọc thì tổng cân đối cộng cả dữ liệu của các test khác.
    listing = client.get(
        "/api/v1/opening-balances",
        headers=viewer,
        params={
            "fiscal_year_id": context.fiscal_year_id,
            "ledger": 0,
            "branch_id": context.branch_id,
        },
    )
    assert listing.status_code == 200, listing.text
    body = listing.json()
    assert body["balanced"] is True
    assert Decimal(body["total_debit"]) == Decimal("500000.00")
    payable_row = next(item for item in body["items"] if item["detail_kind"] == 3)
    assert payable_row["partner_code"] == vendor_code
    assert payable_row["account_code"] == "331"

    invoices = client.get(f"/api/v1/opening-balances/{payable_row['id']}/invoices", headers=viewer)
    assert invoices.status_code == 200, invoices.text
    (invoice,) = invoices.json()["items"]
    assert invoice["invoice_no"] == "HD-API"
    assert Decimal(invoice["amount"]) == Decimal("500000.00")
