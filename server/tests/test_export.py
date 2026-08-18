"""Xuất danh mục ra Excel và nhập lại (lát 3C-2, bước 16 — FR-SYS-014, FR-NFR-061).

Khẳng định trung tâm của cả tệp là **vòng khép kín**: xuất ra rồi nhập lại thì
không lỗi và không sinh bản ghi thứ hai. Nó đáng được đo trên **toàn bộ** registry
chứ không vài danh mục: bộ xuất dùng chung một `TemplateDescriptor` cho hai mươi
hình dạng bảng khác nhau, và đó đúng dạng điểm mù mà lát 3B-1 đã ghi tên ("mã
dùng chung cho N thực thể, test chỉ chạy vài hình dạng") — lát 3C-1 trả giá cho
nó bằng bốn danh mục không nhập được dòng nào.

Ba thứ khác được đo riêng vì vòng khép kín **không** bắt được chúng:

* **`false` phải sống sót ở đường tạo mới.** Round-trip trên cùng dữ liệu kế toán
  đi vào đường *sửa*, nơi ô trống nghĩa là "giữ nguyên" (H91) — nên một ô boolean
  bị bỏ trống vẫn cho ra kết quả đúng ở đó và sai khi tệp được nhập vào một dữ
  liệu kế toán khác.
* **Phạm vi chi nhánh.** Cần một chi nhánh **thứ hai** mới phân biệt được
  "lọc đúng" với "không lọc gì" — nợ L5 của lát 3C-1.
* **Trần.** Phải bằng đúng trần của bước nhập, nếu không hệ thống sinh ra tệp mà
  chính nó từ chối.
"""

from __future__ import annotations

import time
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook
from sqlalchemy import Engine, delete, select
from sqlalchemy.orm import Session, sessionmaker

from catalog_api_support import branch_ids, ensure_branches, unique_code
from import_support import import_source, minimal_row, spec_of, workbook_bytes
from ket.kernel.datasets.provisioning import DatasetRef
from ket.kernel.errors import ExportTooManyRowsError
from ket.kernel.excel import exporter
from ket.kernel.excel.descriptors import INSTRUCTIONS_SHEET, CellKind, template_for
from ket.kernel.excel.exporter import FALSE_CELL, TRUE_CELL, build_export
from ket.kernel.excel.reader import (
    MAX_UNCOMPRESSED_BYTES,
    ensure_structure,
    uncompressed_size,
)
from ket.kernel.excel.report import ImportMode
from ket.kernel.excel.sql import FALSE_WORDS, TRUE_WORDS
from ket.kernel.master_data.models.partner import Partner
from ket.kernel.master_data.models.payment_term import PaymentTerm
from ket.kernel.master_data.models.warehouse import Warehouse
from ket.kernel.master_data.registry import REGISTRY
from ket.kernel.master_data.service import MasterDataService
from ket.kernel.persistence.unit_of_work import RequestScope, unit_of_work

pytestmark = pytest.mark.db

WAREHOUSES = "warehouses"
ITEMS = "items"
UNITS = "units_of_measure"
PARTNERS = "partners"

BRANCH_ONE = "CN_EXPORT_1"
BRANCH_TWO = "CN_EXPORT_2"


@pytest.fixture
def scope(dataset_alpha: DatasetRef) -> RequestScope:
    return RequestScope(dataset_schema=dataset_alpha.schema_name, user_id=1, branch_ids=())


@pytest.fixture
def two_branches(
    session_factory: sessionmaker[Session], dataset_alpha: DatasetRef
) -> dict[str, int]:
    """**Hai** chi nhánh, không một.

    Nợ L5 của lát 3C-1: mọi test phạm vi chi nhánh đến giờ dựng đúng một chi
    nhánh, nên chúng phân biệt được "dùng chung" với "riêng" mà **không** phân
    biệt được "riêng của tôi" với "riêng của người khác". Một bộ lọc bị xóa hẳn
    vẫn để chúng xanh — và đó chính là cụm đột biến sống sót ở vòng review thứ hai.
    """
    ensure_branches(session_factory, dataset_alpha, [BRANCH_ONE, BRANCH_TWO])
    return branch_ids(session_factory, dataset_alpha, [BRANCH_ONE, BRANCH_TWO])


def _import(
    session_factory: sessionmaker[Session],
    dataset: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
    slug: str,
    rows: list[list[object]],
    *,
    mode: ImportMode = ImportMode.CREATE_ONLY,
    branch_id: int | None = None,
) -> None:
    report = import_source(
        session_factory,
        dataset,
        scope,
        tmp_path,
        slug,
        workbook_bytes(slug, rows),
        mode=mode,
        commit=True,
        branch_id=branch_id,
    )
    assert report.is_valid and report.committed, report.errors


def _export(
    session_factory: sessionmaker[Session],
    scope: RequestScope,
    slug: str,
    *,
    branch_id: int | None = None,
    include_inactive: bool = True,
) -> bytes:
    spec = spec_of(slug)
    with unit_of_work(session_factory, scope) as session:
        return build_export(
            session,
            spec,
            template_for(spec),
            branch_id=branch_id,
            include_inactive=include_inactive,
        )


def _data_rows(content: bytes, slug: str) -> list[tuple[object, ...]]:
    """Dòng dữ liệu của tệp xuất ra, đọc lại bằng chính openpyxl."""
    descriptor = template_for(spec_of(slug))
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook[descriptor.sheet_name]
        return [
            row
            for row in sheet.iter_rows(min_row=2, max_col=len(descriptor.columns), values_only=True)
            if any(value is not None for value in row)
        ]
    finally:
        workbook.close()


def _column_index(slug: str, field: str) -> int:
    descriptor = template_for(spec_of(slug))
    return next(index for index, item in enumerate(descriptor.columns) if item.field == field)


# --------------------------------------------------------- vòng khép kín


@pytest.mark.parametrize("slug", [spec.slug for spec in REGISTRY.specs()])
def test_export_then_import_again_is_clean_for_every_catalog(
    slug: str,
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
    app_engine: Engine,
) -> None:
    """Tiêu chí phase 3: *"xuất danh mục rồi import lại: 0 lỗi, 0 bản ghi trùng"*.

    Duyệt toàn bộ registry vì hai mươi danh mục có hai mươi hình dạng cột, và
    một cột kiểu lạ ở danh mục thứ mười một là thứ không ai nhớ thêm test cho.
    Danh mục thêm ở phase 5 hay phase 9 vào đây miễn phí.

    Chế độ `create_and_update`: tệp xuất ra chứa những mã **đã có**, nên
    `create_only` sẽ báo trùng mã cho từng dòng — đúng theo H80, và đó là lý do
    tiêu chí này đòi đường cập nhật phải tồn tại.
    """
    spec = spec_of(slug)
    code = unique_code(f"RT{slug[:4].upper()}")
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        slug,
        [minimal_row(spec, code, f"Bản ghi {slug}")],
    )

    content = _export(session_factory, scope, slug)
    report = import_source(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        slug,
        BytesIO(content),
        mode=ImportMode.CREATE_AND_UPDATE,
        commit=True,
    )

    assert report.is_valid, report.errors
    assert report.committed
    assert report.create_rows == 0, "tệp xuất ra chỉ chứa mã đã có — không được tạo thêm dòng nào"
    # `create_rows == 0` **một mình là một khẳng định rỗng** (review vòng 1, H1):
    # một bộ xuất ghi 0 dòng cũng thỏa nó, và đột biến `for row in rows` →
    # `for row in []` để cả hai mươi tham số xanh. Phải khẳng định tệp **có** dòng
    # ấy, và mọi ô của nó bằng đúng thứ đã nhập vào.
    exported = _data_rows(content, slug)
    (line,) = [item for item in exported if item[_column_index(slug, "code")] == code]
    assert line[_column_index(slug, "name")] == f"Bản ghi {slug}"
    # Mọi dòng của tệp phải đi vào đường **cập nhật** — không phải "đúng một
    # dòng": dataset dùng chung giữa các tệp test nên danh mục có thể đã mang
    # bản ghi của test khác, và một con số cứng ở đây chỉ đúng khi chạy riêng.
    # Đây mới là bất biến: số dòng xuất ra = số dòng được cập nhật, tạo mới 0.
    assert report.update_rows == len(exported)


def test_the_exported_file_passes_the_structure_check(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
) -> None:
    """Tệp xuất ra **là** một tệp mẫu đã điền: cùng sheet, cùng tiêu đề, cùng thứ tự.

    Kiểm thẳng bằng `ensure_structure` — chính hàm mà lượt nhập gọi (FR-SYS-082)
    — thay vì so hai danh sách chuỗi trong test: so tay là bản sao thứ ba của
    cùng một luật, và nó sẽ đồng ý với một tệp mà bước nhập từ chối.
    """
    spec = spec_of(WAREHOUSES)
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec, unique_code("KHO_ST"), "Kho cấu trúc")],
    )
    content = _export(session_factory, scope, WAREHOUSES)

    ensure_structure(BytesIO(content), template_for(spec))
    assert INSTRUCTIONS_SHEET in load_workbook(BytesIO(content), read_only=True).sheetnames


def test_reference_columns_export_codes_not_ids(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
) -> None:
    """H79: ô tra cứu trong tệp là **mã**. Một id xuất ra là một tệp không nhập lại được.

    Vật tư hàng hóa là ca đáng đo nhất — nó có hai cột tra cứu trỏ vào hai danh
    mục khác nhau (`base_unit_id`, `warehouse_id`), nên nó bắt được cả lỗi "nối
    nhầm bí danh" mà một danh mục chỉ có `parent_code` không bắt được.
    """
    unit_code = unique_code("DVT_REF")
    warehouse_code = unique_code("KHO_REF")
    item_code = unique_code("VT_REF")
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        UNITS,
        [minimal_row(spec_of(UNITS), unit_code, "Cái")],
    )
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec_of(WAREHOUSES), warehouse_code, "Kho tham chiếu")],
    )
    row = minimal_row(spec_of(ITEMS), item_code, "Hàng hóa tham chiếu")
    descriptor = template_for(spec_of(ITEMS))
    for index, column in enumerate(descriptor.columns):
        if column.field == "nature":
            row[index] = "goods"
        elif column.field == "base_unit_code":
            row[index] = unit_code
        elif column.field == "warehouse_code":
            row[index] = warehouse_code
    _import(session_factory, dataset_alpha, scope, tmp_path, ITEMS, [row])

    exported = _data_rows(_export(session_factory, scope, ITEMS), ITEMS)
    (line,) = [item for item in exported if item[_column_index(ITEMS, "code")] == item_code]
    assert line[_column_index(ITEMS, "base_unit_code")] == unit_code
    assert line[_column_index(ITEMS, "warehouse_code")] == warehouse_code


# ------------------------------------------------------- ô boolean `false`


def test_the_boolean_words_written_out_are_words_the_importer_accepts() -> None:
    """Hai hằng số của bộ xuất phải nằm trong hai bộ từ khóa của bộ nhập.

    Cổng rẻ nhất của tệp này, và nó canh đúng chỗ dễ trôi: ai đó đổi `FALSE_CELL`
    thành `"-"` cho gọn mắt sẽ không thấy gì hỏng cho tới khi một tệp xuất ra
    được nhập vào và mọi bản ghi ngừng theo dõi sống lại.
    """
    assert TRUE_CELL.lower() in TRUE_WORDS
    assert FALSE_CELL.lower() in FALSE_WORDS


def test_an_inactive_record_stays_inactive_when_imported_as_a_new_row(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
) -> None:
    """`false` xuất ra phải là một **từ**, không phải ô trống (đoạn "Ba" của exporter).

    Đo ở đường **tạo mới** chứ không đường sửa, và đó là cả điểm của test: ô
    trống ở đường sửa nghĩa là "giữ nguyên" (H91) nên nó cho ra kết quả đúng dù
    bộ xuất có ghi gì; ở đường tạo mới thì ô trống trên cột "Còn theo dõi" nghĩa
    là **còn theo dõi** (`sql.is_active_expression`). Tình huống thật: xuất danh
    mục ở một bản cài rồi nhập vào bản cài của chi nhánh mới.
    """
    spec = spec_of(WAREHOUSES)
    source_code = unique_code("KHO_OFF")
    row = minimal_row(spec, source_code, "Kho đã ngừng")
    row[_column_index(WAREHOUSES, "is_active")] = "không"
    _import(session_factory, dataset_alpha, scope, tmp_path, WAREHOUSES, [row])

    exported = _data_rows(_export(session_factory, scope, WAREHOUSES), WAREHOUSES)
    (line,) = [item for item in exported if item[_column_index(WAREHOUSES, "code")] == source_code]
    assert line[_column_index(WAREHOUSES, "is_active")] == FALSE_CELL

    # Cùng dòng ấy, mã mới → đi vào đường **tạo mới**.
    fresh_code = unique_code("KHO_OFF2")
    fresh = list(line)
    fresh[_column_index(WAREHOUSES, "code")] = fresh_code
    _import(session_factory, dataset_alpha, scope, tmp_path, WAREHOUSES, [list(fresh)])

    with unit_of_work(session_factory, scope) as session:
        created = session.scalar(select(Warehouse).where(Warehouse.code == fresh_code))
    assert created is not None
    assert not created.is_active, "bản ghi đã ngừng theo dõi sống lại sau một vòng xuất–nhập"


def test_inactive_rows_are_left_out_unless_asked_for(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
) -> None:
    spec = spec_of(WAREHOUSES)
    code = unique_code("KHO_HID")
    row = minimal_row(spec, code, "Kho ẩn")
    row[_column_index(WAREHOUSES, "is_active")] = "không"
    _import(session_factory, dataset_alpha, scope, tmp_path, WAREHOUSES, [row])

    default_export = _data_rows(
        _export(session_factory, scope, WAREHOUSES, include_inactive=False), WAREHOUSES
    )
    codes = {item[_column_index(WAREHOUSES, "code")] for item in default_export}
    assert code not in codes

    asked = _data_rows(
        _export(session_factory, scope, WAREHOUSES, include_inactive=True), WAREHOUSES
    )
    assert code in {item[_column_index(WAREHOUSES, "code")] for item in asked}


# ------------------------------------------------- phạm vi chi nhánh (nợ L5)


def test_export_shows_shared_rows_and_this_branch_but_not_the_other_branch(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
    two_branches: dict[str, int],
) -> None:
    """Ba bản ghi, ba phạm vi — và chỉ hai trong ba được xuất (FR-SYS-018).

    Chi nhánh **thứ hai** là thứ làm test này có giá trị: với một chi nhánh, một
    bộ lọc bị xóa hẳn vẫn cho kết quả đúng, vì mọi dòng đều hoặc dùng chung hoặc
    thuộc chính chi nhánh đang đứng. Đây là chỗ trả nợ L5 của lát 3C-1.
    """
    shared = unique_code("KHO_SHARE")
    mine = unique_code("KHO_MINE")
    theirs = unique_code("KHO_THEIRS")
    spec = spec_of(WAREHOUSES)
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec, shared, "Kho dùng chung")],
    )
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec, mine, "Kho của tôi")],
        branch_id=two_branches[BRANCH_ONE],
    )
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec, theirs, "Kho của chi nhánh kia")],
        branch_id=two_branches[BRANCH_TWO],
    )

    exported = _data_rows(
        _export(session_factory, scope, WAREHOUSES, branch_id=two_branches[BRANCH_ONE]),
        WAREHOUSES,
    )
    codes = {item[_column_index(WAREHOUSES, "code")] for item in exported}
    assert shared in codes, "phần dùng chung toàn công ty phải xuất được từ mọi chi nhánh"
    assert mine in codes
    assert theirs not in codes, "xuất được cả danh mục riêng của chi nhánh khác"


# ------------------------------------------------------------------ hình dạng


@pytest.mark.parametrize(
    "amount",
    [
        # Mức người dùng thật gõ: ghi ra dưới dạng **số**, cộng được bằng Excel.
        pytest.param(Decimal("5000000000.000000"), id="hạn-mức-thường"),
        # Kín cả 20 chữ số của `NUMERIC(20, 6)` — quá tầm của dấu phẩy động 64
        # bit, nên bản đầu của bộ xuất trả về `100000000000000` mà không kêu.
        pytest.param(Decimal("99999999999999.999999"), id="kín-cả-cột"),
    ],
)
def test_a_decimal_survives_the_round_trip_at_full_column_precision(
    amount: Decimal,
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
) -> None:
    """Con số phải quay về **nguyên vẹn**, kể cả ở mức cột chứa được tối đa.

    Định dạng xlsx lưu mọi số dưới dạng dấu phẩy động, nên `NUMERIC(20, 6)` của
    `partners.credit_limit` rộng hơn thứ một ô Excel giữ nổi. Đây là chỗ một hạn
    mức nợ đổi giá trị trong im lặng qua một vòng xuất–nhập, nên nó được đo bằng
    hai giá trị: mức thường (phải ra **số**) và mức kín cả cột (phải vẫn đúng).
    """
    spec = spec_of(PARTNERS)
    descriptor = template_for(spec)
    index = _column_index(PARTNERS, "credit_limit")
    assert descriptor.columns[index].kind is CellKind.DECIMAL

    code = unique_code("DT_NUM")
    row = minimal_row(spec, code, "Đối tác hạn mức")
    row[index] = str(amount)
    _import(session_factory, dataset_alpha, scope, tmp_path, PARTNERS, [row])

    exported = _data_rows(_export(session_factory, scope, PARTNERS), PARTNERS)
    (line,) = [item for item in exported if item[_column_index(PARTNERS, "code")] == code]
    assert Decimal(str(line[index])) == amount, "giá trị đã đổi ngay ở bước ghi ra tệp"

    report = import_source(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        PARTNERS,
        BytesIO(_export(session_factory, scope, PARTNERS)),
        mode=ImportMode.CREATE_AND_UPDATE,
        commit=True,
    )
    assert report.is_valid, report.errors
    with unit_of_work(session_factory, scope) as session:
        stored = session.scalar(select(Partner).where(Partner.code == code))
    assert stored is not None
    assert stored.credit_limit == amount


def test_a_catalog_larger_than_the_import_ceiling_is_refused(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Trần của bộ xuất **bằng** trần của bộ nhập, nếu không hệ thống tự mâu thuẫn.

    Hạ trần thay vì dựng 100.000 bản ghi: thứ đang được đo là "có một trần và nó
    chặn", không phải con số cụ thể — và con số cụ thể đã được `reader.MAX_ROWS`
    khẳng định ở chỗ khác.
    """
    spec = spec_of(WAREHOUSES)
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec, unique_code("KHO_CAP"), "Kho chạm trần")],
    )
    monkeypatch.setattr(exporter, "MAX_EXPORT_ROWS", 0)
    with pytest.raises(ExportTooManyRowsError):
        _export(session_factory, scope, WAREHOUSES)


def test_a_file_that_would_be_too_big_for_the_importer_is_refused(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """Bất biến "xuất ra thì nhập lại được" phải đo **byte**, không chỉ đếm dòng.

    Review vòng 2, H1: cổng cũ ghim một hằng số bịa (150 byte/dòng). Đo thật trên
    `partners` với ô tiếng Việt đầy: **9.830 byte/dòng**, tức 10.000 dòng ra
    **93,7 MiB** — vượt `MAX_UNCOMPRESSED_BYTES` = 64 MiB của bộ nhập. Chữ tiếng
    Việt tốn 3 byte UTF-8 mỗi ký tự, nên số byte mỗi dòng đổi theo **nội dung**,
    và mọi hằng số đoán trước đều sai ở một bảng chữ cái nào đó.

    Nay bộ xuất đo chính tệp nó vừa dựng. Test hạ trần byte thay vì dựng một tệp
    94 MiB: thứ được đo là "có một phép kiểm và nó chặn", còn con số 64 MiB đã
    được `reader` khẳng định ở chỗ khác.
    """
    spec = spec_of(WAREHOUSES)
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec, unique_code("KHO_BYTE"), "Kho đo byte")],
    )
    monkeypatch.setattr(exporter, "MAX_UNCOMPRESSED_BYTES", 1)
    with pytest.raises(ExportTooManyRowsError):
        _export(session_factory, scope, WAREHOUSES)


def test_a_text_heavy_catalog_at_the_row_ceiling_is_refused_by_the_byte_guard(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
) -> None:
    """Danh mục nặng ký tự ở mức trần dòng **vượt** trần byte — và phải bị chặn.

    Đo thật (review vòng 2 + cổng này): `partners` 22 cột với ô tiếng Việt đầy
    cho ra ~7.550 byte/dòng, tức ~72 MiB ở 10.000 dòng — vượt
    `MAX_UNCOMPRESSED_BYTES` = 64 MiB của bộ nhập. Chữ tiếng Việt tốn 3 byte
    UTF-8 mỗi ký tự, nên số byte mỗi dòng đổi theo **nội dung**: không hằng số
    "byte mỗi dòng" nào đoán trước được, và bản đầu của cổng này ghim đúng một
    con số bịa (150).

    Quyết định (người dùng chốt 2026-08-18): giữ trần 10.000 dòng và để **phép
    đo lúc chạy** giữ bất biến. Phần lớn danh mục có tên ngắn nên 10.000 chạy
    tốt; ca nặng ký tự nhận một câu từ chối nói rõ việc phải làm, thay vì một tệp
    mà chính hệ thống từ chối nhập lại.

    Ngoại suy từ mẫu nhỏ rồi hạ trần byte để kiểm chính đường từ chối ấy — dựng
    đủ 10.000 dòng `partners` mất hàng chục giây và làm bẩn dataset dùng chung.
    """
    widest = max(
        REGISTRY.specs(),
        key=lambda item: sum(column.max_length or 0 for column in template_for(item).columns),
    )
    descriptor = template_for(widest)
    rows = []
    for index in range(20):
        row = minimal_row(widest, unique_code(f"BYTE{index}"), "Đối tượng đo dung lượng")
        for position, column in enumerate(descriptor.columns):
            # Ô chữ đầy tới trần bằng ký tự tiếng Việt: ca **tốn byte nhất** mà
            # một tệp khách hàng thật có thể tạo ra.
            if column.kind is CellKind.TEXT and column.max_length and row[position] is None:
                row[position] = "ữ" * min(column.max_length, 255)
        rows.append(row)
    _import(session_factory, dataset_alpha, scope, tmp_path, widest.slug, rows)

    content = _export(session_factory, scope, widest.slug)
    measured = uncompressed_size(BytesIO(content))
    assert measured is not None
    per_row = measured / max(len(_data_rows(content, widest.slug)), 1)
    projected = per_row * exporter.MAX_EXPORT_ROWS

    if projected <= MAX_UNCOMPRESSED_BYTES:
        pytest.fail(
            f"danh mục {widest.slug!r} nay vừa trần byte ở {exporter.MAX_EXPORT_ROWS} dòng "
            f"(~{projected / 1024 / 1024:.1f} MiB) — phép đo đã đổi, hãy đọc lại quyết định "
            f"giữ trần 10.000 dòng"
        )
    # Ca ấy **phải** bị chặn, và bị chặn bằng một lỗi nghiệp vụ đọc được chứ
    # không bằng một tệp hỏng ở lần nhập sau.
    monkey = min(measured - 1, MAX_UNCOMPRESSED_BYTES)
    with pytest.MonkeyPatch.context() as patch:
        patch.setattr(exporter, "MAX_UNCOMPRESSED_BYTES", monkey)
        with pytest.raises(ExportTooManyRowsError):
            _export(session_factory, scope, widest.slug)


def test_a_text_cell_starting_with_an_equals_sign_is_not_written_as_a_formula(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
) -> None:
    """Review vòng 1, C2 — hai lỗi trong một, và cả hai đều im lặng.

    openpyxl suy kiểu ô từ nội dung, nên một cái tên bắt đầu bằng `=` thành ô
    **công thức**. Hậu quả một: dữ liệu do một người dùng nhập vào trở thành lệnh
    chạy trên máy người mở tệp (tệp danh mục thì được gửi qua email suốt). Hậu
    quả hai: `reader.py` đọc với `data_only=True`, tức lấy *kết quả* của công
    thức — mà tệp ta ghi ra không có kết quả nào tính sẵn, nên ô đọc lại thành
    `None` và vòng xuất–nhập xóa trắng đúng ô đó.
    """
    code = unique_code("KHO_FX")
    hostile = "=cmd|'/c calc'!A0"
    # Dựng bản ghi thẳng qua service, **không** qua một lượt nhập: tệp nhập do
    # test tự ghi cũng đi qua openpyxl, nên chính nó biến ô ấy thành công thức và
    # `reader` (đọc `data_only=True`) nhận `None` — tức lượt nhập đổ trước khi
    # test chạm tới thứ nó muốn đo. Cái bẫy ở phía tệp **người dùng** soạn là một
    # sự thật riêng, không phải thứ khẳng định này nói về.
    with unit_of_work(session_factory, scope) as session:
        service: MasterDataService[Warehouse] = MasterDataService(session, Warehouse)
        service.create(code=code, name=hostile)

    content = _export(session_factory, scope, WAREHOUSES)
    with ZipFile(BytesIO(content)) as archive:
        sheets = [name for name in archive.namelist() if name.startswith("xl/worksheets/")]
        xml = "".join(archive.read(name).decode() for name in sheets)
    assert "<f>" not in xml, "ô chuỗi đã được ghi thành công thức Excel"

    exported = _data_rows(content, WAREHOUSES)
    (line,) = [item for item in exported if item[_column_index(WAREHOUSES, "code")] == code]
    assert line[_column_index(WAREHOUSES, "name")] == hostile, "giá trị mất khi đọc lại"


def test_a_shared_record_cannot_be_updated_from_a_branch(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
    two_branches: dict[str, int],
) -> None:
    """Review vòng 1, C1: vòng xuất–nhập ở chi nhánh **nhân đôi** mã dùng chung.

    Bộ xuất lọc theo phạm vi đọc (gồm dòng dùng chung), bộ nhập so khớp theo
    phạm vi ghi (đúng chi nhánh — H86), nên mã dùng chung không khớp gì và được
    tạo mới thành bản ghi riêng của chi nhánh. Không một dòng lỗi nào, và từ đó
    hai bản ghi cùng mã trôi vào mọi báo cáo.

    Nay nó là một dòng lỗi ở **chế độ cập nhật** — nơi ý định đọc được từ tệp là
    "sửa thứ đã có". Chế độ chỉ-tạo-mới giữ nguyên hành vi của H86; test ngay
    dưới canh điều đó.
    """
    shared_code = unique_code("KHO_CHUNG")
    spec = spec_of(WAREHOUSES)
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec, shared_code, "Kho dùng chung")],
    )
    branch = two_branches[BRANCH_ONE]
    content = _export(session_factory, scope, WAREHOUSES, branch_id=branch)

    report = import_source(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        BytesIO(content),
        mode=ImportMode.CREATE_AND_UPDATE,
        commit=True,
        branch_id=branch,
    )
    assert not report.is_valid
    assert not report.committed, "không được tạo bản sao riêng chi nhánh của mã dùng chung"
    codes = {error.code for error in report.errors}
    assert codes == {"import.shared_record_not_editable_from_branch"}

    with unit_of_work(session_factory, scope) as session:
        rows = session.scalars(select(Warehouse).where(Warehouse.code == shared_code)).all()
    assert len(rows) == 1, "mã dùng chung đã bị nhân đôi thành bản ghi riêng chi nhánh"


def test_a_branch_may_still_create_its_own_record_beside_a_shared_code(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
    two_branches: dict[str, int],
) -> None:
    """Đối chứng cho test trên: H86 **không** bị đảo.

    FR-SYS-018 cho phép chi nhánh khai bản ghi riêng mang cùng mã với một mã dùng
    chung, và `POST /api/v1/master/{slug}` vẫn làm được. Phép kiểm mới chỉ áp cho
    chế độ cập nhật, nên đường ấy còn nguyên — nếu không, bản sửa C1 sẽ lặng lẽ
    gỡ mất một quyền mà người dùng đang dùng.
    """
    shared_code = unique_code("KHO_RIENG")
    spec = spec_of(WAREHOUSES)
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec, shared_code, "Kho dùng chung")],
    )
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        WAREHOUSES,
        [minimal_row(spec, shared_code, "Kho riêng của chi nhánh")],
        branch_id=two_branches[BRANCH_ONE],
    )
    with unit_of_work(session_factory, scope) as session:
        rows = session.scalars(select(Warehouse).where(Warehouse.code == shared_code)).all()
    assert {row.branch_id for row in rows} == {None, two_branches[BRANCH_ONE]}


def test_exporting_a_full_catalog_stays_inside_a_request(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
) -> None:
    """Ngưỡng cứng cho quyết định "xuất chạy đồng bộ" (FR-NFR-043, FR-NFR-044).

    Quyết định ấy chỉ đúng chừng nào một danh mục đầy còn xuất xong trong thời
    gian của một request. Review vòng 1 đo được 100.000 dòng × 15 cột tốn 7,9
    giây CPU — đó là lý do trần hạ xuống `MAX_EXPORT_ROWS`. Cổng này giữ con số
    ấy khỏi trôi lại: nới trần lên mà quên đo thì nó đỏ, chứ không phải người
    dùng đầu tiên phát hiện bằng một request treo.

    Ngưỡng đặt rộng (30s) vì máy CI chậm hơn máy dev nhiều lần; thứ nó chặn là
    một bậc độ lớn, không phải một dao động.
    """
    spec = spec_of(WAREHOUSES)
    # Một tiền tố chung cho cả lượt, để dọn được bằng **một** câu lệnh sau đó.
    prefix = unique_code("KHO_PERF")
    rows = [
        minimal_row(spec, f"{prefix}_{index}", f"Kho hiệu năng {index}")
        for index in range(exporter.MAX_EXPORT_ROWS // 10)
    ]
    try:
        _import(session_factory, dataset_alpha, scope, tmp_path, WAREHOUSES, rows)

        started = time.monotonic()
        content = _export(session_factory, scope, WAREHOUSES)
        elapsed = time.monotonic() - started

        assert len(_data_rows(content, WAREHOUSES)) >= len(rows)
        assert elapsed < 30, f"xuất {len(rows)} dòng mất {elapsed:.1f}s — quá lâu cho một request"
    finally:
        # Dọn **bắt buộc**: dataset dùng chung giữa mọi tệp test, và một nghìn
        # bản ghi kho ở lại sẽ đẩy bản ghi của test khác rơi khỏi trang đầu của
        # đường đọc có phân trang — test ấy đỏ vì một lý do không liên quan gì
        # tới nó, và người đọc mất cả buổi để lần ra.
        with unit_of_work(session_factory, scope) as session:
            session.execute(delete(Warehouse).where(Warehouse.code.like(f"{prefix}_%")))


def test_a_lookup_code_that_now_points_elsewhere_is_refused(
    session_factory: sessionmaker[Session],
    dataset_alpha: DatasetRef,
    scope: RequestScope,
    tmp_path: Path,
    two_branches: dict[str, int],
) -> None:
    """Nợ H4 vòng 1: khóa ngoại đổi **im lặng** qua một vòng xuất–nhập.

    Dựng đúng probe của review: điều khoản `PT` dùng chung (id 1) ← đối tác của
    chi nhánh trỏ vào; rồi chi nhánh khai `PT` **riêng** cùng mã (id 2 — hợp lệ
    theo H86). Xuất đối tác ở chi nhánh rồi nhập lại nguyên xi: mã trong ô không
    đổi một ký tự, nhưng nó nay tra ra bản ghi của chi nhánh.

    Trước bản sửa: `payment_term_id` 1 → 2, `is_valid=True`, không cảnh báo —
    hạn nợ của mọi hóa đơn sau đó tính theo một điều khoản khác.
    """
    branch = two_branches[BRANCH_ONE]
    term_code = unique_code("DK_H4")
    partner_code = unique_code("DT_H4")
    term_spec = spec_of("payment_terms")
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        "payment_terms",
        [minimal_row(term_spec, term_code, "Điều khoản dùng chung")],
    )
    partner_spec = spec_of(PARTNERS)
    row = minimal_row(partner_spec, partner_code, "Đối tác chi nhánh")
    row[_column_index(PARTNERS, "payment_term_code")] = term_code
    _import(session_factory, dataset_alpha, scope, tmp_path, PARTNERS, [row], branch_id=branch)
    # Chi nhánh khai điều khoản **riêng** trùng mã — H86 cho phép, và đó là điều
    # kiện làm cho cùng một mã tra ra hai bản ghi khác nhau.
    _import(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        "payment_terms",
        [minimal_row(term_spec, term_code, "Điều khoản riêng chi nhánh")],
        branch_id=branch,
    )

    content = _export(session_factory, scope, PARTNERS, branch_id=branch)
    report = import_source(
        session_factory,
        dataset_alpha,
        scope,
        tmp_path,
        PARTNERS,
        BytesIO(content),
        mode=ImportMode.CREATE_AND_UPDATE,
        commit=True,
        branch_id=branch,
    )
    moved = [error for error in report.errors if error.code == "import.reference_moved"]
    assert moved, f"khóa ngoại đổi mà không ai kêu: {report.errors}"
    assert not report.committed

    with unit_of_work(session_factory, scope) as session:
        stored = session.scalar(select(Partner).where(Partner.code == partner_code))
    assert stored is not None
    with unit_of_work(session_factory, scope) as session:
        shared_term = session.scalar(
            select(PaymentTerm).where(
                PaymentTerm.code == term_code, PaymentTerm.branch_id.is_(None)
            )
        )
    assert shared_term is not None
    assert stored.payment_term_id == shared_term.id, "điều khoản của đối tác đã bị đổi"
