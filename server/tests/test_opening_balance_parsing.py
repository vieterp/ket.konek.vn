"""Đọc và kiểm hình thức workbook số dư ban đầu (4C) — nhóm không cần DB.

Hai hợp đồng được đo:

* **Tệp mẫu tải về chính là tệp được chấp nhận** (FR-SYS-082 áp cho số dư):
  workbook do `build_opening_template` sinh ra phải qua được chính
  `ensure_structure` mà lượt nhập dùng.
* **Luật hình thức từng dòng** (FR-OPB-005/006, BR-OPB-05): một-bên-Nợ-hoặc-Có,
  quy đổi khớp nguyên tệ × tỷ giá, ngày chứng từ trước đầu năm, bên ứng trước
  không mang số hóa đơn.
"""

from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import Workbook

from ket.kernel.errors import ImportSheetMissingError
from ket.kernel.excel.descriptors import TemplateDescriptor
from ket.kernel.excel.reader import ensure_structure
from ket.posting.opening_balances.parsing import ParsedOpeningRow, parse_workbook
from ket.posting.opening_balances.report import OpeningImportReport
from ket.posting.opening_balances.template import (
    ACCOUNT_SHEET,
    EMPLOYEE_ADVANCE_SHEET,
    PAYABLE_SHEET,
    RECEIVABLE_SHEET,
    SHEETS,
    build_opening_template,
)

YEAR_START = date(2026, 1, 1)


def workbook_of(sheets: dict[TemplateDescriptor, list[list[object]]]) -> BytesIO:
    """Workbook chỉ gồm các sheet được nêu — sheet vắng là một phần của kịch bản."""
    workbook = Workbook()
    default_sheet = workbook.active
    assert default_sheet is not None
    workbook.remove(default_sheet)
    for descriptor, rows in sheets.items():
        sheet = workbook.create_sheet(descriptor.sheet_name)
        sheet.append(list(descriptor.headers))
        for row in rows:
            sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def parse(
    sheets: dict[TemplateDescriptor, list[list[object]]],
) -> tuple[list[ParsedOpeningRow], set[tuple[str, int]], OpeningImportReport]:
    report = OpeningImportReport(
        fiscal_year_id=1, ledger=0, file_name="test.xlsx", content_hash="0" * 64
    )
    rows, failed = parse_workbook(
        workbook_of(sheets),
        base_currency="VND",
        money_scale=2,
        fiscal_year_start=YEAR_START,
        report=report,
    )
    return rows, failed, report


def error_codes(report: OpeningImportReport) -> set[str]:
    return {error.code for error in report.errors}


def test_template_round_trips_through_structure_check() -> None:
    content = build_opening_template()
    for descriptor in SHEETS.values():
        ensure_structure(BytesIO(content), descriptor)


def test_base_currency_row_normalises_rate_and_fc() -> None:
    rows, failed, report = parse({ACCOUNT_SHEET: [["111", None, None, None, None, 500_000, None]]})
    assert not failed and report.is_valid
    (row,) = rows
    assert row.exchange_rate == Decimal(1)
    assert row.debit == Decimal(500_000)
    # Quy ước posting engine: dòng đồng tiền hạch toán có nguyên tệ = quy đổi.
    assert row.debit_fc == Decimal(500_000)


def test_foreign_row_computes_converted_amount_when_blank() -> None:
    rows, failed, report = parse({ACCOUNT_SHEET: [["112", "USD", 25_000, 100, None, None, None]]})
    assert not failed and report.is_valid
    (row,) = rows
    assert row.debit == Decimal("2500000.00")
    assert row.debit_fc == Decimal(100)


def test_foreign_row_rejects_mismatched_conversion() -> None:
    _, failed, report = parse({ACCOUNT_SHEET: [["112", "USD", 25_000, 100, None, 2_600_000, None]]})
    assert failed
    assert "opening.conversion_mismatch" in error_codes(report)


def test_foreign_row_requires_positive_rate() -> None:
    _, _, report = parse({ACCOUNT_SHEET: [["112", "USD", None, 100, None, None, None]]})
    assert "opening.rate_required" in error_codes(report)


def test_base_currency_row_rejects_foreign_rate() -> None:
    _, _, report = parse({ACCOUNT_SHEET: [["111", "VND", 25_000, None, None, 100, None]]})
    assert "opening.rate_for_base_currency" in error_codes(report)


def test_row_with_both_sides_is_rejected() -> None:
    _, _, report = parse({ACCOUNT_SHEET: [["111", None, None, None, None, 100, 200]]})
    assert "opening.both_sides" in error_codes(report)


def test_row_without_amount_is_rejected() -> None:
    _, _, report = parse({ACCOUNT_SHEET: [["111", None, None, None, None, None, None]]})
    assert "opening.no_amount" in error_codes(report)


def test_negative_amount_is_rejected() -> None:
    _, _, report = parse({ACCOUNT_SHEET: [["111", None, None, None, None, -5, None]]})
    assert "opening.negative_amount" in error_codes(report)


def test_missing_account_code_is_rejected() -> None:
    _, _, report = parse({ACCOUNT_SHEET: [[None, None, None, None, None, 100, None]]})
    assert "opening.required" in error_codes(report)


def test_text_that_is_not_a_number_is_rejected() -> None:
    _, _, report = parse({ACCOUNT_SHEET: [["111", None, None, None, None, "abc", None]]})
    assert "opening.not_a_number" in error_codes(report)


def test_receivable_natural_side_requires_invoice_no_and_date() -> None:
    _, _, report = parse(
        {RECEIVABLE_SHEET: [["KH01", "131", None, None, None, None, None, None, None, 1_000, None]]}
    )
    codes = error_codes(report)
    assert "opening.invoice_no_required" in codes
    assert "opening.invoice_date_required" in codes


def test_invoice_date_must_be_before_fiscal_year_start() -> None:
    _, _, report = parse(
        {
            RECEIVABLE_SHEET: [
                ["KH01", "131", None, None, "HD1", "15/01/2026", None, None, None, 1_000, None]
            ]
        }
    )
    assert "opening.invoice_date_in_year" in error_codes(report)


@pytest.mark.parametrize("raw", ["2025-12-31", "31/12/2025", "31-12-2025"])
def test_invoice_date_formats_are_accepted(raw: str) -> None:
    rows, failed, report = parse(
        {RECEIVABLE_SHEET: [["KH01", "131", None, None, "HD1", raw, None, None, None, 1_000, None]]}
    )
    assert not failed, [error.model_dump() for error in report.errors]
    assert rows[0].invoice_date == date(2025, 12, 31)


def test_unreadable_date_is_rejected() -> None:
    _, _, report = parse(
        {
            RECEIVABLE_SHEET: [
                ["KH01", "131", None, None, "HD1", "31.12.2025", None, None, None, 1_000, None]
            ]
        }
    )
    assert "opening.invalid_date" in error_codes(report)


def test_counter_side_must_not_carry_invoice_fields() -> None:
    """Khoản khách ứng trước (dư Có 131) không mang số hóa đơn."""
    _, _, report = parse(
        {
            RECEIVABLE_SHEET: [
                ["KH01", "131", None, None, "HD1", "01/12/2025", None, None, None, None, 700]
            ]
        }
    )
    assert "opening.invoice_on_counter_side" in error_codes(report)


def test_payable_natural_side_is_credit() -> None:
    """Bên "còn nợ" của sheet phải trả là bên Có — dòng Có mang chi tiết hóa đơn."""
    rows, failed, _ = parse(
        {
            PAYABLE_SHEET: [
                ["NCC01", "331", None, None, "HD9", "20/11/2025", None, None, None, None, 2_000]
            ]
        }
    )
    assert not failed
    assert rows[0].is_natural_side


def test_employee_advance_requires_document_fields() -> None:
    _, _, report = parse(
        {EMPLOYEE_ADVANCE_SHEET: [["NV01", "141", None, None, None, None, None, None, 300, None]]}
    )
    assert "opening.invoice_no_required" in error_codes(report)


def test_absent_sheets_are_skipped() -> None:
    rows, failed, report = parse({ACCOUNT_SHEET: [["111", None, None, None, None, 100, None]]})
    assert not failed
    assert report.total_rows == 1
    assert {row.kind for row in rows} == {0}


def test_workbook_without_any_opening_sheet_is_rejected() -> None:
    workbook = Workbook()
    sheet = workbook.active
    assert sheet is not None
    sheet.title = "Sheet1"
    sheet.append(["A"])
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    report = OpeningImportReport(
        fiscal_year_id=1, ledger=0, file_name="test.xlsx", content_hash="0" * 64
    )
    with pytest.raises(ImportSheetMissingError):
        parse_workbook(
            buffer,
            base_currency="VND",
            money_scale=2,
            fiscal_year_start=YEAR_START,
            report=report,
        )


def test_text_longer_than_column_limit_is_rejected_at_validate() -> None:
    """Review 4C, H1: lượt kiểm phải từ chối thứ mà cột `varchar` sẽ làm crash lượt ghi."""
    _, _, report = parse(
        {
            RECEIVABLE_SHEET: [
                [
                    "KH01",
                    "131",
                    "USDT",  # quá varchar(3)
                    25_000,
                    "X" * 63,  # quá varchar(50)
                    "01/12/2025",
                    None,
                    100,
                    None,
                    2_500_000,
                    None,
                ]
            ]
        }
    )
    codes = error_codes(report)
    assert "opening.too_long" in codes
    too_long_columns = {error.column for error in report.errors if error.code == "opening.too_long"}
    assert {"Loại tiền", "Số hóa đơn/chứng từ"} <= too_long_columns


def test_more_decimals_than_money_scale_is_rejected() -> None:
    """Review 4C, H2: DB sẽ tròn `100.999` thành `101.00` im lặng — phải chặn từ lượt kiểm."""
    _, _, report = parse({ACCOUNT_SHEET: [["111", None, None, None, None, "100.999", None]]})
    assert "opening.too_many_decimals" in error_codes(report)


def test_base_currency_rate_written_as_one_point_zero_is_accepted() -> None:
    """Review 4C, L4: `1.0` là tỷ giá 1 — so bằng số, không bằng chuỗi."""
    rows, failed, _ = parse({ACCOUNT_SHEET: [["111", "VND", "1.0", None, None, 100, None]]})
    assert not failed
    assert rows[0].exchange_rate == Decimal(1)


def test_invoice_dated_exactly_on_year_start_is_rejected() -> None:
    """Biên `==` ngày đầu năm thuộc về năm mới, không phải số dư đầu kỳ (BR-OPB-05)."""
    _, _, report = parse(
        {
            RECEIVABLE_SHEET: [
                ["KH01", "131", None, None, "HD1", "01/01/2026", None, None, None, 1_000, None]
            ]
        }
    )
    assert "opening.invoice_date_in_year" in error_codes(report)
