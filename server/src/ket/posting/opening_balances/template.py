"""Tệp mẫu Excel của số dư ban đầu — bốn sheet cho nhóm 0–4 (FR-OPB-002).

Một workbook nhiều sheet chứ không bốn tệp: kế toán chuyển từ phần mềm cũ nhập
số dư **một lần cho cả năm**, và bốn tệp là bốn lần tải lên với bốn lượt kiểm
phải tự nhớ đã chạy đủ chưa. Sheet vắng mặt = nhóm đó không đụng tới trong lượt
nhập này (người dùng hay xóa sheet không dùng — cùng lý do `INSTRUCTIONS_SHEET`
được phép vắng ở khung danh mục).

Nhóm **ngân hàng (kind 1) chưa có sheet**: chi tiết "từng tài khoản ngân hàng"
đòi danh mục tài khoản ngân hàng của chính doanh nghiệp, mà danh mục ấy thuộc
module ngân hàng (phase 6 — `bank_vouchers.bank_account_id`). Số dư 112 nhập
tạm ở sheet "Số dư tài khoản"; phase 6 thêm sheet ngân hàng cạnh danh mục của
nó, đúng bài RT-24 đã tách nhóm 5–9 về phase 8.

Tái dùng `ColumnDescriptor`/`TemplateDescriptor` của khung danh mục: `reader`
kiểm cấu trúc và đọc dòng theo đúng descriptor này, nên "tệp mẫu tải về" và
"tệp được chấp nhận" là một hợp đồng (FR-SYS-082) — không phải hai bản chép.
"""

from __future__ import annotations

from io import BytesIO
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from ket.kernel.excel.descriptors import (
    INSTRUCTIONS_SHEET,
    CellKind,
    ColumnDescriptor,
    TemplateDescriptor,
)
from ket.kernel.excel.template import (
    HEADER_FILL,
    HEADER_WIDTH,
    OPTIONAL_FONT,
    REQUIRED_FONT,
)
from ket.posting.opening_balances.models import OpeningDetailKind

ACCOUNT_CODE_COLUMN: Final[ColumnDescriptor] = ColumnDescriptor(
    field="account_code",
    header="Số hiệu tài khoản",
    kind=CellKind.TEXT,
    required=True,
    max_length=20,
    note="Số hiệu trong hệ thống tài khoản của chế độ kế toán năm (TT200/TT133).",
)

_CURRENCY_COLUMNS: Final[tuple[ColumnDescriptor, ...]] = (
    ColumnDescriptor(
        field="currency_code",
        header="Loại tiền",
        kind=CellKind.TEXT,
        max_length=3,
        note="Để trống = đồng tiền hạch toán của năm (VND).",
    ),
    ColumnDescriptor(
        field="exchange_rate",
        header="Tỷ giá",
        kind=CellKind.DECIMAL,
        note="Bắt buộc khi có ngoại tệ; số quy đổi phải bằng nguyên tệ × tỷ giá.",
    ),
)

_AMOUNT_COLUMNS: Final[tuple[ColumnDescriptor, ...]] = (
    ColumnDescriptor(
        field="debit_fc",
        header="Dư Nợ nguyên tệ",
        kind=CellKind.DECIMAL,
        note="Chỉ điền khi dòng là ngoại tệ.",
    ),
    ColumnDescriptor(
        field="credit_fc",
        header="Dư Có nguyên tệ",
        kind=CellKind.DECIMAL,
        note="Chỉ điền khi dòng là ngoại tệ.",
    ),
    ColumnDescriptor(
        field="debit",
        header="Dư Nợ quy đổi",
        kind=CellKind.DECIMAL,
        note="Số tiền theo đồng tiền hạch toán. Mỗi dòng chỉ điền một bên Nợ hoặc Có.",
    ),
    ColumnDescriptor(
        field="credit",
        header="Dư Có quy đổi",
        kind=CellKind.DECIMAL,
        note="Số tiền theo đồng tiền hạch toán. Mỗi dòng chỉ điền một bên Nợ hoặc Có.",
    ),
)


def _invoice_columns(
    *, no_header: str, date_header: str, with_due_date: bool
) -> tuple[ColumnDescriptor, ...]:
    """Bộ cột chứng từ của các sheet công nợ (FR-OPB-003, BR-OPB-05)."""
    columns = [
        ColumnDescriptor(
            field="invoice_no",
            header=no_header,
            kind=CellKind.TEXT,
            max_length=50,
            note="Bắt buộc cho dòng còn nợ; dòng trả trước/ứng trước để trống.",
        ),
        ColumnDescriptor(
            field="invoice_date",
            header=date_header,
            kind=CellKind.TEXT,
            note="Định dạng ngày/tháng/năm hoặc năm-tháng-ngày; phải trước ngày đầu năm tài chính.",
        ),
    ]
    if with_due_date:
        columns.append(
            ColumnDescriptor(
                field="due_date",
                header="Hạn thanh toán",
                kind=CellKind.TEXT,
                note="Để trống nếu không theo dõi hạn.",
            )
        )
    return tuple(columns)


ACCOUNT_SHEET: Final[TemplateDescriptor] = TemplateDescriptor(
    slug="opening-account",
    sheet_name="Số dư tài khoản",
    columns=(ACCOUNT_CODE_COLUMN, *_CURRENCY_COLUMNS, *_AMOUNT_COLUMNS),
)

RECEIVABLE_SHEET: Final[TemplateDescriptor] = TemplateDescriptor(
    slug="opening-receivable",
    sheet_name="Phải thu khách hàng",
    columns=(
        ColumnDescriptor(
            field="partner_code",
            header="Mã khách hàng",
            kind=CellKind.TEXT,
            required=True,
            max_length=50,
            note="Mã trong danh mục đối tác, bản ghi có đánh dấu là khách hàng.",
        ),
        ACCOUNT_CODE_COLUMN,
        *_CURRENCY_COLUMNS,
        *_invoice_columns(
            no_header="Số hóa đơn/chứng từ", date_header="Ngày hóa đơn", with_due_date=True
        ),
        *_AMOUNT_COLUMNS,
    ),
)

PAYABLE_SHEET: Final[TemplateDescriptor] = TemplateDescriptor(
    slug="opening-payable",
    sheet_name="Phải trả nhà cung cấp",
    columns=(
        ColumnDescriptor(
            field="partner_code",
            header="Mã nhà cung cấp",
            kind=CellKind.TEXT,
            required=True,
            max_length=50,
            note="Mã trong danh mục đối tác, bản ghi có đánh dấu là nhà cung cấp.",
        ),
        ACCOUNT_CODE_COLUMN,
        *_CURRENCY_COLUMNS,
        *_invoice_columns(
            no_header="Số hóa đơn/chứng từ", date_header="Ngày hóa đơn", with_due_date=True
        ),
        *_AMOUNT_COLUMNS,
    ),
)

EMPLOYEE_ADVANCE_SHEET: Final[TemplateDescriptor] = TemplateDescriptor(
    slug="opening-employee-advance",
    sheet_name="Tạm ứng nhân viên",
    columns=(
        ColumnDescriptor(
            field="employee_code",
            header="Mã nhân viên",
            kind=CellKind.TEXT,
            required=True,
            max_length=50,
            note="Mã trong danh mục nhân viên.",
        ),
        ACCOUNT_CODE_COLUMN,
        *_CURRENCY_COLUMNS,
        *_invoice_columns(
            no_header="Số chứng từ tạm ứng", date_header="Ngày tạm ứng", with_due_date=False
        ),
        *_AMOUNT_COLUMNS,
    ),
)

SHEETS: Final[dict[int, TemplateDescriptor]] = {
    OpeningDetailKind.ACCOUNT: ACCOUNT_SHEET,
    OpeningDetailKind.RECEIVABLE: RECEIVABLE_SHEET,
    OpeningDetailKind.PAYABLE: PAYABLE_SHEET,
    OpeningDetailKind.EMPLOYEE_ADVANCE: EMPLOYEE_ADVANCE_SHEET,
}
"""Nhóm số dư → sheet của nó, theo đúng thứ tự nhóm trong SRS 02 §1.1."""

TEMPLATE_FILE_NAME: Final[str] = "mau-nhap-so-du-ban-dau.xlsx"


def _write_instructions(workbook: Workbook) -> None:
    """Sheet Hướng dẫn cho cả bốn nhóm — sinh từ chính các descriptor.

    Không dùng `template.write_instructions_sheet` vì hàm đó nhận `CatalogSpec`;
    ở đây không có danh mục nào — nhưng nội dung sinh cùng một cách: từ
    descriptor, nên hướng dẫn không thể mô tả một cột đã đổi.
    """
    sheet = workbook.create_sheet(INSTRUCTIONS_SHEET)
    sheet.append(["Tệp mẫu nhập số dư ban đầu"])
    sheet.append([])
    sheet.append(["Mỗi nhóm số dư một sheet; sheet không dùng có thể xóa."])
    sheet.append(["Không đổi tên sheet, không đổi tên cột, không đổi thứ tự cột."])
    sheet.append(["Cột có dấu * là cột bắt buộc nhập. Mỗi dòng chỉ điền một bên Nợ hoặc Có."])
    sheet.append(
        [
            "Số dư tiền gửi ngân hàng (TK 112) tạm nhập ở sheet Số dư tài khoản; "
            "chi tiết theo từng tài khoản ngân hàng sẽ có cùng phân hệ Ngân hàng."
        ]
    )
    for descriptor in SHEETS.values():
        sheet.append([])
        sheet.append([f"Sheet {descriptor.sheet_name!r}", ""])
        for column in descriptor.columns:
            note_parts = [
                "Bắt buộc." if column.required else "Có thể để trống.",
                column.note,
            ]
            sheet.append([column.display_header, " ".join(part for part in note_parts if part)])
    sheet.column_dimensions["A"].width = HEADER_WIDTH
    sheet.column_dimensions["B"].width = 90


def _write_header(workbook: Workbook, descriptor: TemplateDescriptor) -> None:
    """Một sheet dữ liệu: đúng một dòng tiêu đề, cùng trình bày với tệp mẫu danh mục."""
    sheet = workbook.create_sheet(descriptor.sheet_name)
    sheet.append(list(descriptor.headers))
    for index, column in enumerate(descriptor.columns, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = REQUIRED_FONT if column.required else OPTIONAL_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = HEADER_WIDTH
    sheet.freeze_panes = "A2"


def build_opening_template() -> bytes:
    """Workbook hoàn chỉnh: Hướng dẫn + bốn sheet dữ liệu, sẵn cho phản hồi HTTP."""
    workbook = Workbook(write_only=False)
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)
    _write_instructions(workbook)
    for descriptor in SHEETS.values():
        _write_header(workbook, descriptor)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
