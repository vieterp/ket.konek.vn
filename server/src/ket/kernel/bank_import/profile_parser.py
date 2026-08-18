"""Đọc một tệp sao kê theo hồ sơ định dạng của ngân hàng (RT-26, FR-BNK-032).

Đầu ra là `StatementLine` — dòng sao kê đã chuẩn hóa, **không** phụ thuộc nhà
băng nào nữa. Phase 6 nhận danh sách ấy và biến thành chứng từ thu/chi, đối
chiếu với sổ tiền gửi; nó không bao giờ phải biết Vietcombank viết ngày kiểu gì.

Hai loại lỗi, tách bạch có chủ đích:

* **Lỗi cấu trúc** → ném ngay, dừng cả lượt. Không tìm thấy cột `"Số tiền ghi
  có"` nghĩa là hồ sơ không khớp tệp; báo từng dòng một sẽ cho ra vài nghìn lỗi
  giống hệt nhau cho **một** việc phải sửa. Cùng lý do `reader.ensure_structure`
  dừng trước khi đọc dòng nào.
* **Lỗi từng dòng** → gom vào `issues`, đọc tiếp. Một ô ngày hỏng ở dòng 47 là
  chuyện của dòng 47, và người dùng cần thấy **tất cả** những dòng như vậy trong
  một lượt để sửa một lần.

**Quy ước dấu của `amount`: dương = tiền VÀO, âm = tiền RA.** Một quy ước, khai
ở đây, cho cả ba hình dạng nguồn (hai cột nợ/có, một cột có dấu, một cột chỉ
ghi chi). Không có nó thì mỗi nơi đọc phải tự hỏi lại, và chỗ trả lời sai sẽ là
chỗ đảo chiều một khoản tiền.
"""

from __future__ import annotations

import csv
from collections.abc import Iterator, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import TextIOWrapper
from itertools import islice
from numbers import Real
from typing import IO, Final

from openpyxl import load_workbook

from ket.kernel.bank_import.profile_models import (
    AmountSignRule,
    BankStatementProfile,
    StatementFileKind,
)
from ket.kernel.errors import (
    BankStatementColumnAmbiguousError,
    BankStatementColumnMissingError,
    BankStatementFileUnreadableError,
    BankStatementFormatUnsupportedError,
)
from ket.kernel.excel.reader import MAX_UNCOMPRESSED_BYTES, uncompressed_size

MAX_ROWS: Final[int] = 100_000
"""Trần số dòng dữ liệu của một tệp sao kê. Cùng bậc với `excel/reader.MAX_ROWS`
và vì cùng lý do: một tệp bất thường không được ăn hết bộ nhớ máy chủ."""

CSV_ENCODING: Final[str] = "utf-8-sig"
"""`utf-8-sig` chứ không `utf-8`: tệp CSV do Excel trên Windows xuất ra gần như
luôn mang BOM, và với `utf-8` thuần thì ba byte ấy dính vào **tiêu đề cột đầu
tiên** — cột "Ngày giao dịch" thành "\\ufeffNgày giao dịch" và phép tra cột đổ
với một thông báo trông như hồ sơ khai sai. Bộ giải mã này ăn BOM nếu có và
không đổi gì nếu không."""

_DEFAULT_CSV_DELIMITER: Final[str] = ","


@dataclass(frozen=True)
class StatementLine:
    """Một dòng sao kê đã chuẩn hóa."""

    row_number: int
    """Số dòng **như trong tệp gốc** — người dùng đối chiếu bằng mắt trên tệp đó."""

    booked_on: date
    amount: Decimal
    """Dương = tiền vào, âm = tiền ra. Xem docstring module."""

    reference: str | None
    description: str | None
    balance: Decimal | None


@dataclass(frozen=True)
class StatementIssue:
    """Một dòng không đọc được, kèm đủ thứ để người dùng tìm ra nó."""

    row_number: int
    column: str | None
    code: str
    message: str
    value: str | None = None


@dataclass(frozen=True)
class StatementParseResult:
    lines: tuple[StatementLine, ...]
    issues: tuple[StatementIssue, ...]

    @property
    def is_clean(self) -> bool:
        return not self.issues


def parse_statement(source: IO[bytes], profile: BankStatementProfile) -> StatementParseResult:
    """Đọc tệp sao kê theo hồ sơ, trả về dòng đã chuẩn hóa + danh sách dòng hỏng."""
    kind = StatementFileKind(profile.file_kind)
    if kind is StatementFileKind.MT940:
        raise BankStatementFormatUnsupportedError(
            "Bản cài này chưa đọc được sao kê dạng MT940", file_kind=kind.value
        )
    rows = _xlsx_rows(source) if kind is StatementFileKind.XLSX else _csv_rows(source, profile)
    return _parse_rows(rows, profile)


def _ensure_within_zip_budget(source: IO[bytes]) -> None:
    """Từ chối tệp mà **nội dung đã giải nén** vượt trần, trước khi mở nó.

    Bài học R2-4 của lát 3C-1, mang sang đây: `MAX_ROWS` chỉ chặn **sau** khi
    openpyxl đã tự tính vùng dữ liệu của sheet, mà chi phí nằm đúng ở bước ấy.
    Đo được trên một tệp .xlsx 27 KB nén / 17 MiB giải nén không khai
    `<dimension>`: **1,4 giây CPU** trước khi dòng đầu tiên được phát ra.

    Module Bank của phase 6 sẽ nhận tệp người dùng tải lên qua mạng, nên trần
    này phải có **trước** lúc đó — không phải sau lần đầu một worker bị treo.
    """
    total = uncompressed_size(source)
    if total is not None and total > MAX_UNCOMPRESSED_BYTES:
        raise BankStatementFileUnreadableError(
            "Tệp sao kê giải nén ra lớn hơn mức cho phép — hãy xuất lại kỳ ngắn hơn",
            reason="uncompressed_too_large",
        )


def _xlsx_rows(source: IO[bytes]) -> list[tuple[object, ...]]:
    """Mọi dòng của sheet **đầu tiên**, giữ nguyên kiểu ô.

    Sheet đầu chứ không tra theo tên: tên sheet của sao kê do ngân hàng đặt và
    đổi theo từng lần xuất ("Sheet1", "sao_ke_202608", tên tài khoản). Thêm một
    ô cấu hình cho nó là bắt người quản trị khai lại một giá trị đổi mỗi tháng.

    Giữ nguyên kiểu ô (không `str()` sớm) vì Excel lưu ngày dưới dạng **ngày
    thật**: ép nó qua chuỗi rồi phân tích lại bằng `date_format` là tự tạo ra
    một chỗ để sai, và nó sai đúng vào tệp đã hoàn toàn đúng.
    """
    _ensure_within_zip_budget(source)
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as error:
        raise BankStatementFileUnreadableError(
            "Không đọc được tệp sao kê — hãy lưu lại ở định dạng .xlsx rồi thử lại",
            reason=type(error).__name__,
        ) from error
    try:
        sheet = workbook[workbook.sheetnames[0]]
        return list(islice(sheet.iter_rows(values_only=True), MAX_ROWS))
    finally:
        workbook.close()


def _csv_rows(source: IO[bytes], profile: BankStatementProfile) -> list[tuple[object, ...]]:
    delimiter = profile.csv_delimiter or _DEFAULT_CSV_DELIMITER
    try:
        text = TextIOWrapper(source, encoding=CSV_ENCODING, newline="")
        try:
            reader = csv.reader(text, delimiter=delimiter)
            return [tuple(row) for row in islice(reader, MAX_ROWS)]
        finally:
            # `detach()` chứ không để wrapper tự thu gom: lúc bị thu gom nó
            # **đóng luôn** stream gốc, nên chỗ gọi ở phase 6 (băm nội dung rồi
            # đọc lại lần hai) sẽ gặp `ValueError: I/O operation on closed file`
            # ở một thời điểm không đoán trước được (R3-L3).
            text.detach()
    except (UnicodeDecodeError, csv.Error) as error:
        raise BankStatementFileUnreadableError(
            "Không đọc được tệp sao kê — kiểm tra lại dấu ngăn cột và bảng mã của tệp",
            reason=type(error).__name__,
        ) from error


def _header_index(header: Sequence[object], wanted: str) -> int:
    """Vị trí của một cột theo **tên**, so sau khi chuẩn hóa khoảng trắng.

    Chuẩn hóa vì tiêu đề sao kê hay mang khoảng trắng thừa và xuống dòng mềm từ
    ô đã gộp — từ chối một tệp vì `"Số tiền "` khác `"Số tiền"` là đúng chữ và
    sai tinh thần, giống hệt lý do ở `excel/reader._headers_of`.
    """
    normalized = [_squash(item) for item in header]
    target = _squash(wanted)
    if not target:
        # Ô tiêu đề trống ở cuối dòng là chuyện bình thường của sao kê, nên một
        # tên cột rỗng trong **hồ sơ** khớp mọi ô ấy và biến cả tệp thành "cột
        # trùng tên" — một câu vô nghĩa về tệp, cho một lỗi của hồ sơ (R3-M2).
        raise BankStatementColumnMissingError(
            "Hồ sơ khai một tên cột rỗng — hãy sửa hồ sơ định dạng của ngân hàng này",
            expected="",
            found="",
        )
    matches = [index for index, item in enumerate(normalized) if item == target]
    if not matches:
        raise BankStatementColumnMissingError(
            f"Tệp sao kê không có cột {wanted!r}",
            expected=wanted,
            found=", ".join(item for item in normalized if item)[:500],
        )
    if len(matches) > 1:
        # Ngân hàng xuất kèm một cột "Số tiền quy đổi" trùng tên là chuyện có
        # thật. Lấy cột đầu là đúng loại lỗi mà việc định vị **theo tên** sinh ra
        # để tránh — mọi dòng vẫn đọc được, mọi con số lấy từ nhầm cột.
        raise BankStatementColumnAmbiguousError(
            f"Tệp sao kê có {len(matches)} cột cùng tên {wanted!r} — hãy sửa hồ sơ "
            "hoặc bỏ bớt cột trùng trong tệp",
            column=wanted,
            positions=", ".join(str(index + 1) for index in matches),
        )
    return matches[0]


def _squash(value: object) -> str:
    return " ".join(str(value).split()) if value is not None else ""


@dataclass(frozen=True)
class _Layout:
    """Vị trí (chỉ số cột) đã phân giải từ tên, một lần cho cả tệp."""

    booked_on: int
    debit: int | None
    credit: int | None
    amount: int | None
    reference: int | None
    description: int | None
    balance: int | None


def _resolve_layout(header: Sequence[object], profile: BankStatementProfile) -> _Layout:
    def optional(name: str | None) -> int | None:
        return None if name is None else _header_index(header, name)

    return _Layout(
        booked_on=_header_index(header, profile.date_col),
        debit=optional(profile.debit_col),
        credit=optional(profile.credit_col),
        amount=optional(profile.amount_col),
        reference=optional(profile.ref_col),
        description=optional(profile.description_col),
        balance=optional(profile.balance_col),
    )


def _parse_rows(
    rows: Sequence[tuple[object, ...]], profile: BankStatementProfile
) -> StatementParseResult:
    header_index = profile.header_row - 1
    if header_index >= len(rows):
        raise BankStatementColumnMissingError(
            f"Tệp sao kê không có dòng {profile.header_row} để lấy tiêu đề cột",
            expected=profile.date_col,
            found="",
        )
    layout = _resolve_layout(rows[header_index], profile)
    lines: list[StatementLine] = []
    issues: list[StatementIssue] = []
    for offset, row in enumerate(rows[header_index + 1 :]):
        number = profile.header_row + 1 + offset
        if all(_squash(cell) == "" for cell in row):
            # Sao kê hầu như luôn có một khối chân bảng (tổng phát sinh, chữ ký)
            # cách phần dữ liệu bằng vài dòng trống. Báo lỗi cho chúng là báo lỗi
            # cho thứ người dùng không sửa được.
            continue
        parsed = _parse_row(row, number, layout, profile)
        if isinstance(parsed, StatementLine):
            lines.append(parsed)
        else:
            issues.extend(parsed)
    return StatementParseResult(lines=tuple(lines), issues=tuple(issues))


def _parse_row(
    row: tuple[object, ...],
    number: int,
    layout: _Layout,
    profile: BankStatementProfile,
) -> StatementLine | list[StatementIssue]:
    """Một dòng → `StatementLine`, hoặc danh sách lỗi của **chính** dòng ấy.

    Gom lỗi rồi trả cả cụm chứ không dừng ở lỗi đầu: một dòng vừa sai ngày vừa
    sai số tiền thường là một dòng bị dán lệch cột, và thấy cả hai lỗi cùng lúc
    là thứ giúp người dùng nhận ra điều đó.
    """
    issues: list[StatementIssue] = []
    booked_on = _read_date(_cell(row, layout.booked_on), number, profile, issues)
    amount = _read_amount(row, number, layout, profile, issues)
    balance = (
        _read_decimal(_cell(row, layout.balance), number, "số dư", profile, issues)
        if layout.balance is not None
        else None
    )
    if issues:
        return issues
    if booked_on is None or amount is None:  # pragma: no cover — `issues` đã rỗng nên cả hai có
        raise AssertionError("dòng không lỗi phải có cả ngày lẫn số tiền")
    return StatementLine(
        row_number=number,
        booked_on=booked_on,
        amount=amount,
        reference=_text(_cell(row, layout.reference)),
        description=_text(_cell(row, layout.description)),
        balance=balance,
    )


def _cell(row: tuple[object, ...], index: int | None) -> object:
    if index is None or index >= len(row):
        return None
    return row[index]


def _text(value: object) -> str | None:
    squashed = _squash(value)
    return squashed or None


def _read_date(
    value: object,
    number: int,
    profile: BankStatementProfile,
    issues: list[StatementIssue],
) -> date | None:
    """Ô ngày → `date`. Ô đã là ngày thì dùng thẳng, không đi qua `date_format`."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _text(value)
    if text is None:
        issues.append(
            StatementIssue(
                row_number=number,
                column=profile.date_col,
                code="bank_statement.date_missing",
                message="Dòng không có ngày giao dịch",
            )
        )
        return None
    try:
        return datetime.strptime(text, profile.date_format).date()  # noqa: DTZ007 — ngày sổ sách, không mốc thời gian
    except ValueError:
        issues.append(
            StatementIssue(
                row_number=number,
                column=profile.date_col,
                code="bank_statement.date_invalid",
                message=f"Ngày không đúng khuôn {profile.date_format!r}",
                value=text,
            )
        )
        return None


def _read_amount(
    row: tuple[object, ...],
    number: int,
    layout: _Layout,
    profile: BankStatementProfile,
    issues: list[StatementIssue],
) -> Decimal | None:
    """Số tiền đã mang **dấu theo quy ước của module** (dương = vào)."""
    if layout.amount is not None:
        value = _read_decimal(
            _cell(row, layout.amount), number, profile.amount_col, profile, issues
        )
        if value is None:
            _require_amount(number, profile.amount_col, issues)
            return None
        rule = AmountSignRule(profile.sign_rule) if profile.sign_rule else AmountSignRule.SIGNED
        if rule is AmountSignRule.DEBIT_POSITIVE:
            if value < 0:
                return _sign_conflict(number, profile.amount_col, value, issues)
            return -value
        return value

    debit = _read_decimal(_cell(row, layout.debit), number, profile.debit_col, profile, issues)
    credit = _read_decimal(_cell(row, layout.credit), number, profile.credit_col, profile, issues)
    if debit is None and credit is None:
        _require_amount(number, profile.debit_col or profile.credit_col, issues)
        return None
    # Gom **cả hai** lỗi dấu rồi mới dừng: `return` sớm ở cột nợ khiến người dùng
    # sửa xong ô thứ nhất mới biết ô thứ hai cũng sai (R3-L4), trái tinh thần
    # "gom lỗi rồi trả cả cụm" của `_parse_row`.
    negative = [
        (profile.debit_col, debit) if debit is not None and debit < 0 else None,
        (profile.credit_col, credit) if credit is not None and credit < 0 else None,
    ]
    found = [item for item in negative if item is not None]
    if found:
        for name, amount in found:
            _sign_conflict(number, name, amount, issues)
        return None
    if debit is not None and credit is not None:
        # Cả hai cột cùng có số nghĩa là dòng bị dán lệch, hoặc hồ sơ khai nhầm
        # hai cột. Cộng chúng lại (`credit - debit`) sẽ ra một con số hợp lệ cho
        # một dòng vô nghĩa — đúng loại lỗi im mà cả module này dựng ra để chặn.
        issues.append(
            StatementIssue(
                row_number=number,
                column=profile.debit_col,
                code="bank_statement.both_sides_filled",
                message="Dòng có số tiền ở cả cột ghi nợ lẫn cột ghi có",
            )
        )
        return None
    return -debit if debit is not None else credit


def _sign_conflict(
    number: int, column: str | None, value: Decimal, issues: list[StatementIssue]
) -> Decimal | None:
    """Ô mang dấu mà hồ sơ không mô tả được — **hỏi**, không tự sửa.

    Bản đầu dùng `-abs(...)`: một ô `-1.000.000` ở cột ghi **nợ** vẫn thành tiền
    ra. Gốc của ca ấy là bút toán **hoàn/điều chỉnh** — ngân hàng ghi số âm ở cột
    chi để nói "trả lại tiền vào tài khoản" — nên sau khi ép dấu, một khoản thu 5
    triệu vào sổ thành một khoản chi 5 triệu: sai 10 triệu ở đối chiếu và ở báo
    cáo lưu chuyển tiền tệ của phase 10.

    Ép dấu là **im lặng sửa dữ liệu người dùng**, đúng thứ quy ước dấu ở docstring
    module được viết ra để không làm. "Hồ sơ khai sai" và "tệp có bút toán hoàn"
    là hai việc khác nhau, và chỉ người dùng phân biệt được.
    """
    issues.append(
        StatementIssue(
            row_number=number,
            column=column,
            code="bank_statement.sign_conflicts_with_profile",
            message="Số tiền mang dấu âm ở một cột mà hồ sơ khai là luôn dương — "
            "hãy kiểm tra lại hồ sơ định dạng hoặc dòng này",
            value=str(value),
        )
    )
    return None


def _require_amount(number: int, column: str | None, issues: list[StatementIssue]) -> None:
    """Chỉ báo "thiếu số tiền" khi **chưa** có lỗi nào khác của cùng ô đó.

    Không có phép kiểm này thì một ô `"1.2.3"` sinh hai lỗi cho một việc sửa:
    "không phải số" rồi "thiếu số tiền".
    """
    if any(
        item.row_number == number and item.code == "bank_statement.amount_invalid"
        for item in issues
    ):
        return
    issues.append(
        StatementIssue(
            row_number=number,
            column=column,
            code="bank_statement.amount_missing",
            message="Dòng không có số tiền",
        )
    )


def _read_decimal(
    value: object,
    number: int,
    column: str | None,
    profile: BankStatementProfile,
    issues: list[StatementIssue],
) -> Decimal | None:
    """Ô số tiền → `Decimal`, theo dấu phân cách khai trong hồ sơ.

    Ô rỗng trả `None` **không kèm lỗi**: ở sao kê hai cột thì mỗi dòng luôn để
    trống một trong hai, và đó là hình dạng bình thường chứ không phải thiếu sót.
    Câu hỏi "dòng này có số tiền không" được trả lời ở `_read_amount`, nơi nhìn
    thấy cả hai cột cùng lúc.
    """
    if isinstance(value, bool):
        # `bool` là `int` trong Python, nên không chặn riêng thì `True` thành `1`.
        return _invalid_amount(number, column, "true" if value else "false", issues)
    if isinstance(value, int):
        return Decimal(value)
    if isinstance(value, Decimal):
        # **Trước** nhánh `Real`: `repr(Decimal("12.5"))` là chuỗi
        # `"Decimal('12.5')"`, và `Decimal(...)` của nó ném `InvalidOperation`
        # **ngoài** khối `try` bên dưới. Bản sửa C2 gộp hai kiểu vào một nhánh và
        # mở đúng lỗ này (R3-H2) — hôm nay openpyxl/csv không trả `Decimal` nên
        # chưa nổ, nhưng chữ ký nhận `object` và phase 6 nạp từ nguồn thứ hai.
        return _finite_or_invalid(value, number, column, issues)
    if isinstance(value, Real):
        # Ô **số** đã là một con số phân giải xong — dấu phân cách của hồ sơ chỉ
        # mô tả cách một **chuỗi** được viết ra, nên áp chúng ở đây là phá dữ liệu.
        #
        # Bản đầu chỉ bắt `int | Decimal`, mà openpyxl trả `float` cho mọi ô không
        # nguyên: `12500000.50` rơi xuống nhánh chuỗi, thành `"12500000.5"`, rồi
        # `thousand_sep = "."` xóa nốt dấu chấm → **125000005**, sai đúng mười lần,
        # không một cảnh báo nào. Ô số nguyên thì đúng, nên lỗi chỉ nổ ở cột số dư
        # và tài khoản ngoại tệ — chỗ ít ai đối chiếu ngay.
        #
        # `Real` chứ không tên kiểu dấu phẩy động: ADR-015 cấm tên ấy trong mã
        # nghiệp vụ (có cổng `test_no_float_in_domain` đếm), và `repr` giữ đúng
        # chữ số người dùng nhìn thấy — cùng cách `excel/reader._cell_text` làm.
        return _finite_or_invalid(Decimal(repr(value)), number, column, issues)
    text = _text(value)
    if text is None:
        return None
    normalized = text.replace(profile.thousand_sep, "") if profile.thousand_sep else text
    normalized = normalized.replace(profile.decimal_sep, ".")
    try:
        parsed = Decimal(normalized)
    except InvalidOperation:
        return _invalid_amount(number, column, text, issues)
    return _finite_or_invalid(parsed, number, column, issues)


def _finite_or_invalid(
    value: Decimal, number: int, column: str | None, issues: list[StatementIssue]
) -> Decimal | None:
    """Chặn `NaN`/`±Infinity` — **một** điểm thoát cho mọi nhánh của `_read_decimal`.

    Cả ba đều hợp lệ về cú pháp `Decimal`, và PostgreSQL `numeric` **nhận** cả
    ba. Một dòng như vậy vào sổ tiền gửi ở phase 6 thì mọi phép cộng sau đó ra
    `NaN` và mọi phép so `> 0` trả `false` — một dòng sao kê làm hỏng số dư của
    cả kỳ, và không có gì trong báo cáo chỉ ra dòng nào gây ra.

    Đặt ở đây chứ không trong nhánh chuỗi, vì bản đầu làm đúng thế và nhánh
    `Real` thêm sau **đi vòng qua nó**: một ô `<v>1e999</v>` trong sheet XML cho
    ra `StatementLine(amount=Infinity)` với `issues` rỗng (R3-H3). Hai bản sửa
    của cùng một hàm, và bản thứ hai mở lại đúng lỗ bản thứ nhất vá.
    """
    if not value.is_finite():
        return _invalid_amount(number, column, str(value), issues)
    return value


def _invalid_amount(
    number: int, column: str | None, value: str, issues: list[StatementIssue]
) -> Decimal | None:
    """Ghi nhận một ô không phải số tiền, và trả `None` cho chỗ gọi dùng lại.

    Trả về thay vì chỉ ghi nhận để hai chỗ gọi viết được `return _invalid_amount(...)`:
    tách thành hai câu lệnh là chỗ để một nhánh quên mất câu `return None` và
    đi tiếp với một giá trị rác.
    """
    issues.append(
        StatementIssue(
            row_number=number,
            column=column,
            code="bank_statement.amount_invalid",
            message="Ô này không phải một số tiền",
            value=value,
        )
    )
    return None


def iter_lines(result: StatementParseResult) -> Iterator[StatementLine]:
    """Dòng đọc được, theo thứ tự trong tệp — tiện cho phase 6 duyệt."""
    yield from result.lines
