"""Xuất một danh mục ra Excel (FR-SYS-014, FR-NFR-061, bước 16).

Dùng lại **đúng** `TemplateDescriptor` của bước nhập liệu, và đó là toàn bộ giá
trị của tệp này. Bất biến nó giữ: *tệp xuất ra là một tệp mẫu đã điền* — cùng
dòng tiêu đề, cùng thứ tự cột, cùng sheet Hướng dẫn — nên xuất rồi nhập lại
không bao giờ đổ ở phép kiểm cấu trúc (FR-SYS-082). Một bộ xuất giữ danh sách
cột riêng sẽ lệch, và nó lệch đúng vào lúc ai đó thêm một cột: tệp khách hàng
vừa tải về bị chính máy chủ từ chối.

Ba quyết định đáng ghi lại:

**Một — truy vấn là một câu `SELECT` có `JOIN`, không phải vòng lặp ORM.** Ô tra
cứu trong tệp là **mã** chứ không phải id (H79), nên mỗi cột khóa ngoại cần một
`LEFT JOIN` về danh mục đích. Đọc từng bản ghi rồi `record.base_unit.code` là
N+1 truy vấn trên 10.000 dòng — đúng thứ LD-14 và bảng §Risk của phase 3 gọi tên.

**Hai — trần đủ hẹp để tệp xuất ra luôn nhập lại được** (`MAX_EXPORT_ROWS`).
Xuất ra một tệp lớn hơn mức hệ thống nhập lại được là tự phá bất biến ở trên.
Trần theo *số dòng* của bộ nhập không đủ: bộ nhập còn một trần theo *byte sau
giải nén*, và 42.000 dòng đã chạm nó — xem `MAX_EXPORT_ROWS`.

**Ba — ô boolean `false` phải là một từ, không phải ô trống.** Ở đường **sửa**,
ô trống nghĩa là "không nói gì về cột này" (H87/H91), nên `false` viết thành ô
trống sẽ round-trip đúng ở đường sửa và **sai** ở đường tạo mới: `is_active`
trống là `true` (`sql.is_active_expression`). Một bản ghi đã ngừng theo dõi,
xuất ra rồi nhập vào một dữ liệu kế toán khác, sẽ sống lại. Nên `false` viết
thành `"không"` — một từ nằm trong `sql.FALSE_WORDS`, và là từ người dùng đọc
được ngay.
"""

from __future__ import annotations

from collections.abc import Iterator, Sequence
from decimal import Decimal
from io import BytesIO
from typing import Any, Final

from openpyxl import Workbook
from openpyxl.cell.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import ColumnElement, Select, Table, and_, func, select
from sqlalchemy.orm import Session
from sqlalchemy.sql.selectable import NamedFromClause

from ket.kernel.errors import ExportTooManyRowsError
from ket.kernel.excel.descriptors import CellKind, ColumnDescriptor, TemplateDescriptor
from ket.kernel.excel.reader import MAX_UNCOMPRESSED_BYTES, uncompressed_size
from ket.kernel.excel.sql import table_for, table_of, visible_to
from ket.kernel.excel.template import (
    HEADER_FILL,
    HEADER_WIDTH,
    OPTIONAL_FONT,
    REQUIRED_FONT,
    write_instructions_sheet,
)
from ket.kernel.master_data.registry import CatalogSpec

TRUE_CELL: Final[str] = "x"
FALSE_CELL: Final[str] = "không"
"""Cách viết một ô boolean ra tệp — xem đoạn "Ba" ở docstring module.

Cả hai từ đều nằm trong bộ từ khóa mà bước nhập nhận (`sql.TRUE_WORDS` /
`sql.FALSE_WORDS`), và `tests/test_export.py` canh đúng điều đó thay vì tin vào
việc hai hằng số này được đọc lại lúc sửa."""

MAX_EXPORT_ROWS: Final[int] = 10_000
"""Trần số dòng của một tệp xuất — **không** bằng `reader.MAX_ROWS` (100.000).

Bản đầu của lát này dùng chung trần với bộ nhập, với lập luận "xuất ra thứ nhập
lại được". Lập luận đúng, con số sai: review vòng 1 đo `partners` (22 cột, ô
đầy) và thấy **42.000 dòng đã vượt trần 64 MiB sau giải nén** của chính bộ nhập
(`reader.MAX_UNCOMPRESSED_BYTES`) — tức trần theo *số dòng* không bao được trần
theo *byte*, và hệ thống vẫn sinh ra được tệp chính nó từ chối.

10.000 là mức FR-NFR-043 đặt ra cho một lượt nhập. Nó cũng giữ được quyết định
"xuất chạy đồng bộ": 100.000 dòng × 15 cột tốn **7,9 giây CPU** ngay trong một
request, còn ở mức này là **≈1,5 giây**.

Trần theo số dòng vẫn **không đủ một mình** — xem `_ensure_importable`. Một danh
mục nhiều cột chữ với ô tiếng Việt đầy cho ra hơn 64 MiB ngay ở 10.000 dòng, nên
phép đo byte sau khi dựng tệp mới là thứ giữ bất biến.

Danh mục lớn hơn mức này xuất theo từng nhánh — `?parent_id=` của đường đọc danh
sách là chỗ 3D sẽ nối vào."""

STREAM_BATCH: Final[int] = 1_000
"""Số dòng mỗi lượt lấy từ DB. Đủ lớn để không phải một vòng round-trip mỗi
dòng, đủ nhỏ để 10.000 dòng không nằm trọn trong bộ nhớ cùng lúc với sổ Excel
đang dựng."""


def _reference_alias(column: ColumnDescriptor, index: int) -> NamedFromClause:
    """Bảng đích của một cột tra cứu, đặt bí danh riêng cho từng cột.

    Bí danh chứ không dùng thẳng bảng: `parent_code` trỏ về **chính danh mục
    này** (tự nối), và một danh mục có thể trỏ hai lần vào cùng một bảng đích
    (vật tư hàng hóa trỏ `warehouses` một lần hôm nay, và phase 8 sẽ thêm vai trò
    thứ hai). Không có bí danh thì câu lệnh nối nhầm vào chính nó.
    """
    return table_of(column.reference_slug).alias(f"ref_{index}")


def export_statement(
    spec: CatalogSpec,
    descriptor: TemplateDescriptor,
    *,
    branch_id: int | None,
    include_inactive: bool,
) -> Select[Any]:
    """Câu lệnh đọc dữ liệu của một danh mục theo đúng hình dạng tệp mẫu.

    Phạm vi đọc là `visible_to` — **không** `owned_by`. Xuất là một thao tác
    đọc, và một chi nhánh phải xuất được cả phần dùng chung toàn công ty mà màn
    hình của họ vẫn hiện. (Phạm vi **ghi** thì hẹp hơn, và `sql.owned_by` giải
    thích vì sao hai thứ đó không được là một.)

    Sắp theo `path`: đó là thứ tự cây — nút cha luôn đứng trước con cháu của nó,
    vì `path` của cha là tiền tố của `path` con. Người đọc tệp thấy đúng cấu trúc
    họ thấy trên màn hình, và dòng `Mã nhóm cha` luôn trỏ ngược lên một dòng phía
    trên chứ không xuống dưới.
    """
    table = table_for(spec.model)
    statement = select().select_from(table)
    for index, column in enumerate(descriptor.columns):
        if column.kind is CellKind.CODE_REF:
            alias = _reference_alias(column, index)
            source_field = column.target_field or column.field
            statement = statement.add_columns(alias.c.code).join(
                alias,
                # Kèm `visible_to` chứ không chỉ so `id`: hôm nay ba lớp canh
                # (router, nhập liệu, gộp) không cho một khóa ngoại trỏ ra ngoài
                # tầm nhìn của chi nhánh tồn tại, nên nhánh này chưa tới được.
                # Nhưng nếu nó tới được thì hậu quả là một mã của chi nhánh khác
                # hiện ra trong tệp xuất — tức đường rò dữ liệu, và nó rò qua
                # đúng lớp lọc mà `where` bên dưới được viết ra để giữ.
                and_(table.c[source_field] == alias.c.id, visible_to(alias, branch_id)),
                isouter=True,
            )
        else:
            statement = statement.add_columns(table.c[column.field])
    for condition in _scope(table, branch_id=branch_id, include_inactive=include_inactive):
        statement = statement.where(condition)
    return statement.order_by(table.c.path)


def _scope(
    table: Table, *, branch_id: int | None, include_inactive: bool
) -> tuple[ColumnElement[bool], ...]:
    """Điều kiện "tệp xuất gồm dòng nào" — **một** bản, hai chỗ đọc.

    Bản đầu để `export_statement` và `_count_statement` mỗi bên tự viết bộ lọc
    của mình. Review vòng 1 chỉ ra: hai đột biến làm lệch chúng đều sống sót, và
    hậu quả của việc lệch là một trần đếm trên tập này rồi dựng tệp từ tập kia —
    tức trần không còn canh đúng thứ nó tưởng đang canh.
    """
    conditions = [visible_to(table, branch_id)]
    if not include_inactive:
        conditions.append(table.c.is_active.is_(True))
    return tuple(conditions)


def _count_statement(
    spec: CatalogSpec, *, branch_id: int | None, include_inactive: bool
) -> Select[tuple[int]]:
    """Đếm trước khi dựng tệp, để chạm trần thì không tốn công dựng gì."""
    table = table_for(spec.model)
    statement = select(func.count()).select_from(table)
    for condition in _scope(table, branch_id=branch_id, include_inactive=include_inactive):
        statement = statement.where(condition)
    return statement


MAX_EXACT_DIGITS: Final[int] = 15
"""Số chữ số có nghĩa mà một ô Excel giữ lại **nguyên vẹn**.

Định dạng xlsx lưu mọi số dưới dạng dấu phẩy động 64 bit, thứ round-trip chính
xác được 15 chữ số thập phân có nghĩa (16–17 chữ số thì tùy giá trị). Cột của
danh mục rộng hơn thế: `partners.credit_limit` là `NUMERIC(20, 6)`.

Lấy mức **bảo đảm** chứ không mức tối đa: đoán rộng thêm hai chữ số chỉ đổi được
vài ô từ dạng chuỗi sang dạng số, còn đoán hụt thì một con số đổi giá trị trong
im lặng. Cái giá của việc thận trọng ở đây rẻ hơn hẳn cái giá của việc sai."""


def _survives_excel_storage(value: int | Decimal) -> bool:
    """Con số này có quay về nguyên vẹn sau một vòng qua tệp xlsx không?

    Một hạn mức nợ `99999999999999.999999` ghi ra dưới dạng số sẽ đọc lại thành
    `100000000000000` — **không** lỗi, không cảnh báo, và một vòng xuất rồi nhập
    lại vừa đổi hạn mức nợ của một khách hàng.

    Đếm chữ số trên `Decimal` chứ không thử ép qua dấu phẩy động rồi so: ADR-015
    cấm **tên** `float` trong mã nghiệp vụ (có cổng `test_no_float_in_domain`
    đếm), và lệ đó không nên có ngoại lệ chỉ vì chỗ này tình cờ không tính tiền.
    Phép đếm cũng nói đúng thứ đang được hỏi.

    `normalize()` trước khi đếm: `Decimal` giữ số 0 ở đuôi như chữ số có nghĩa,
    nên `5000000000.000000` đọc thô là 16 chữ số trong khi nó là một con số một
    chữ số có nghĩa mà mọi ô Excel giữ được thừa sức.
    """
    return len(Decimal(value).normalize().as_tuple().digits) <= MAX_EXACT_DIGITS


def _cell_value(column: ColumnDescriptor, value: object) -> str | int | Decimal | None:
    """Giá trị từ DB → giá trị ghi vào ô Excel.

    Số ghi dưới dạng **số** khi ghi được: người nhận tệp cộng cột đó bằng Excel,
    và một cột số lưu thành text là cột không cộng được (và mang tam giác xanh
    báo lỗi ở mọi ô). `Decimal` đi thẳng vào openpyxl — nó nằm trong bộ kiểu số
    openpyxl nhận, nên không phải qua `float`.

    Ngoại lệ là những giá trị mà chính định dạng xlsx không giữ nổi: chúng ghi
    thành **chuỗi**, nơi chúng chính xác tuyệt đối. Đánh đổi có chủ đích — mất
    khả năng cộng bằng Excel ở đúng những ô ấy, đổi lấy việc không có con số nào
    lặng lẽ đổi giá trị trên đường xuất–nhập. Bước nhập nhận cả hai dạng như
    nhau: `staging.py` ép ô về kiểu cột bằng SQL, từ một chuỗi.
    """
    if value is None:
        return None
    if column.kind is CellKind.BOOLEAN or isinstance(value, bool):
        return TRUE_CELL if value else FALSE_CELL
    if isinstance(value, int | Decimal):
        return value if _survives_excel_storage(value) else str(value)
    return str(value)


def _as_cell(sheet: Worksheet, value: str | int | Decimal | None) -> object:
    """Giá trị → ô, và **ô chuỗi luôn là chuỗi** (review vòng 1, C2).

    openpyxl suy kiểu ô từ nội dung: một chuỗi bắt đầu bằng `=` thành ô **công
    thức** (`data_type = "f"`). Hai hậu quả, cả hai đều thật:

    * **Tiêm công thức.** Tên đối tác `=cmd|'/c calc'!A0` do một người dùng nhập
      vào hệ thống sẽ chạy trên máy của người mở tệp xuất ra. Đây là đường dữ
      liệu của người này thành lệnh trên máy người khác, và tệp danh mục thì
      được gửi qua email suốt.
    * **Mất dữ liệu im lặng.** `reader.py` mở tệp với `data_only=True`, tức đọc
      *kết quả* của công thức — mà tệp do chúng ta ghi ra không có kết quả nào
      được tính sẵn, nên ô đọc lại thành `None`. Vòng xuất–nhập vì thế xóa trắng
      ô đó, hoặc báo "Cột 'Tên' bắt buộc nhập" cho một dòng có tên.

    Ép `data_type` **sau** khi gán `value`: chính lượt gán là chỗ openpyxl suy ra
    `"f"`, nên đặt trước sẽ bị ghi đè.

    Ô số và ô rỗng đi thẳng: chúng phải giữ đúng kiểu số để cộng được trong Excel
    (xem `_cell_value`), và không có đường nào để một số thành công thức.
    """
    if not isinstance(value, str):
        return value
    cell = WriteOnlyCell(sheet, value=value)
    cell.data_type = "s"
    return cell


def _write_data_sheet(
    workbook: Workbook,
    descriptor: TemplateDescriptor,
    rows: Iterator[Sequence[object]],
) -> None:
    """Sheet dữ liệu: dòng tiêu đề của tệp mẫu, rồi dữ liệu.

    Dòng tiêu đề lấy từ `descriptor.headers` — cùng nguồn mà `reader.ensure_structure`
    so lại lúc nhập, nên hai bên không thể lệch.
    """
    sheet = workbook.create_sheet(descriptor.sheet_name)
    for index in range(1, len(descriptor.columns) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = HEADER_WIDTH
    sheet.freeze_panes = "A2"
    header: list[object] = []
    for column in descriptor.columns:
        cell = WriteOnlyCell(sheet, value=column.display_header)
        cell.font = REQUIRED_FONT if column.required else OPTIONAL_FONT
        cell.fill = HEADER_FILL
        header.append(cell)
    sheet.append(header)
    for row in rows:
        # `strict=True`: câu lệnh `SELECT` được dựng từ **chính** `descriptor.columns`
        # ngay trên, nên hai độ dài lệch nhau là một lỗi lập trình — và nó phải nổ ở
        # đây chứ không lặng lẽ cắt cụt vài cột cuối của mọi dòng trong tệp.
        sheet.append(
            [
                _as_cell(sheet, _cell_value(column, value))
                for column, value in zip(descriptor.columns, row, strict=True)
            ]
        )


def build_export(
    session: Session,
    spec: CatalogSpec,
    descriptor: TemplateDescriptor,
    *,
    branch_id: int | None,
    include_inactive: bool,
) -> bytes:
    """Tệp .xlsx của một danh mục dưới dạng byte.

    `write_only=True`: sổ ghi-một-lượt không giữ đối tượng ô nào sau khi ghi,
    nên 10.000 dòng × 15 cột không thành 150.000 đối tượng Python trong bộ nhớ
    máy chủ — cùng lý do `read_only=True` ở `reader.py`.
    """
    total = session.execute(
        _count_statement(spec, branch_id=branch_id, include_inactive=include_inactive)
    ).scalar_one()
    if total > MAX_EXPORT_ROWS:
        raise ExportTooManyRowsError(
            f"Danh mục có {total} dòng, vượt trần {MAX_EXPORT_ROWS} dòng của một tệp — "
            "hãy xuất theo từng nhánh",
            max_rows=MAX_EXPORT_ROWS,
            total=total,
        )
    statement = export_statement(
        spec, descriptor, branch_id=branch_id, include_inactive=include_inactive
    )
    workbook = Workbook(write_only=True)
    write_instructions_sheet(workbook, spec, descriptor)
    result = session.execute(statement).yield_per(STREAM_BATCH)
    _write_data_sheet(workbook, descriptor, (row for row in result))
    buffer = BytesIO()
    workbook.save(buffer)
    content = buffer.getvalue()
    _ensure_importable(content, total)
    return content


def _ensure_importable(content: bytes, total: int) -> None:
    """Tệp vừa dựng có nằm trong **trần byte** của bộ nhập không?

    Trần theo số dòng (`MAX_EXPORT_ROWS`) một mình **không đủ**, và đó là bài học
    đắt nhất của hai vòng review: bộ nhập còn `MAX_UNCOMPRESSED_BYTES` = 64 MiB,
    mà 10.000 dòng `partners` với ô tiếng Việt đầy đo được **93,7 MiB** — chữ
    tiếng Việt tốn 3 byte UTF-8 mỗi ký tự, nên số byte mỗi dòng đổi theo *nội
    dung*, không theo số cột. Một hằng số "byte mỗi dòng" đoán trước là thứ vòng
    review đã bắt được là bịa.

    Nên đo **chính tệp vừa dựng**. Tốn công dựng rồi mới từ chối, nhưng chỉ ở ca
    bệnh lý, và đổi lại bất biến trở thành thứ không phụ thuộc vào bảng chữ cái
    hay số cột của danh mục nào.
    """
    size = uncompressed_size(BytesIO(content))
    if size is not None and size > MAX_UNCOMPRESSED_BYTES:
        raise ExportTooManyRowsError(
            f"Tệp xuất ra ({size // (1024 * 1024)} MB sau giải nén) lớn hơn mức bộ nhập "
            f"nhận lại được — hãy xuất theo từng nhánh",
            max_bytes=MAX_UNCOMPRESSED_BYTES,
            total=total,
        )


def export_file_name(spec: CatalogSpec) -> str:
    """Tên tệp gợi ý — ASCII thuần, cùng lý do với `template.template_file_name`."""
    return f"danh-muc-{spec.slug.replace('_', '-')}.xlsx"
