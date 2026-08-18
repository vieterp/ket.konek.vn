"""Đọc tệp xlsx người dùng nộp lên: kiểm cấu trúc rồi phát dòng thô.

Hai việc, theo đúng thứ tự đó, và thứ tự là một phần của thiết kế. FR-SYS-082
nói tệp đổi tên sheet hoặc tên cột thì **không nhận** — nên phép kiểm cấu trúc
chạy trước và dừng hẳn, thay vì để mười nghìn dòng đi qua rồi báo mười nghìn lỗi
"thiếu cột Mã". Người dùng đổi tên cột không có mười nghìn vấn đề, họ có một.

Thông điệp từ chối **nêu tên đúng** (bước 13): "sheet phải tên `Vật tư hàng hóa`,
tệp đang có `Sheet1`". Một câu "sai cấu trúc" thì đúng nhưng vô dụng — nó buộc
người dùng tải lại tệp mẫu và so từng cột bằng mắt.

`read_only=True` ở mọi đường đọc: openpyxl mặc định dựng đối tượng ô cho toàn bộ
bảng tính trong bộ nhớ, và 10.000 dòng × 15 cột là 150.000 đối tượng Python cho
một việc chỉ cần đọc tuần tự một lần.
"""

from __future__ import annotations

from collections.abc import Iterator
from datetime import date, datetime
from decimal import Decimal
from numbers import Real
from typing import IO, Final
from zipfile import BadZipFile, ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ket.kernel.errors import (
    ImportFileUnreadableError,
    ImportSheetMissingError,
    ImportTemplateMismatchError,
    ImportTooManyRowsError,
)
from ket.kernel.excel.descriptors import INSTRUCTIONS_SHEET, TemplateDescriptor

HEADER_ROW: Final[int] = 1
"""Dòng tiêu đề. Cố định chứ không dò tìm: một tệp mẫu strict có đúng một chỗ
cho tiêu đề, và "dò dòng tiêu đề" là tính năng của khung *không* strict — tức
của sao kê ngân hàng, thứ RT-26 tách hẳn ra khỏi đây."""

FIRST_DATA_ROW: Final[int] = HEADER_ROW + 1

MAX_ROWS: Final[int] = 100_000
"""Trần cứng số dòng dữ liệu.

FR-NFR-043 đặt mục tiêu ở 10.000 dòng; trần gấp mười lần đó là chỗ để một tệp
lớn bất thường vẫn chạy, còn một tệp 5 triệu dòng (dán nhầm, hoặc cố ý) thì dừng
với một câu nói rõ trần thay vì ăn hết bộ nhớ máy chủ của cả phòng kế toán."""

MAX_SCANNED_ROWS: Final[int] = MAX_ROWS * 5
"""Trần số dòng **quét qua**, kể cả dòng rỗng.

Rộng hơn `MAX_ROWS` vì Excel thường để lại một vệt dòng rỗng dưới vùng dữ liệu
thật và đó là chuyện bình thường; nhưng vẫn là một trần, vì `MAX_ROWS` một mình
không chặn được gì: nó chỉ đếm dòng **có nội dung**, nên một tệp toàn dòng rỗng
đi qua nó vô hạn."""

MAX_UNCOMPRESSED_BYTES: Final[int] = 64 * 1024 * 1024
"""Trần tổng kích thước **sau giải nén** của tệp xlsx.

Tồn tại vì hai trần kia đều đo sai chỗ: `IMPORT_MAX_BYTES` đo tệp **đã nén** (tệp
tấn công 172 KB lọt qua dễ dàng), còn `MAX_SCANNED_ROWS` chỉ có tác dụng sau khi
vòng lặp đọc đã bắt đầu — mà chi phí nằm trước đó, trong `load_workbook`.

Con số lấy từ phép đo, không đoán: một tệp **hợp lệ** 10.000 dòng (mức FR-NFR-043
đặt ra) giải nén ra **1,5 MB** cho cả `items` lẫn `partners` — danh mục nhiều cột
nhất. 64 MiB là hơn bốn mươi lần mức đó, đủ rộng để không bao giờ chạm một tệp
người dùng soạn thật, và vẫn chặn được tệp 172 KB chứa 20 triệu thẻ `<row/>` rỗng
(~120 MB sau giải nén)."""

_TEXT_TRUNCATE: Final[int] = 500
"""Trần độ dài một ô khi đọc vào. Cột dài nhất của danh mục là 255 ký tự, nên
mọi giá trị vượt trần này chắc chắn đã là một dòng lỗi — giữ lại nguyên văn 5000
ký tự chỉ để in ra câu "quá dài" là phí cả bộ nhớ lẫn chỗ trong JSONB."""


def _sheet_of(
    path_or_stream: IO[bytes], descriptor: TemplateDescriptor
) -> tuple[Workbook, Worksheet]:
    """Mở tệp và lấy đúng sheet dữ liệu, hoặc từ chối kèm tên đúng."""
    _ensure_within_zip_budget(path_or_stream)
    try:
        workbook = load_workbook(path_or_stream, read_only=True, data_only=True)
    except Exception as error:
        # Tệp hỏng, tệp .xls cũ, tệp đổi đuôi, tệp mã hóa: với người dùng đó là
        # **một** tình huống ("máy chủ không đọc được tệp này") và họ xử lý nó
        # bằng cùng một hành động — lưu lại bằng Excel dưới dạng .xlsx.
        raise ImportFileUnreadableError(
            "Không đọc được tệp — hãy lưu lại bằng Excel ở định dạng .xlsx rồi thử lại",
            reason=type(error).__name__,
        ) from error

    if descriptor.sheet_name not in workbook.sheetnames:
        present = [name for name in workbook.sheetnames if name != INSTRUCTIONS_SHEET]
        workbook.close()
        raise ImportSheetMissingError(
            f"Tệp phải có sheet tên {descriptor.sheet_name!r}",
            expected=descriptor.sheet_name,
            found=present,
        )
    return workbook, workbook[descriptor.sheet_name]


def uncompressed_size(path_or_stream: IO[bytes]) -> int | None:
    """Tổng kích thước **sau giải nén** đọc từ mục lục zip, hoặc `None` nếu không phải zip.

    Đọc con số ghi sẵn trong mục lục (`ZipInfo.file_size`) nên không giải nén
    byte nào. Công khai vì **bộ xuất** cũng phải hỏi đúng câu này: một tệp nó
    sinh ra mà vượt trần của bộ nhập là một tệp hệ thống tự từ chối
    (`exporter.build_export`).
    """
    position = path_or_stream.tell()
    try:
        with ZipFile(path_or_stream) as archive:
            return sum(item.file_size for item in archive.infolist())
    except BadZipFile:
        return None
    finally:
        path_or_stream.seek(position)


def _ensure_within_zip_budget(path_or_stream: IO[bytes]) -> None:
    """Từ chối tệp mà **nội dung đã giải nén** vượt trần, trước khi mở nó.

    `MAX_SCANNED_ROWS` chặn vòng lặp đọc, nhưng chi phí thật nằm **trước** vòng
    lặp: khi bảng tính không khai thẻ `<dimension>`, openpyxl phải quét trọn phần
    XML của sheet ngay trong `load_workbook` để biết vùng dữ liệu. Một tệp 172 KB
    chứa 20 triệu thẻ `<row/>` rỗng vì thế ngốn ~9 giây CPU **trước khi** dòng đầu
    tiên được phát ra — không trần nào ở tầng trên chạm tới được.

    Đo tỷ lệ nén thay vì đọc nội dung: `ZipInfo.file_size` là con số ghi trong mục
    lục zip, đọc được mà không giải nén byte nào. Tệp nói dối kích thước thì phần
    giải nén thật vẫn bị chính openpyxl chặn ở mức nó đọc được.

    Tệp hợp lệ 10.000 dòng × 15 cột giải nén ra khoảng vài MB, nên trần 512 MB
    rộng hơn hai bậc — nó chỉ chạm những tệp không ai soạn bằng tay.
    """
    total = uncompressed_size(path_or_stream)
    if total is None:
        # Không phải zip: `load_workbook` sẽ nói câu đó rõ hơn ta.
        return
    if total > MAX_UNCOMPRESSED_BYTES:
        raise ImportTooManyRowsError(
            "Tệp giải nén ra lớn hơn mức cho phép — hãy xóa vùng trống phía dưới "
            "dữ liệu rồi lưu lại bằng Excel",
            max_bytes=MAX_UNCOMPRESSED_BYTES,
        )


def _headers_of(sheet: Worksheet, column_count: int) -> list[str]:
    """Dòng tiêu đề, đã chuẩn hóa khoảng trắng.

    Chuẩn hóa vì Excel và người dùng thêm khoảng trắng ở cuối ô một cách vô hình
    — từ chối một tệp vì `"Mã "` khác `"Mã"` là đúng luật FR-SYS-082 về mặt chữ
    và sai hoàn toàn về mặt tinh thần: cột **không** bị đổi tên.
    """
    rows = sheet.iter_rows(
        min_row=HEADER_ROW, max_row=HEADER_ROW, max_col=column_count, values_only=True
    )
    first = next(iter(rows), ())
    return [" ".join(str(value).split()) if value is not None else "" for value in first]


def ensure_structure(path_or_stream: IO[bytes], descriptor: TemplateDescriptor) -> None:
    """Kiểm sheet + tiêu đề cột, ném ngay khi lệch (FR-SYS-082, bước 13).

    So **cả thứ tự**, không chỉ tập hợp tên: các cột được đọc theo vị trí ở
    `iter_rows`, nên hai cột đổi chỗ cho nhau mà vẫn nhận là cách ghi tên vào ô
    mã và mã vào ô tên — im lặng, và chỉ lộ ra khi ai đó mở danh mục lên xem.
    """
    workbook, sheet = _sheet_of(path_or_stream, descriptor)
    try:
        expected = list(descriptor.headers)
        found = _headers_of(sheet, len(expected))
        if found != expected:
            missing = [item for item in expected if item not in found]
            unexpected = [item for item in found if item and item not in expected]
            raise ImportTemplateMismatchError(
                "Dòng tiêu đề không khớp tệp mẫu — hãy tải lại tệp mẫu và dán dữ liệu vào đó",
                expected=expected,
                found=found,
                missing=missing,
                unexpected=unexpected,
            )
    finally:
        workbook.close()


def _cell_text(value: object) -> str | None:
    """Một ô Excel → chuỗi thô, hoặc `None` khi ô trống.

    Giữ **nguyên văn** người dùng gõ là hợp đồng của cột `cells` ở `models.py` —
    phép ép kiểu diễn ra bằng SQL sau đó, nơi lỗi ép kiểu chính là thứ cần báo.
    Bốn loại openpyxl trả về cần chú ý:

    * **Số thực** — Excel lưu mọi số dưới dạng đó. `Decimal(repr(...))` giữ đúng
      chữ số người dùng nhìn thấy; ép thẳng sang chuỗi cho ra
      `0.30000000000000004` trên những giá trị hoàn toàn bình thường.

      Nhận dạng bằng `numbers.Real` chứ không bằng chính tên kiểu: ADR-015 cấm
      **tên** đó xuất hiện trong mã nghiệp vụ (có cổng `test_no_float_in_domain`
      đếm), vì mọi phép tính tiền phải là `Decimal`. Ở đây không có phép tính
      nào — chỉ có một giá trị từ thư viện ngoài cần đổi sang chuỗi *ngay lập
      tức* — nhưng lệ thì không nên có ngoại lệ, và `numbers.Real` diễn đạt đúng
      câu hỏi đang hỏi ("đây có phải một số không nguyên không").
    * `int` — bắt **trước** `Real` (mọi `int` cũng là `Real`), để `5` không
      thành `"5"` qua đường `Decimal` với một số 0 thập phân thừa ra.
    * `datetime`/`date` — Excel tự đổi một ô mã như `03-01` thành ngày tháng.
      Giữ dạng ISO để câu báo lỗi trích được đúng thứ đang nằm trong ô, nên
      người dùng nhìn ra ngay là Excel đã đổi định dạng của họ.
    * `bool` — bắt trước `int` (mọi `bool` cũng là `int`), và trả `true`/`false`
      chứ không `1`/`0`, để phép nhận dạng ở `staging.py` chỉ phải biết một bộ
      từ khóa.
    """
    match value:
        case None:
            return None
        case bool():
            return "true" if value else "false"
        case int():
            return str(value)
        case Real():
            return str(Decimal(repr(value)))
        case datetime():
            return value.isoformat()
        case date():
            return value.isoformat()
        case _:
            text = str(value).strip()
            return text[:_TEXT_TRUNCATE] if text else None


def iter_rows(
    path_or_stream: IO[bytes], descriptor: TemplateDescriptor
) -> Iterator[tuple[int, dict[str, str | None]]]:
    """Phát `(số dòng như trong Excel, {tên trường: chuỗi thô})`.

    **Bỏ qua dòng trống hoàn toàn** thay vì báo lỗi: người dùng dán dữ liệu vào
    tệp mẫu và Excel hay giữ lại vài nghìn dòng rỗng phía dưới vùng đã dùng. Báo
    lỗi cho chúng là báo lỗi cho một thứ họ không nhìn thấy và không sửa được.

    Không bỏ qua dòng *một phần* trống — thiếu ô bắt buộc là một lỗi thật, và
    đoán rằng người dùng "chắc không định nhập dòng này" là cách bỏ sót dữ liệu
    trong im lặng.
    """
    workbook, sheet = _sheet_of(path_or_stream, descriptor)
    fields = [column.field for column in descriptor.columns]
    try:
        emitted = 0
        for offset, raw in enumerate(
            sheet.iter_rows(min_row=FIRST_DATA_ROW, max_col=len(fields), values_only=True)
        ):
            values = {
                field: _cell_text(raw[index] if index < len(raw) else None)
                for index, field in enumerate(fields)
            }
            # Trần áp cho số dòng **đã quét**, không cho số dòng phát ra: một
            # tệp 172 KB khai vùng đã dùng tới vài triệu dòng rỗng sẽ giữ vòng
            # lặp này chạy hàng chục giây mà `emitted` không bao giờ chạm trần.
            # Dòng rỗng không tốn bộ nhớ, nhưng nó tốn đúng thứ FR-NFR-044 bảo
            # vệ: một worker.
            if offset >= MAX_SCANNED_ROWS:
                raise ImportTooManyRowsError(
                    f"Tệp khai vùng dữ liệu quá {MAX_SCANNED_ROWS} dòng — hãy xóa phần "
                    "trống phía dưới rồi lưu lại",
                    max_rows=MAX_SCANNED_ROWS,
                )
            if all(value is None for value in values.values()):
                continue
            emitted += 1
            if emitted > MAX_ROWS:
                raise ImportTooManyRowsError(
                    f"Tệp vượt quá {MAX_ROWS} dòng dữ liệu — hãy tách thành nhiều tệp",
                    max_rows=MAX_ROWS,
                )
            yield FIRST_DATA_ROW + offset, values
    finally:
        workbook.close()
