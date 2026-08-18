"""Sinh tệp mẫu Excel cho một danh mục (FR-SYS-080, FR-SYS-083, bước 12).

Hai sheet: **Hướng dẫn** rồi tới sheet dữ liệu. Thứ tự đó có chủ đích — Excel mở
sheet đầu tiên, và thứ người dùng cần đọc trước khi dán dữ liệu là hướng dẫn.

Cột bắt buộc mang dấu `*` **trong chính tiêu đề** (`ColumnDescriptor.display_header`),
không ở một dòng chú thích riêng: người dùng dán dữ liệu bằng cách nhìn dòng
tiêu đề, và một chú thích ở dòng khác là chú thích họ đã cuộn qua.

`write_only=True`: cùng lý do với `read_only` ở `reader.py` — lát 3C-2 sẽ dùng
lại chính bộ này để **xuất** dữ liệu, và một danh mục vật tư 10.000 dòng không
được biến thành 150.000 đối tượng ô trong bộ nhớ máy chủ.
"""

from __future__ import annotations

from io import BytesIO
from typing import Final

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ket.kernel.excel.descriptors import (
    INSTRUCTIONS_SHEET,
    TemplateDescriptor,
    instructions_for,
)
from ket.kernel.master_data.registry import CatalogSpec

# Bốn hằng số trình bày dưới đây **công khai** vì bộ xuất (`exporter.py`, bước
# 16) dựng cùng dòng tiêu đề ấy trên một sổ `write_only`. Chép lại chúng ở đó là
# chép một quyết định trình bày sang chỗ thứ hai, và hai chỗ sẽ lệch — tệp mẫu
# tải về trông khác tệp dữ liệu xuất ra, dù cả hai được cho là cùng một biểu mẫu.
HEADER_FILL: Final[PatternFill] = PatternFill("solid", fgColor="E8EEF7")
REQUIRED_FONT: Final[Font] = Font(bold=True, color="B4232C")
"""Cột bắt buộc đậm và đỏ, ngoài dấu `*`.

Hai tín hiệu cho cùng một thông tin, có chủ đích: dấu `*` sống sót qua thao tác
sao chép sang một bảng tính khác (thứ người dùng làm suốt), còn màu thì nhìn
thấy ngay mà không phải đọc."""

OPTIONAL_FONT: Final[Font] = Font(bold=True)

HEADER_WIDTH: Final[int] = 22
_NOTE_WIDTH: Final[int] = 90

XLSX_MEDIA_TYPE: Final[str] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
"""Kiểu MIME của tệp .xlsx.

Ở kernel chứ ở một router: **hai** router trả tệp .xlsx (nhập liệu trả tệp mẫu,
xuất trả dữ liệu), và bản đầu để `routers/exports.py` import hằng số từ
`routers/imports.py` — một router phụ thuộc router chỉ vì một chuỗi."""


def write_instructions_sheet(
    workbook: Workbook, spec: CatalogSpec, descriptor: TemplateDescriptor
) -> None:
    """Sheet **Hướng dẫn** — mỗi cột một dòng, sinh từ descriptor.

    Chỉ dùng `append` + `column_dimensions` nên nó chạy được trên **cả** sổ
    thường lẫn sổ `write_only`: bộ xuất cần đúng sheet này để tệp xuất ra vẫn là
    một tệp mẫu hoàn chỉnh, còn `reader` thì bỏ qua nó theo tên (`INSTRUCTIONS_SHEET`).
    """
    sheet = workbook.create_sheet(INSTRUCTIONS_SHEET)
    sheet.append([f"Tệp mẫu nhập liệu — {spec.title}"])
    sheet.append([])
    sheet.append(
        [
            f"Dán dữ liệu vào sheet {descriptor.sheet_name!r}, bắt đầu từ dòng 2.",
        ]
    )
    sheet.append(["Không đổi tên sheet, không đổi tên cột, không đổi thứ tự cột."])
    sheet.append(["Cột có dấu * là cột bắt buộc nhập."])
    sheet.append([])
    sheet.append(["Cột", "Giải thích"])
    for header, note in instructions_for(spec, descriptor):
        sheet.append([header, note])
    sheet.column_dimensions["A"].width = HEADER_WIDTH
    sheet.column_dimensions["B"].width = _NOTE_WIDTH


def _write_header(workbook: Workbook, descriptor: TemplateDescriptor) -> None:
    """Sheet dữ liệu — đúng một dòng tiêu đề, không dòng dữ liệu mẫu.

    Không kèm dòng ví dụ: một dòng ví dụ trong tệp mẫu là một dòng người dùng sẽ
    quên xóa, và nó sẽ được nhập vào danh mục thật dưới cái tên "Nguyễn Văn A".
    """
    sheet = workbook.create_sheet(descriptor.sheet_name)
    sheet.append(list(descriptor.headers))
    for index, column in enumerate(descriptor.columns, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = REQUIRED_FONT if column.required else OPTIONAL_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.column_dimensions[get_column_letter(index)].width = HEADER_WIDTH
    # Khóa dòng tiêu đề: người dùng cuộn tới dòng 4.000 vẫn phải biết cột nào là
    # cột nào. Không có nó thì họ dán lệch cột — sai sót mà phép kiểm cấu trúc
    # không bắt được, vì dòng tiêu đề vẫn đúng nguyên.
    sheet.freeze_panes = "A2"


def build_template(spec: CatalogSpec, descriptor: TemplateDescriptor) -> bytes:
    """Tệp mẫu hoàn chỉnh dưới dạng byte, sẵn sàng cho một phản hồi HTTP."""
    workbook = Workbook(write_only=False)
    # `Workbook()` tạo sẵn một sheet trống tên "Sheet"; bỏ nó đi trước khi thêm
    # hai sheet thật, nếu không tệp mẫu có ba sheet và sheet thừa là thứ người
    # dùng sẽ hỏi để làm gì. `active` khai kiểu là có thể `None` nên phải hỏi —
    # trên thực tế sổ vừa tạo luôn có đúng một sheet.
    default_sheet = workbook.active
    if default_sheet is not None:
        workbook.remove(default_sheet)
    write_instructions_sheet(workbook, spec, descriptor)
    _write_header(workbook, descriptor)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def template_file_name(spec: CatalogSpec) -> str:
    """Tên tệp gợi ý cho lượt tải về.

    ASCII thuần từ `slug`, không phải `spec.title` có dấu: tên tệp đi vào header
    `Content-Disposition`, và tiếng Việt có dấu ở đó là chỗ mỗi trình duyệt giải
    mã một kiểu. Người dùng biết mình vừa bấm tải mẫu của danh mục nào.
    """
    return f"mau-nhap-lieu-{spec.slug.replace('_', '-')}.xlsx"
